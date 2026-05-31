"""Room-by-room rehab cost estimation with 6-level finish system (Rehab Levels).

Generates full rehab budgets, wholetail comparisons, and project timelines.
Regional pricing calibrated for Knoxville / East Tennessee market.

Usage:
  python src/main.py rehab --address "123 Main St, Knoxville, TN 37918"
  python src/main.py rehab --address "123 Main St" --tier 2 --scope full --region knoxville
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── Regional multipliers (Birmingham = 1.00 baseline, pivoted 2026-05-29) ────
# Per-sqft rates in PER_SQFT_BY_SIZE are calibrated to Birmingham metro directly.
# Other regions scale relative to that anchor (the user's investor sheet provides
# Birmingham-specific numbers; this pivot lets us read them verbatim).
REGIONAL_MULTIPLIERS = {
    # Alabama
    "birmingham": 1.00,      # Jefferson + Shelby (Hoover, Vestavia, Alabaster) — BASELINE
    "huntsville": 1.04,      # Madison — tech/aerospace labor premium
    "albertville": 0.95,     # Marshall County — rural
    # Tennessee
    "knoxville": 0.96,       # ~4% below Birmingham
    "blount": 0.93,
    "nashville": 1.03,
    "chattanooga": 0.98,
    # Default fallback
    "national": 1.09,        # ~9% above Birmingham (national avg)
}
DEFAULT_REGION = "auto"  # resolver picks the right region from city+state

# City → region routing. Covers all cities active in the SiftStack pipeline today
# (Jefferson + Shelby + Madison + Marshall AL, Knox + Blount TN). Unknown cities
# fall back to the state-level default in resolve_region().
_CITY_TO_REGION = {
    # Jefferson County AL (Birmingham metro)
    "birmingham": "birmingham", "bessemer": "birmingham", "hoover": "birmingham",
    "vestavia hills": "birmingham", "mountain brook": "birmingham", "homewood": "birmingham",
    "trussville": "birmingham", "hueytown": "birmingham", "pinson": "birmingham",
    "pleasant grove": "birmingham", "center point": "birmingham", "fultondale": "birmingham",
    "gardendale": "birmingham", "adamsville": "birmingham", "tarrant": "birmingham",
    "forestdale": "birmingham", "brighton": "birmingham", "irondale": "birmingham",
    "leeds": "birmingham", "midfield": "birmingham", "fairfield": "birmingham",
    # Shelby County AL (Birmingham southern bedroom communities — same labor pool)
    "alabaster": "birmingham", "helena": "birmingham", "pelham": "birmingham",
    "calera": "birmingham", "columbiana": "birmingham", "chelsea": "birmingham",
    "indian springs": "birmingham", "vincent": "birmingham", "wilsonville": "birmingham",
    "montevallo": "birmingham",
    # Madison County AL (Huntsville metro)
    "huntsville": "huntsville", "madison": "huntsville", "owens cross roads": "huntsville",
    "gurley": "huntsville", "new hope": "huntsville", "new market": "huntsville",
    "triana": "huntsville", "harvest": "huntsville", "meridianville": "huntsville",
    # Marshall County AL (Albertville/Boaz/Guntersville/Arab)
    "albertville": "albertville", "boaz": "albertville", "guntersville": "albertville",
    "arab": "albertville", "union grove": "albertville", "douglas": "albertville",
    "grant": "albertville", "horton": "albertville",
    # Knox County TN
    "knoxville": "knoxville", "powell": "knoxville", "farragut": "knoxville",
    "karns": "knoxville", "halls": "knoxville", "corryton": "knoxville",
    # Blount County TN
    "maryville": "blount", "alcoa": "blount", "friendsville": "blount",
    "louisville": "blount", "rockford": "blount", "townsend": "blount",
}

# State-level fallback when city isn't in the routing table.
_STATE_TO_REGION = {
    "AL": "birmingham",   # Most AL pipeline activity is Birmingham metro
    "TN": "knoxville",    # Legacy SiftStack TN scope
}


def resolve_region(city: str = "", state: str = "", region_override: str = "") -> str:
    """Pick the right REGIONAL_MULTIPLIERS key from property location.

    - If `region_override` is set AND not "auto", honor it (explicit user choice).
    - Else look up the city (case-insensitive) in `_CITY_TO_REGION`.
    - Else fall back to state-level default (`_STATE_TO_REGION`).
    - Else "national" (1.00 multiplier — neutral).
    """
    if region_override and region_override.lower() != "auto":
        return region_override.lower()
    city_key = (city or "").strip().lower()
    if city_key in _CITY_TO_REGION:
        return _CITY_TO_REGION[city_key]
    state_key = (state or "").strip().upper()
    if state_key in _STATE_TO_REGION:
        return _STATE_TO_REGION[state_key]
    return "national"

# ── 6 Rehab Levels (Birmingham investor sheet, 2026-05-29) ────────────────────
# Cosmetic-scope LEVELS ONLY. Big 6 systems (roof, HVAC, HWH, plumbing, electrical,
# foundation) are ALWAYS added separately via the big_tickets parameter — they
# are NEVER included in the per-sqft base. This avoids double-counting and matches
# the walkthrough discipline: "price the Big 6 first, then add cosmetics."
# Renamed from "Tier" to "Rehab Level" 2026-05-29 to avoid confusion with ZIP tiers.
REHAB_LEVEL_NAMES = {
    1: "Low Rehab (Rental-Almost)",
    2: "Mid Rehab (Salvageable Materials)",
    3: "Full Rehab (Interior Cosmetics)",
    4: "Full + Exterior Cosmetics",
    5: "Full + Some Big 6",
    6: "Gut Job (Interior Only)",
}
# Backward-compat alias (avoid touching every old import). New code uses REHAB_LEVEL_NAMES.
TIER_NAMES = REHAB_LEVEL_NAMES

# Per-sqft rates by size bucket — smaller houses cost MORE per sqft because
# fixed kitchen/bath overhead is spread over fewer sqft. Anchor data from the
# user's investor sheet (900 sqft and 1,035 sqft columns); larger sizes
# extrapolated to taper as fixed-cost overhead becomes a smaller share.
# All values are BIRMINGHAM-TARGETED ($/sqft); regional multiplier scales them.
PER_SQFT_BY_SIZE = {
    "small":    {1: 25, 2: 35, 3: 45, 4: 50, 5: 55, 6: 67},  # <950 sqft
    "standard": {1: 15, 2: 25, 3: 35, 4: 40, 5: 45, 6: 62},  # 950-1,200 (sheet anchor)
    "mid":      {1: 13, 2: 22, 3: 32, 4: 37, 5: 42, 6: 58},  # 1,200-1,700
    "large":    {1: 15, 2: 25, 3: 30, 4: 40, 5: 45, 6: 62},  # 1,700+ (corrected 2026-05-30)
}


def _size_bucket(sqft: int) -> str:
    if sqft < 950:
        return "small"
    if sqft < 1200:
        return "standard"
    if sqft < 1700:
        return "mid"
    return "large"


def per_sqft_for(sqft: int, tier: int) -> float:
    """Birmingham-target per-sqft rate for given size + tier."""
    return PER_SQFT_BY_SIZE[_size_bucket(sqft)][tier]


# ── Big 6 (always-add when seen at walkthrough — investor's rule) ──────────────
# Per Google investor sheet — Foundation, Roof, HVAC, Water Heater, Plumbing,
# Electrical. Cost values are Birmingham midpoints from sheet's range data;
# scaled by regional multiplier (these are mostly labor-driven).
# Roof is special: uses sqft × $5.00/roofing-sqft (where roofing sqft ≈ living × 1.20).
ROOF_PER_ROOFING_SQFT = 5.00  # Birmingham mid (sheet range $3.50-$7.00)

BIG_6_COSTS = {
    "roof":         None,    # computed from sqft (see big_6_cost helper)
    "hvac":         8500,    # sheet range $5K-$12K
    "hwh":          1250,    # sheet range $1K-$1.5K (water heater — NEW per sheet)
    "plumbing":     7500,    # sheet range $5K-$10K (full repipe + fixtures)
    "electrical":   7500,    # sheet range $5K-$10K (panel + wiring scope)
    "foundation":   4000,    # sheet rate $350-$450/pier × ~10 piers
}
BIG_6_LABELS = {
    "roof": "Roof (Big 6)",
    "hvac": "HVAC (Big 6)",
    "hwh": "Water Heater (Big 6)",
    "plumbing": "Plumbing (Big 6)",
    "electrical": "Electrical (Big 6)",
    "foundation": "Foundation (Big 6)",
}

# Optional adders (NOT in Big 6, but commonly needed)
OPTIONAL_ADDERS = {
    "septic":  10000,    # sheet range $5K-$15K — for rural/non-sewer properties
    "sewer":   10000,    # $5K-$15K — lateral replacement (separate from Big 6 plumbing)
    "windows": 10000,    # $6K-$14K — full house vinyl replacement
    "rewire":  15000,    # $10K-$20K — knob-and-tube; on top of Big 6 electrical
}
OPTIONAL_ADDER_LABELS = {
    "septic": "Septic system",
    "sewer": "Sewer lateral",
    "windows": "Windows (full)",
    "rewire": "Full rewire (knob-and-tube)",
}

# Single lookup for any adder name (used by walkthrough flag)
ALL_ADDERS = {**{k: BIG_6_LABELS[k] for k in BIG_6_COSTS},
              **OPTIONAL_ADDER_LABELS}


def big_6_cost(item: str, sqft: int = 0) -> int:
    """Return Birmingham-target cost for any Big 6 or optional adder item."""
    if item == "roof":
        roof_sqft = int(sqft * 1.20) if sqft else 0
        return round(ROOF_PER_ROOFING_SQFT * roof_sqft)
    if item in BIG_6_COSTS:
        return BIG_6_COSTS[item]
    if item in OPTIONAL_ADDERS:
        return OPTIONAL_ADDERS[item]
    return 0


def _contingency_pct(year_built: int, tier: int) -> float:
    """Pick contingency % from property age (preferred) or tier (fallback).

    Per Fast Walkthrough Formula Step 3:
      - Newer / light cosmetic   → 10%
      - 1960s-1990s moderate     → 15%
      - Pre-1960 / distressed    → 25% (mid of 20-30%)
    When year_built is unknown, fall back to tier-based defaults (1-6 scale).
    """
    if year_built:
        if year_built < 1960:
            return 0.25
        if year_built < 1991:
            return 0.15
        return 0.10
    # 6-tier fallback
    return {1: 0.10, 2: 0.10, 3: 0.15, 4: 0.15, 5: 0.20, 6: 0.25}[tier]

# ── Room cost tables ──────────────────────────────────────────────────
# Each room category has cost ranges per tier: {tier: (materials, labor)}
# All costs in USD, national average, before regional multiplier

KITCHEN_COSTS = {
    1: {"demo": 500, "cabinets": 2500, "countertops": 800, "appliances": 1500,
        "fixtures": 300, "backsplash": 0, "flooring": 0, "paint": 200, "labor": 2000},
    2: {"demo": 750, "cabinets": 5000, "countertops": 2000, "appliances": 3000,
        "fixtures": 500, "backsplash": 500, "flooring": 0, "paint": 300, "labor": 4000},
    3: {"demo": 1000, "cabinets": 8000, "countertops": 4000, "appliances": 5000,
        "fixtures": 800, "backsplash": 1200, "flooring": 0, "paint": 400, "labor": 6000},
    4: {"demo": 1500, "cabinets": 15000, "countertops": 8000, "appliances": 8000,
        "fixtures": 1500, "backsplash": 2500, "flooring": 0, "paint": 500, "labor": 10000},
}

MASTER_BATH_COSTS = {
    1: {"demo": 400, "vanity": 300, "toilet": 150, "tub_shower": 400, "tile": 500,
        "fixtures": 200, "paint": 150, "labor": 1500},
    2: {"demo": 600, "vanity": 800, "toilet": 250, "tub_shower": 1200, "tile": 1500,
        "fixtures": 400, "paint": 200, "labor": 3000},
    3: {"demo": 800, "vanity": 1500, "toilet": 400, "tub_shower": 2500, "tile": 3000,
        "fixtures": 800, "paint": 250, "labor": 5000},
    4: {"demo": 1000, "vanity": 3000, "toilet": 600, "tub_shower": 5000, "tile": 5000,
        "fixtures": 1500, "paint": 300, "labor": 8000},
}

SECONDARY_BATH_COSTS = {
    1: {"demo": 300, "vanity": 200, "toilet": 150, "tub_shower": 350, "tile": 400,
        "fixtures": 150, "paint": 100, "labor": 1200},
    2: {"demo": 500, "vanity": 500, "toilet": 250, "tub_shower": 800, "tile": 1000,
        "fixtures": 300, "paint": 150, "labor": 2500},
    3: {"demo": 700, "vanity": 1000, "toilet": 350, "tub_shower": 1800, "tile": 2000,
        "fixtures": 500, "paint": 200, "labor": 3500},
    4: {"demo": 900, "vanity": 2000, "toilet": 500, "tub_shower": 3500, "tile": 3500,
        "fixtures": 1000, "paint": 250, "labor": 6000},
}

# Per-sqft costs for whole-house items
FLOORING_PER_SQFT = {1: 2.50, 2: 4.50, 3: 7.00, 4: 12.00}   # materials
FLOORING_LABOR_PER_SQFT = {1: 1.50, 2: 2.00, 3: 2.50, 4: 3.50}
PAINT_PER_SQFT = {1: 0.50, 2: 0.75, 3: 1.00, 4: 1.50}         # materials (walls + trim)
PAINT_LABOR_PER_SQFT = {1: 0.80, 2: 1.00, 3: 1.20, 4: 1.50}

# Fixed-cost items
WINDOWS_PER_UNIT = {1: 250, 2: 400, 3: 650, 4: 1000}  # per window (materials + labor)
# Birmingham-metro asphalt-shingle roof replacement, calibrated 2026-05-28.
# Per-sqft is roofing-sqft (NOT floor-sqft — roof area runs 15-35% larger than floor
# due to pitch, eaves, valleys; we use 1.20x as the working multiplier).
# Tier mapping (per market reference data):
#   1 = basic 3-tab shingles ($4.5K-$6.5K typical)
#   2 = architectural shingles, most common ($6K-$9K typical)
#   3 = upgraded architectural / starter impact-resistant ($7.5K-$9.5K)
#   4 = full impact-resistant / premium ($10.5K+)
# Includes: tear-off, synthetic underlayment, drip edge, ridge vent, cleanup, warranty.
# Add separately if decking replacement is needed at inspection.
ROOF_PER_SQFT = {1: 4.00, 2: 6.00, 3: 7.25, 4: 9.25}  # per ROOFING sqft (≈ floor sqft × 1.20)
HVAC_COSTS = {1: 4000, 2: 6000, 3: 8500, 4: 12000}     # full system replacement
ELECTRICAL_COSTS = {1: 2000, 2: 4000, 3: 7000, 4: 12000}  # panel + rewire
PLUMBING_COSTS = {1: 1500, 2: 3000, 3: 5000, 4: 8000}    # repipe + fixtures
FOUNDATION_COSTS = {1: 2000, 2: 5000, 3: 10000, 4: 20000}  # structural repair

EXTERIOR_COSTS = {
    1: {"siding": 0, "paint": 1500, "landscaping": 500, "driveway": 0, "labor": 1000},
    2: {"siding": 3000, "paint": 2500, "landscaping": 1500, "driveway": 0, "labor": 2000},
    3: {"siding": 6000, "paint": 4000, "landscaping": 3000, "driveway": 2000, "labor": 4000},
    4: {"siding": 12000, "paint": 6000, "landscaping": 6000, "driveway": 5000, "labor": 7000},
}

# Timeline estimates (weeks) per category per tier
TIMELINE_WEEKS = {
    "kitchen": {1: 1, 2: 2, 3: 3, 4: 5},
    "bathrooms": {1: 1, 2: 2, 3: 3, 4: 4},
    "flooring": {1: 1, 2: 1, 3: 2, 4: 2},
    "paint": {1: 1, 2: 1, 3: 1, 4: 2},
    "windows": {1: 0.5, 2: 1, 3: 1, 4: 2},
    "roof": {1: 1, 2: 1, 3: 1.5, 4: 2},
    "hvac": {1: 0.5, 2: 1, 3: 1, 4: 1.5},
    "electrical": {1: 0.5, 2: 1, 3: 1.5, 4: 2},
    "plumbing": {1: 0.5, 2: 1, 3: 1.5, 4: 2},
    "foundation": {1: 1, 2: 2, 3: 3, 4: 4},
    "exterior": {1: 1, 2: 2, 3: 3, 4: 4},
}

# ── Data structures ───────────────────────────────────────────────────


@dataclass
class RoomEstimate:
    """Cost estimate for a single room/category."""
    category: str = ""
    tier: int = 2
    materials: float = 0.0
    labor: float = 0.0
    total: float = 0.0
    line_items: dict = field(default_factory=dict)
    weeks: float = 0.0
    notes: str = ""


@dataclass
class RehabEstimate:
    """Full rehab estimate for a property."""
    address: str = ""
    tier: int = 2
    scope: str = "full"  # "full" or "wholetail"
    region: str = DEFAULT_REGION
    regional_multiplier: float = 0.88
    sqft: int = 0
    bedrooms: int = 0
    bathrooms: float = 0.0
    year_built: int = 0
    rooms: list = field(default_factory=list)
    total_materials: float = 0.0
    total_labor: float = 0.0
    total_cost: float = 0.0
    total_weeks: float = 0.0
    permits_cost: float = 0.0
    contingency_pct: float = 0.10  # 10% contingency
    contingency_cost: float = 0.0
    grand_total: float = 0.0


# ── Estimation engine ─────────────────────────────────────────────────

def _calc_room(category: str, cost_table: dict, tier: int, multiplier: float,
               quantity: int = 1) -> RoomEstimate:
    """Calculate cost for a room category from its cost table."""
    tier_costs = cost_table.get(tier, cost_table.get(2, {}))
    labor = tier_costs.get("labor", 0) * multiplier * quantity
    materials = sum(v for k, v in tier_costs.items() if k != "labor") * multiplier * quantity

    return RoomEstimate(
        category=category,
        tier=tier,
        materials=round(materials),
        labor=round(labor),
        total=round(materials + labor),
        line_items={k: round(v * multiplier * quantity) for k, v in tier_costs.items()},
        weeks=TIMELINE_WEEKS.get(category.lower().split()[0], {}).get(tier, 1) * quantity,
    )


def _calc_per_sqft(category: str, sqft: int, mat_table: dict, labor_table: dict,
                   tier: int, multiplier: float, timeline_key: str = "") -> RoomEstimate:
    """Calculate cost for a per-sqft category."""
    mat_rate = mat_table.get(tier, mat_table.get(2, 0))
    labor_rate = labor_table.get(tier, labor_table.get(2, 0))
    materials = round(sqft * mat_rate * multiplier)
    labor = round(sqft * labor_rate * multiplier)

    return RoomEstimate(
        category=category,
        tier=tier,
        materials=materials,
        labor=labor,
        total=materials + labor,
        line_items={"materials_per_sqft": round(mat_rate * multiplier, 2),
                    "labor_per_sqft": round(labor_rate * multiplier, 2),
                    "sqft": sqft},
        weeks=TIMELINE_WEEKS.get(timeline_key or category.lower(), {}).get(tier, 1),
    )


def _calc_fixed(category: str, cost_table: dict, tier: int, multiplier: float,
                timeline_key: str = "") -> RoomEstimate:
    """Calculate cost for a fixed-cost category (HVAC, electrical, etc.)."""
    total = round(cost_table.get(tier, cost_table.get(2, 0)) * multiplier)
    # Rough 60/40 labor/materials split for mechanical work
    labor = round(total * 0.6)
    materials = total - labor

    return RoomEstimate(
        category=category,
        tier=tier,
        materials=materials,
        labor=labor,
        total=total,
        line_items={"total_installed": total},
        weeks=TIMELINE_WEEKS.get(timeline_key or category.lower(), {}).get(tier, 1),
    )


def estimate_rehab(address: str = "", sqft: int = 0, bedrooms: int = 3,
                   bathrooms: float = 2.0, year_built: int = 0,
                   tier: int = 2, scope: str = "full",
                   region: str = DEFAULT_REGION,
                   big_tickets: list[str] | None = None,
                   city: str = "", state: str = "") -> RehabEstimate:
    """Generate a Fast Walkthrough rehab estimate.

    Formula:
      Step 1: base = sqft × PER_SQFT_BY_TIER[tier] × regional_multiplier
      Step 2: + sum of opt-in big-ticket items (roof, hvac, etc.)
      Step 3: × (1 + contingency_pct), where contingency is age-driven
              (pre-1960 = 25%, 1960-1990 = 15%, post-1990 = 10%)

    Args:
        address, city, state: Used by resolve_region when region == "auto"
        sqft: Living sqft (defaults to 1500 if unknown)
        tier: 1=Lipstick, 2=Moderate, 3=Heavy, 4=Full Gut
        scope: "full" or "wholetail" (forces Tier 1, lipstick-only, no big tickets)
        big_tickets: opt-in items, keys from BIG_TICKET_COSTS
                     (e.g. ["roof","hvac"]). Defaults: none for Tier 1/2,
                     roof+hvac+panel for Tier 3, all for Tier 4.
    """
    # Resolve region from location if caller passed "auto" / blank
    region = resolve_region(city, state, region)
    multiplier = REGIONAL_MULTIPLIERS.get(region.lower(), REGIONAL_MULTIPLIERS["national"])
    tier = max(1, min(6, tier))  # 6-tier system

    if not sqft:
        sqft = 1500  # fallback when subject data missing

    # Wholetail = Rehab Level 1 (Low Rehab), no Big 6, simple contingency
    if scope == "wholetail":
        tier = 1
        big_tickets = []

    # Per walkthrough rule — Big 6 are ALWAYS opt-in adders, never auto.
    # The per-sqft tier base covers cosmetic scope only; Big 6 items get added
    # explicitly when seen at walkthrough. Investor's rule: "price the Big 6 first."
    if big_tickets is None:
        big_tickets = []

    rooms = []

    # ── Step 1: Base scope (size-bucket × tier × regional multiplier) ──
    base_per_sqft = per_sqft_for(sqft, tier) * multiplier
    base_total = round(sqft * base_per_sqft)
    rooms.append(RoomEstimate(
        category=f"Base Scope — Rehab Level {tier}: {REHAB_LEVEL_NAMES[tier]}",
        tier=tier,
        materials=round(base_total * 0.45),
        labor=round(base_total * 0.55),
        total=base_total,
        line_items={"per_sqft": round(base_per_sqft, 2), "sqft": sqft,
                    "size_bucket": _size_bucket(sqft)},
        weeks={1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 8}[tier],
    ))

    # ── Step 2: Big 6 + Optional adders (opt-in only) ──
    for key in big_tickets:
        if key not in BIG_6_COSTS and key not in OPTIONAL_ADDERS:
            logger.warning("Unknown adder item: %s (valid: %s)",
                           key, list(BIG_6_COSTS.keys()) + list(OPTIONAL_ADDERS.keys()))
            continue
        cost = round(big_6_cost(key, sqft) * multiplier)
        label = BIG_6_LABELS.get(key) or OPTIONAL_ADDER_LABELS.get(key, key)
        rooms.append(RoomEstimate(
            category=label,
            tier=tier,
            materials=round(cost * 0.5),
            labor=round(cost * 0.5),
            total=cost,
            line_items={"adder": key, "applied_multiplier": round(multiplier, 2)},
            weeks={"roof": 1, "hvac": 1, "hwh": 0.25, "plumbing": 1.5,
                   "electrical": 1, "foundation": 2, "septic": 1.5,
                   "sewer": 1, "windows": 1, "rewire": 2}.get(key, 1),
        ))

    # ── Step 3: Subtotals + age-based contingency ──
    total_materials = sum(r.materials for r in rooms)
    total_labor = sum(r.labor for r in rooms)
    subtotal = total_materials + total_labor
    total_weeks = sum(r.weeks for r in rooms) * 0.6  # parallel work assumption

    cont_pct = _contingency_pct(year_built, tier)
    contingency = round(subtotal * cont_pct)
    permits = round(subtotal * 0.03)  # ~3% permits (kept from prior model)
    grand_total = subtotal + permits + contingency

    estimate = RehabEstimate(
        address=address,
        tier=tier,
        scope=scope,
        region=region,
        regional_multiplier=multiplier,
        sqft=sqft,
        bedrooms=bedrooms,
        bathrooms=bathrooms,
        year_built=year_built,
        rooms=rooms,
        total_materials=round(total_materials),
        total_labor=round(total_labor),
        total_cost=round(subtotal),
        total_weeks=round(total_weeks, 1),
        permits_cost=permits,
        contingency_pct=cont_pct,
        contingency_cost=contingency,
        grand_total=round(grand_total),
    )

    logger.info("Rehab estimate for %s: %s scope, Rehab Level %d (%s), Total $%s, ~%.0f weeks",
                address or "property", scope, tier, REHAB_LEVEL_NAMES[tier],
                f"{grand_total:,.0f}", total_weeks)

    return estimate


def estimate_wholetail(address: str = "", sqft: int = 0, bedrooms: int = 3,
                       bathrooms: float = 2.0, year_built: int = 0,
                       tier: int = 2, region: str = DEFAULT_REGION,
                       city: str = "", state: str = "") -> RehabEstimate:
    """Generate a wholetail (cosmetic-only / lipstick) estimate."""
    return estimate_rehab(address, sqft, bedrooms, bathrooms, year_built,
                          tier=1, scope="wholetail", region=region,
                          city=city, state=state)


# ── Excel report generation ──────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_MONEY_FMT = '#,##0'
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _fmt(val: float) -> str:
    return f"${val:,.0f}"


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_col_widths(ws, min_w=12, max_w=35):
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)


def generate_rehab_report(full_est: RehabEstimate, wholetail_est: RehabEstimate | None = None,
                          output_path: str = "") -> str:
    """Generate a 10-tab Excel workbook rehab estimate report."""
    wb = Workbook()

    # ── Tab 1: Budget Matrix (Quick Reference) ────────────────────
    # Single-glance view: subject's tier × size matrix with subject row highlighted.
    ws0 = wb.active
    ws0.title = "Budget Matrix"
    ws0.cell(row=1, column=1, value="Rehab Budget Matrix — Birmingham Metro").font = _TITLE_FONT
    ws0.cell(row=2, column=1, value=full_est.address or "Subject Property").font = _SUBTITLE_FONT
    ws0.cell(row=3, column=1,
        value=f"Subject: {full_est.sqft:,} sqft / {full_est.bedrooms}bd / {full_est.bathrooms}ba"
              + (f" / built {full_est.year_built}" if full_est.year_built else "")
              + f" — region: {full_est.region.title()} (×{full_est.regional_multiplier:.2f})"
    ).font = _LABEL_FONT

    # Subject's specific computed totals — called out prominently
    row = 5
    ws0.cell(row=row, column=1, value="── SUBJECT PROPERTY ESTIMATES ──").font = Font(
        name="Calibri", bold=True, size=13, color="2F5496")
    row += 1
    ws0.cell(row=row, column=1, value=f"Full Rehab (Rehab Level {full_est.tier} — {REHAB_LEVEL_NAMES[full_est.tier]})").font = _LABEL_FONT
    ws0.cell(row=row, column=2, value=full_est.grand_total).number_format = _MONEY_FMT
    ws0.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=14, color="006100")
    ws0.cell(row=row, column=2).fill = _GREEN_FILL
    row += 1
    if wholetail_est:
        ws0.cell(row=row, column=1, value=f"Wholetail (Rehab Level 1 — {REHAB_LEVEL_NAMES[1]})").font = _LABEL_FONT
        ws0.cell(row=row, column=2, value=wholetail_est.grand_total).number_format = _MONEY_FMT
        ws0.cell(row=row, column=2).font = Font(name="Calibri", bold=True, size=14, color="006100")
        ws0.cell(row=row, column=2).fill = _GREEN_FILL
        row += 1

    # Reference matrix
    row += 2
    ws0.cell(row=row, column=1, value="── BIRMINGHAM METRO REFERENCE MATRIX ──").font = Font(
        name="Calibri", bold=True, size=13, color="2F5496")
    row += 1
    ws0.cell(row=row, column=1, value="Typical rehab budget by house size + condition tier").font = _LABEL_FONT
    row += 2

    # Header row — 6 tiers
    headers = ["House Size", "1. Low Rehab\n(Rental-Almost)", "2. Mid Rehab\n(Salvageable)",
               "3. Full Rehab\n(Int. Cosmetics)", "4. Full + Ext.\nCosmetics",
               "5. Full + Some\nBig 6", "6. Gut Job\n(Interior)"]
    for col, h in enumerate(headers, 1):
        cell = ws0.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    matrix_header_row = row
    row += 1

    # Reference rows computed live from PER_SQFT_BY_SIZE × the subject's region
    # multiplier — shows what each tier would cost AT the subject's market.
    bucket_anchors = [("<950 sf",   "small",    900),
                      ("1,000 sf",  "standard", 1000),
                      ("1,200 sf",  "mid",      1200),  # 1200 → mid bucket
                      ("1,500 sf",  "mid",      1500),
                      ("2,000 sf",  "large",    2000)]
    yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    subj_bucket = _size_bucket(full_est.sqft or 0)
    for row_label, bucket_key, sample_sf in bucket_anchors:
        is_subject_row = (bucket_key == subj_bucket)
        ws0.cell(row=row, column=1, value=row_label).font = (
            Font(name="Calibri", bold=True, size=11) if is_subject_row else _LABEL_FONT)
        ws0.cell(row=row, column=1).border = _THIN_BORDER
        for tier in range(1, 7):
            base = PER_SQFT_BY_SIZE[bucket_key][tier] * sample_sf * full_est.regional_multiplier
            cell = ws0.cell(row=row, column=tier + 1, value=f"${base/1000:.0f}K")
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER
            if is_subject_row:
                cell.fill = yellow
                cell.font = Font(name="Calibri", bold=True, size=11)
        if is_subject_row:
            ws0.cell(row=row, column=1).fill = yellow
        row += 1

    row += 1
    ws0.cell(row=row, column=1,
        value=f"↑ Subject is {full_est.sqft:,} sqft → '{subj_bucket}' size bucket (highlighted). "
              f"Rates shown are BASE only — Big 6 adders not included."
    ).font = _LABEL_FONT
    row += 2

    # ── Big 6 reference block ─────────────────────────────────────
    ws0.cell(row=row, column=1, value="── BIG 6 ADDERS (always-add when seen at walkthrough) ──").font = Font(
        name="Calibri", bold=True, size=13, color="2F5496")
    row += 1
    ws0.cell(row=row, column=1, value="Investor rule: walk the house, price the Big 6 first. Cosmetics don't kill deals — hidden systems do.").font = _LABEL_FONT
    row += 2
    ws0.cell(row=row, column=1, value="Item").font = _HEADER_FONT
    ws0.cell(row=row, column=1).fill = _HEADER_FILL
    ws0.cell(row=row, column=2, value="Birmingham Cost").font = _HEADER_FONT
    ws0.cell(row=row, column=2).fill = _HEADER_FILL
    ws0.cell(row=row, column=3, value="Sheet Range").font = _HEADER_FONT
    ws0.cell(row=row, column=3).fill = _HEADER_FILL
    ws0.cell(row=row, column=4, value="CLI Key").font = _HEADER_FONT
    ws0.cell(row=row, column=4).fill = _HEADER_FILL
    row += 1
    subj_sqft = full_est.sqft or 1032
    big_6_display = [
        ("Roof",              f"${big_6_cost('roof', subj_sqft) * full_est.regional_multiplier:,.0f}",
                              "$3.50-$7.00/sqft of roof area", "roof"),
        ("HVAC",              f"${BIG_6_COSTS['hvac'] * full_est.regional_multiplier:,.0f}",
                              "$5K-$12K", "hvac"),
        ("Water Heater (HWH)", f"${BIG_6_COSTS['hwh'] * full_est.regional_multiplier:,.0f}",
                              "$1K-$1.5K", "hwh"),
        ("Plumbing",          f"${BIG_6_COSTS['plumbing'] * full_est.regional_multiplier:,.0f}",
                              "$5K-$10K", "plumbing"),
        ("Electrical",        f"${BIG_6_COSTS['electrical'] * full_est.regional_multiplier:,.0f}",
                              "$5K-$10K", "electrical"),
        ("Foundation",        f"${BIG_6_COSTS['foundation'] * full_est.regional_multiplier:,.0f}",
                              "$350-$450/pier × ~10 piers", "foundation"),
    ]
    for label, cost, range_, key in big_6_display:
        ws0.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws0.cell(row=row, column=2, value=cost).font = _VALUE_FONT
        ws0.cell(row=row, column=3, value=range_).font = _LABEL_FONT
        ws0.cell(row=row, column=4, value=key).font = Font(name="Consolas", size=10)
        row += 1
    row += 1
    ws0.cell(row=row, column=1, value="── OPTIONAL ADDERS ──").font = Font(name="Calibri", bold=True, size=11, color="555555")
    row += 1
    optional_display = [
        ("Septic system", f"${OPTIONAL_ADDERS['septic'] * full_est.regional_multiplier:,.0f}", "$5K-$15K (rural/non-sewer)", "septic"),
        ("Sewer lateral", f"${OPTIONAL_ADDERS['sewer'] * full_est.regional_multiplier:,.0f}", "$5K-$15K", "sewer"),
        ("Windows (full)", f"${OPTIONAL_ADDERS['windows'] * full_est.regional_multiplier:,.0f}", "$6K-$14K", "windows"),
        ("Full rewire (K&T)", f"${OPTIONAL_ADDERS['rewire'] * full_est.regional_multiplier:,.0f}", "$10K-$20K", "rewire"),
    ]
    for label, cost, range_, key in optional_display:
        ws0.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws0.cell(row=row, column=2, value=cost).font = _VALUE_FONT
        ws0.cell(row=row, column=3, value=range_).font = _LABEL_FONT
        ws0.cell(row=row, column=4, value=key).font = Font(name="Consolas", size=10)
        row += 1

    row += 2
    ws0.cell(row=row, column=1, value="Formula:").font = Font(name="Calibri", bold=True, size=11)
    row += 1
    for note in [
        "  Step 1: base = sqft × per-sqft (size-bucket × tier) × regional multiplier",
        "  Step 2: + Big 6 / optional adders (--big-tickets roof,hvac,...) — ALWAYS opt-in",
        "  Step 3: × age contingency (pre-1960 = 25%, 1960-1990 = 15%, post-1990 = 10%)",
    ]:
        ws0.cell(row=row, column=1, value=note).font = _LABEL_FONT
        row += 1

    ws0.column_dimensions["A"].width = 22
    for c in ["B", "C", "D", "E", "F", "G"]:
        ws0.column_dimensions[c].width = 18
    ws0.row_dimensions[matrix_header_row].height = 42

    # ── Tab 2: Executive Summary ──────────────────────────────────
    ws = wb.create_sheet("Executive Summary")
    ws.cell(row=1, column=1, value="Rehab Cost Estimate").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=full_est.address or "Subject Property").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _LABEL_FONT

    row = 5
    data = [
        ("Rehab Level", f"Level {full_est.tier} — {REHAB_LEVEL_NAMES[full_est.tier]}"),
        ("Region", f"{full_est.region.title()} (×{full_est.regional_multiplier:.2f})"),
        ("Property Size", f"{full_est.sqft:,} sqft"),
        ("Bed/Bath", f"{full_est.bedrooms}bd / {full_est.bathrooms}ba"),
        ("Year Built", str(full_est.year_built) if full_est.year_built else "N/A"),
        ("", ""),
        ("FULL REHAB", ""),
        ("Materials", _fmt(full_est.total_materials)),
        ("Labor", _fmt(full_est.total_labor)),
        ("Subtotal", _fmt(full_est.total_cost)),
        ("Permits (~3%)", _fmt(full_est.permits_cost)),
        ("Contingency (10%)", _fmt(full_est.contingency_cost)),
        ("GRAND TOTAL", _fmt(full_est.grand_total)),
        ("Est. Timeline", f"{full_est.total_weeks:.0f} weeks"),
    ]

    if wholetail_est:
        data += [
            ("", ""),
            ("WHOLETAIL", ""),
            ("Materials", _fmt(wholetail_est.total_materials)),
            ("Labor", _fmt(wholetail_est.total_labor)),
            ("GRAND TOTAL", _fmt(wholetail_est.grand_total)),
            ("Est. Timeline", f"{wholetail_est.total_weeks:.0f} weeks"),
            ("", ""),
            ("SAVINGS (Wholetail vs Full)", _fmt(full_est.grand_total - wholetail_est.grand_total)),
        ]

    for label, value in data:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _VALUE_FONT
        if label in ("GRAND TOTAL", "FULL REHAB", "WHOLETAIL"):
            cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35

    # ── Tab 2: Wholetail Scope ────────────────────────────────────
    ws2 = wb.create_sheet("Wholetail Scope")
    ws2.cell(row=1, column=1, value="Wholetail Scope (Cosmetic Only)").font = _TITLE_FONT
    ws2.cell(row=2, column=1, value="Paint, flooring, fixtures, landscaping, deep clean").font = _LABEL_FONT

    est = wholetail_est or estimate_wholetail(
        full_est.address, full_est.sqft, full_est.bedrooms,
        full_est.bathrooms, full_est.year_built, region=full_est.region)

    headers = ["Category", "Materials", "Labor", "Total", "Weeks"]
    _write_header_row(ws2, 4, headers)
    for i, room in enumerate(est.rooms, 5):
        ws2.cell(row=i, column=1, value=room.category)
        ws2.cell(row=i, column=2, value=room.materials).number_format = _MONEY_FMT
        ws2.cell(row=i, column=3, value=room.labor).number_format = _MONEY_FMT
        ws2.cell(row=i, column=4, value=room.total).number_format = _MONEY_FMT
        ws2.cell(row=i, column=5, value=room.weeks)
        for c in range(1, 6):
            ws2.cell(row=i, column=c).border = _THIN_BORDER
    total_row = 5 + len(est.rooms)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = _VALUE_FONT
    ws2.cell(row=total_row, column=4, value=est.grand_total).number_format = _MONEY_FMT
    ws2.cell(row=total_row, column=4).font = _VALUE_FONT
    _auto_col_widths(ws2)

    # ── Tab 3: Full Rehab Scope ───────────────────────────────────
    ws3 = wb.create_sheet("Full Rehab Scope")
    ws3.cell(row=1, column=1, value="Full Rehab Scope").font = _TITLE_FONT

    _write_header_row(ws3, 3, headers)
    for i, room in enumerate(full_est.rooms, 4):
        ws3.cell(row=i, column=1, value=room.category)
        ws3.cell(row=i, column=2, value=room.materials).number_format = _MONEY_FMT
        ws3.cell(row=i, column=3, value=room.labor).number_format = _MONEY_FMT
        ws3.cell(row=i, column=4, value=room.total).number_format = _MONEY_FMT
        ws3.cell(row=i, column=5, value=room.weeks)
        for c in range(1, 6):
            ws3.cell(row=i, column=c).border = _THIN_BORDER
    total_row = 4 + len(full_est.rooms)
    ws3.cell(row=total_row, column=1, value="Subtotal").font = _VALUE_FONT
    ws3.cell(row=total_row, column=4, value=full_est.total_cost).number_format = _MONEY_FMT
    ws3.cell(row=total_row + 1, column=1, value="Permits (3%)")
    ws3.cell(row=total_row + 1, column=4, value=full_est.permits_cost).number_format = _MONEY_FMT
    ws3.cell(row=total_row + 2, column=1, value="Contingency (10%)")
    ws3.cell(row=total_row + 2, column=4, value=full_est.contingency_cost).number_format = _MONEY_FMT
    ws3.cell(row=total_row + 3, column=1, value="GRAND TOTAL").font = _VALUE_FONT
    ws3.cell(row=total_row + 3, column=4, value=full_est.grand_total).number_format = _MONEY_FMT
    ws3.cell(row=total_row + 3, column=4).font = _VALUE_FONT
    _auto_col_widths(ws3)

    # ── Tab 4: Room-by-Room Detail ────────────────────────────────
    ws4 = wb.create_sheet("Room-by-Room Detail")
    ws4.cell(row=1, column=1, value="Room-by-Room Line Items").font = _TITLE_FONT

    row = 3
    for room in full_est.rooms:
        ws4.cell(row=row, column=1, value=room.category).font = _SUBTITLE_FONT
        row += 1
        for item, cost in room.line_items.items():
            ws4.cell(row=row, column=1, value=f"  {item.replace('_', ' ').title()}")
            ws4.cell(row=row, column=2, value=cost).number_format = _MONEY_FMT
            ws4.cell(row=row, column=2).border = _THIN_BORDER
            row += 1
        ws4.cell(row=row, column=1, value=f"  Room Total").font = _VALUE_FONT
        ws4.cell(row=row, column=2, value=room.total).number_format = _MONEY_FMT
        ws4.cell(row=row, column=2).font = _VALUE_FONT
        row += 2

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 15

    # ── Tab 5: Materials List ─────────────────────────────────────
    ws5 = wb.create_sheet("Materials List")
    ws5.cell(row=1, column=1, value="Materials Summary").font = _TITLE_FONT
    _write_header_row(ws5, 3, ["Category", "Materials Cost"])
    for i, room in enumerate(full_est.rooms, 4):
        ws5.cell(row=i, column=1, value=room.category)
        ws5.cell(row=i, column=2, value=room.materials).number_format = _MONEY_FMT
    total_row = 4 + len(full_est.rooms)
    ws5.cell(row=total_row, column=1, value="TOTAL MATERIALS").font = _VALUE_FONT
    ws5.cell(row=total_row, column=2, value=full_est.total_materials).number_format = _MONEY_FMT
    _auto_col_widths(ws5)

    # ── Tab 6: Labor Breakdown ────────────────────────────────────
    ws6 = wb.create_sheet("Labor Breakdown")
    ws6.cell(row=1, column=1, value="Labor Summary").font = _TITLE_FONT
    _write_header_row(ws6, 3, ["Category", "Labor Cost", "Est. Weeks"])
    for i, room in enumerate(full_est.rooms, 4):
        ws6.cell(row=i, column=1, value=room.category)
        ws6.cell(row=i, column=2, value=room.labor).number_format = _MONEY_FMT
        ws6.cell(row=i, column=3, value=room.weeks)
    total_row = 4 + len(full_est.rooms)
    ws6.cell(row=total_row, column=1, value="TOTAL LABOR").font = _VALUE_FONT
    ws6.cell(row=total_row, column=2, value=full_est.total_labor).number_format = _MONEY_FMT
    _auto_col_widths(ws6)

    # ── Tab 7: Timeline ───────────────────────────────────────────
    ws7 = wb.create_sheet("Timeline")
    ws7.cell(row=1, column=1, value="Project Timeline").font = _TITLE_FONT
    ws7.cell(row=2, column=1,
             value=f"Estimated {full_est.total_weeks:.0f} weeks (with parallel work)").font = _LABEL_FONT

    _write_header_row(ws7, 4, ["Category", "Duration (weeks)", "Phase"])
    phase_map = {
        "Foundation/Structural": "1 — Structural",
        "Roof": "1 — Structural",
        "Plumbing": "2 — Mechanical",
        "Electrical": "2 — Mechanical",
        "HVAC": "2 — Mechanical",
        "Windows": "2 — Mechanical",
        "Kitchen": "3 — Interior",
        "Master Bathroom": "3 — Interior",
        "Secondary Bathroom(s)": "3 — Interior",
        "Flooring": "4 — Finish",
        "Paint (Interior)": "4 — Finish",
        "Exterior": "5 — Exterior",
    }
    for i, room in enumerate(full_est.rooms, 5):
        ws7.cell(row=i, column=1, value=room.category)
        ws7.cell(row=i, column=2, value=room.weeks)
        ws7.cell(row=i, column=3, value=phase_map.get(room.category, ""))
    _auto_col_widths(ws7)

    # ── Tab 8: Comparison ─────────────────────────────────────────
    ws8 = wb.create_sheet("Comparison")
    ws8.cell(row=1, column=1, value="Wholetail vs Full Rehab Comparison").font = _TITLE_FONT

    wt = wholetail_est or est
    comp_data = [
        ("", "Wholetail", "Full Rehab", "Difference"),
        ("Materials", wt.total_materials, full_est.total_materials,
         full_est.total_materials - wt.total_materials),
        ("Labor", wt.total_labor, full_est.total_labor,
         full_est.total_labor - wt.total_labor),
        ("Permits", wt.permits_cost, full_est.permits_cost,
         full_est.permits_cost - wt.permits_cost),
        ("Contingency", wt.contingency_cost, full_est.contingency_cost,
         full_est.contingency_cost - wt.contingency_cost),
        ("GRAND TOTAL", wt.grand_total, full_est.grand_total,
         full_est.grand_total - wt.grand_total),
        ("Timeline (weeks)", wt.total_weeks, full_est.total_weeks,
         full_est.total_weeks - wt.total_weeks),
    ]
    _write_header_row(ws8, 3, comp_data[0])
    for i, (label, wt_val, full_val, diff) in enumerate(comp_data[1:], 4):
        ws8.cell(row=i, column=1, value=label)
        ws8.cell(row=i, column=2, value=wt_val).number_format = _MONEY_FMT
        ws8.cell(row=i, column=3, value=full_val).number_format = _MONEY_FMT
        ws8.cell(row=i, column=4, value=diff).number_format = _MONEY_FMT
        for c in range(1, 5):
            ws8.cell(row=i, column=c).border = _THIN_BORDER
    _auto_col_widths(ws8)

    # ── Tab 9: Notes & Assumptions ────────────────────────────────
    ws9 = wb.create_sheet("Notes & Assumptions")
    ws9.cell(row=1, column=1, value="Notes & Assumptions").font = _TITLE_FONT
    notes = [
        f"Rehab Level {full_est.tier}: {REHAB_LEVEL_NAMES[full_est.tier]}",
        f"Regional multiplier: {full_est.region.title()} = {full_est.regional_multiplier:.2f}x national avg",
        "",
        "Tier Definitions:",
        "  Tier 1 (Minimum Viable): Cheapest materials, basic function. Rental-ready.",
        "  Tier 2 (Builder Grade): Standard new construction level. Most common for flips.",
        "  Tier 3 (Mid-Range): Granite, hardwood, updated fixtures. Higher ARV neighborhoods.",
        "  Tier 4 (Premium/Custom): High-end finishes, custom work. Luxury market only.",
        "",
        "Assumptions:",
        "  - Permits estimated at 3% of total cost",
        "  - 10% contingency included for unforeseen issues",
        "  - Timeline assumes parallel work (60% of sequential estimate)",
        "  - Foundation/structural work included only for homes built before 1970",
        "  - Window count estimated at 1 per 100 sqft of living area",
        "  - Roof area estimated at 1.20x living area sqft (Birmingham-metro median; range 1.15-1.35)",
        "",
        "Wholetail Scope includes: Kitchen, Bathrooms, Flooring, Paint, Exterior (cosmetic)",
        "Wholetail Excludes: Windows, Roof, HVAC, Electrical, Plumbing, Foundation",
        "",
        "After calibrating with 3-5 closed deals, estimates typically tighten to",
        "within 10-15% of actual contractor SOWs.",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for i, note in enumerate(notes, 3):
        ws9.cell(row=i, column=1, value=note).font = _LABEL_FONT
    ws9.column_dimensions["A"].width = 70

    # Save
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_addr = "".join(c if c.isalnum() or c == "-" else "_" for c in (full_est.address or "property"))[:40]
        output_path = str(config.OUTPUT_DIR / f"rehab_estimate_{safe_addr}_{timestamp}.xlsx")

    wb.save(output_path)
    logger.info("Rehab report saved to %s", output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

def run_rehab_estimate(address: str = "", sqft: int = 0, bedrooms: int = 3,
                       bathrooms: float = 2.0, year_built: int = 0,
                       tier: int = 2, scope: str = "full",
                       region: str = DEFAULT_REGION,
                       output_path: str = "",
                       big_tickets: list[str] | None = None,
                       city: str = "", state: str = "") -> dict:
    """Run rehab estimation and generate report.

    Returns dict with estimates and report path.
    """
    resolved = resolve_region(city, state, region)
    if resolved != region and region != DEFAULT_REGION:
        logger.info("Region resolver: %s → %s (from %s, %s)", region, resolved, city, state)
    logger.info("Estimating rehab for: %s (%s sqft, Rehab Level %d, %s scope, region=%s)",
                address or "property", sqft, tier, scope, resolved)

    full_est = estimate_rehab(address, sqft, bedrooms, bathrooms, year_built,
                              tier=tier, scope="full", region=resolved,
                              big_tickets=big_tickets, city=city, state=state)

    wholetail_est = estimate_wholetail(address, sqft, bedrooms, bathrooms,
                                       year_built, region=resolved,
                                       city=city, state=state)

    # Generate report
    report_path = generate_rehab_report(full_est, wholetail_est, output_path)

    return {
        "full_estimate": full_est,
        "wholetail_estimate": wholetail_est,
        "report_path": report_path,
    }
