"""Deal analysis combining comping + rehab estimation into full deal packages.

Calculates MAO (Maximum Allowable Offer), ROI projections, holding costs,
selling costs, and financing scenarios for flip/wholesale/hold strategies.

Usage:
  python src/main.py analyze-deal --address "123 Main St, Knoxville, TN 37918"
  python src/main.py analyze-deal --address "123 Main St" --purchase-price 150000 --rehab-tier 2
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config
from comp_analyzer import (
    ARVResult, SubjectProperty, fetch_comparable_sales, fetch_pending_sales, fetch_rental_comps, fetch_subject_property,
    calculate_arv, _fmt_money, DEFAULT_RADIUS_MILES, DEFAULT_MONTHS_BACK,
)
from rehab_estimator import (
    RehabEstimate, TIER_NAMES, REHAB_LEVEL_NAMES, estimate_rehab, estimate_wholetail,
    DEFAULT_REGION, PER_SQFT_BY_SIZE, _size_bucket,
)

logger = logging.getLogger(__name__)

# ── Deal defaults ─────────────────────────────────────────────────────
DEFAULT_HARD_MONEY_RATE = 0.12     # 12% annual
DEFAULT_HARD_MONEY_POINTS = 2      # 2 origination points
DEFAULT_HARD_MONEY_TERM = 12       # months
DEFAULT_CONVENTIONAL_RATE = 0.07   # 7% annual (2026 rates)
DEFAULT_CONVENTIONAL_DOWN = 0.20   # 20% down
DEFAULT_INSURANCE_MONTHLY = 150.0
DEFAULT_UTILITIES_MONTHLY = 200.0
DEFAULT_AGENT_COMMISSION = 0.06    # 6% (3% buyer + 3% seller)
DEFAULT_CLOSING_COSTS_PCT = 0.025  # 2.5% of sale price
# Transfer tax varies by state. Lookup keyed on USPS state code; default
# falls through to TN's rate to preserve prior behavior for unknown states.
#   TN: $0.37 per $100 = 0.37%
#   AL: $0.50 per $500 = 0.10%
_TRANSFER_TAX_BY_STATE = {
    "TN": 0.0037,
    "AL": 0.001,
}
DEFAULT_TRANSFER_TAX_PCT = 0.0037


def transfer_tax_pct(state: str) -> float:
    return _TRANSFER_TAX_BY_STATE.get((state or "").upper(), DEFAULT_TRANSFER_TAX_PCT)
# ── Profit Calculator defaults (Google investor sheet, 2026-05-29) ────────────
# Replaces the legacy 75%/70% rules with explicit closing/holding/target-profit
# line items matching how the investor underwrites in the sheet.
#
# Flip / Wholesale Assignment:
#   All-In = ARV × (1 - closing% - holding%) - Repairs
#   MAO    = All-In - WS_Fee - (ARV × target_profit%)
DEFAULT_FLIP_CLOSING_PCT       = 0.07   # Sheet: 7% closing (commission + title + misc)
DEFAULT_FLIP_HOLDING_PCT       = 0.05   # Sheet: 5% holding (carry covering hold period)
DEFAULT_FLIP_TARGET_PROFIT_PCT = 0.15   # 15% target "FOR DISPO" (lowered from 16% on 2026-05-30)
DEFAULT_FLIP_WHOLESALE_FEE     = 15000  # WS fee for flip dispo (bumped from $10K → $15K, 2026-05-30)

# Wholetail (sold as-is with light cosmetic):
#   MAO = ARV × (1 - wholetail_closing%) - Repairs - Buyer_Profit - WS_Fee
DEFAULT_WHOLETAIL_CLOSING_PCT  = 0.10   # Sheet: 10% closing (higher — sold as-is to retail)
DEFAULT_WHOLETAIL_BUYER_PROFIT = 70000  # Sheet: $70K buyer profit baked in
DEFAULT_WHOLETAIL_WS_FEE       = 15000  # Sheet: $15K WS fee

# ── Novation defaults (2026-06-12) ────────────────────────────────────
# Novation = control listing without title. Math: Retail ARV − selling % −
# light rehab − holding − target spread = MAO to seller. Key differences from
# Wholetail: NO purchase closing costs (we never take title), shorter hold
# (cosmetic only + listing period), target spread on top instead of buyer-profit.
DEFAULT_NOVATION_SELLING_PCT      = 0.10   # 8-10% — buyer comm (2.5-3) + listing comm (2.5-3) + title (1.5) + AL transfer 0.10%
DEFAULT_NOVATION_HOLDING          = 800    # 60 days × ~$400/mo (insurance + utilities + lawn) — no mortgage
DEFAULT_NOVATION_TARGET_SPREAD    = 0.20   # 20% of retail ARV — higher than wholesale, lower than flip
DEFAULT_NOVATION_SELLER_BONUS     = 0      # Extra above seller's net to incentivize agreement (deal-by-deal)

# Rental (hold for cashflow):
#   All-In = ARV × (1 - rental_closing% - rental_holding%) - Repairs
#   MAO    = All-In - WS_Fee - (ARV × target_equity%)
DEFAULT_RENTAL_CLOSING_PCT     = 0.01   # Sheet: 1% closing
DEFAULT_RENTAL_HOLDING_PCT     = 0.01   # Sheet: 1% holding
DEFAULT_RENTAL_WS_FEE          = 5000   # Sheet: $5K WS fee
DEFAULT_TARGET_EQUITY_PCT      = 0.20   # Sheet: "make sure EQUITY PERCENTAGE IS AT 20%"

# Market tier rent % thresholds (sheet's reference column).
# Used to flag rent shortfall: actual_monthly_rent / all_in_number should be
# above the floor for the tier the property is in.
RENTAL_MARKET_TIERS = {
    "A": {"rent_pct_min": 0.010, "rent_pct_max": 0.015,
          "description": "Higher Mid Range to Highest. Pop 100k+, top of major cities, populous suburbs.",
          "examples": "Dallas, Hoover-luxury, Mountain Brook"},
    "B": {"rent_pct_min": 0.015, "rent_pct_max": 0.0225,
          "description": "Higher Low Range to Mid. Lower-end areas of major cities or suburbs ≤45 min, sub-$150K ARV.",
          "examples": "Cleveland, St. Louis, Alabaster, Pelham, Bessemer"},
    "C": {"rent_pct_min": 0.025, "rent_pct_max": 0.030,
          "description": "Lower End Ranges. Rural, fewer rental comps, high-crime areas.",
          "examples": "Jackson, Kansas City, rural Marshall AL"},
}

# Backward-compat aliases (keep for any external references; new code uses the explicit names above)
DEFAULT_WHOLESALE_FEE = DEFAULT_FLIP_WHOLESALE_FEE
DEFAULT_FLIP_RULE = 1 - DEFAULT_FLIP_CLOSING_PCT - DEFAULT_FLIP_HOLDING_PCT - DEFAULT_FLIP_TARGET_PROFIT_PCT  # ≈ 0.72
DEFAULT_WHOLESALE_RULE = 1 - DEFAULT_WHOLETAIL_CLOSING_PCT  # ≈ 0.90
DEFAULT_CAP_RATE_TARGET = 0.08
DEFAULT_CASH_ON_CASH_TARGET = 0.10
DEFAULT_VACANCY_RATE = 0.08
DEFAULT_MAINTENANCE_PCT = 0.01
DEFAULT_PROP_MGMT_PCT = 0.10


def classify_market_tier(arv: float = 0, city: str = "", state: str = "") -> str:
    """Rough A/B/C classification for rental rent% targeting.

    Heuristic — refine as needed:
      A: ARV > $300K OR known luxury suburb
      C: rural / unknown / sub-$80K ARV
      B: everything else (most Birmingham metro starter-to-mid stock)
    """
    luxury_cities = {"mountain brook", "vestavia hills", "homewood"}
    rural_states_default_c = set()  # extend if pipeline reaches truly rural
    if (city or "").strip().lower() in luxury_cities or arv >= 300_000:
        return "A"
    if arv and arv < 80_000:
        return "C"
    return "B"

# ── Data structures ───────────────────────────────────────────────────


@dataclass
class MAOResult:
    """Maximum Allowable Offer calculations."""
    flip_mao: float = 0.0
    wholesale_mao: float = 0.0
    hold_mao: float = 0.0
    flip_rule_pct: float = DEFAULT_FLIP_RULE
    wholesale_rule_pct: float = DEFAULT_WHOLESALE_RULE


@dataclass
class HoldingCosts:
    """Monthly and total holding costs."""
    months: float = 0.0
    mortgage_monthly: float = 0.0
    taxes_monthly: float = 0.0
    insurance_monthly: float = DEFAULT_INSURANCE_MONTHLY
    utilities_monthly: float = DEFAULT_UTILITIES_MONTHLY
    total_monthly: float = 0.0
    total: float = 0.0


@dataclass
class SellingCosts:
    """Costs to sell the property."""
    agent_commission: float = 0.0
    closing_costs: float = 0.0
    transfer_tax: float = 0.0
    total: float = 0.0


@dataclass
class FlipProjection:
    """Profit projection for a flip."""
    arv: float = 0.0
    purchase_price: float = 0.0
    rehab_cost: float = 0.0
    holding_costs: float = 0.0
    selling_costs: float = 0.0
    total_investment: float = 0.0
    net_profit: float = 0.0
    roi_pct: float = 0.0
    months_to_complete: float = 0.0


@dataclass
class WholesaleProjection:
    """Profit projection for a wholesale assignment."""
    arv: float = 0.0
    contract_price: float = 0.0
    assignment_fee: float = 0.0
    buyer_rehab_cost: float = 0.0
    buyer_profit_estimate: float = 0.0


@dataclass
class HoldProjection:
    """Cash flow projection for buy-and-hold."""
    purchase_price: float = 0.0
    rehab_cost: float = 0.0
    total_investment: float = 0.0
    estimated_rent_monthly: float = 0.0
    gross_annual_income: float = 0.0
    vacancy_loss: float = 0.0
    effective_income: float = 0.0
    expenses_annual: float = 0.0
    noi: float = 0.0
    cap_rate: float = 0.0
    debt_service_annual: float = 0.0
    cash_flow_annual: float = 0.0
    cash_on_cash: float = 0.0
    equity_year5: float = 0.0


@dataclass
class FinancingScenario:
    """Financing comparison."""
    name: str = ""
    down_payment: float = 0.0
    loan_amount: float = 0.0
    rate: float = 0.0
    monthly_payment: float = 0.0
    total_cost_of_money: float = 0.0
    points_cost: float = 0.0


@dataclass
class DealPackage:
    """Complete deal analysis package."""
    subject: SubjectProperty = field(default_factory=SubjectProperty)
    arv: ARVResult = field(default_factory=ARVResult)
    rehab_full: RehabEstimate = field(default_factory=RehabEstimate)
    rehab_wholetail: RehabEstimate = field(default_factory=RehabEstimate)
    mao: MAOResult = field(default_factory=MAOResult)
    flip: FlipProjection = field(default_factory=FlipProjection)
    wholesale: WholesaleProjection = field(default_factory=WholesaleProjection)
    hold: HoldProjection = field(default_factory=HoldProjection)
    holding_costs: HoldingCosts = field(default_factory=HoldingCosts)
    selling_costs: SellingCosts = field(default_factory=SellingCosts)
    financing: list = field(default_factory=list)
    recommendation: str = ""
    risk_factors: list = field(default_factory=list)
    # Profit Calculator breakdowns (Google investor sheet format, 2026-05-29)
    profit_breakdown: dict = field(default_factory=dict)      # Flip Profit Calc
    wholetail_breakdown: dict = field(default_factory=dict)   # Wholetail Calculator
    novation_breakdown: dict = field(default_factory=dict)    # Novation Calculator (2026-06-12)
    rental_breakdown: dict = field(default_factory=dict)      # Rental Calculator
    market_tier: str = "B"                                    # A / B / C per sheet
    # Comp list preserved for the Comp Analysis tab (ChatARV-style derivation)
    comps: list = field(default_factory=list)
    # Rental comps (active FOR_RENT listings near subject) for Rental Calc tab
    rental_comps: list = field(default_factory=list)
    # Pending sales (PENDING/UNDER_CONTRACT listings near subject) — directional ARV signal
    pending_comps: list = field(default_factory=list)


# ── Calculation engine ────────────────────────────────────────────────

def _calc_monthly_payment(principal: float, annual_rate: float, months: int) -> float:
    """Standard amortization monthly payment."""
    if annual_rate == 0:
        return principal / months if months else 0
    r = annual_rate / 12
    return principal * (r * (1 + r) ** months) / ((1 + r) ** months - 1)


def _estimate_monthly_rent(arv: float, sqft: int, bedrooms: int) -> float:
    """Estimate monthly rent based on 1% rule and property characteristics."""
    # 1% rule as baseline (conservative for Knoxville)
    rent_pct_rule = arv * 0.008  # 0.8% for Knoxville (below 1% rule)
    # Sqft-based estimate
    rent_sqft = sqft * 0.75 if sqft else 0  # ~$0.75/sqft for Knoxville
    # Bedroom-based estimate
    rent_bed = 600 + (bedrooms * 200)  # base $600 + $200/bed
    # Average the estimates
    estimates = [e for e in [rent_pct_rule, rent_sqft, rent_bed] if e > 0]
    return round(sum(estimates) / len(estimates)) if estimates else 1200


def calculate_mao(arv_mid: float, rehab_cost: float,
                  wholesale_fee: float = DEFAULT_FLIP_WHOLESALE_FEE,
                  closing_pct: float = DEFAULT_FLIP_CLOSING_PCT,
                  holding_pct: float = DEFAULT_FLIP_HOLDING_PCT,
                  target_profit_pct: float = DEFAULT_FLIP_TARGET_PROFIT_PCT,
                  wholetail_closing_pct: float = DEFAULT_WHOLETAIL_CLOSING_PCT,
                  wholetail_buyer_profit: float = DEFAULT_WHOLETAIL_BUYER_PROFIT,
                  wholetail_ws_fee: float = DEFAULT_WHOLETAIL_WS_FEE,
                  rental_closing_pct: float = DEFAULT_RENTAL_CLOSING_PCT,
                  rental_holding_pct: float = DEFAULT_RENTAL_HOLDING_PCT,
                  target_equity_pct: float = DEFAULT_TARGET_EQUITY_PCT,
                  rental_ws_fee: float = DEFAULT_RENTAL_WS_FEE,
                  ) -> MAOResult:
    """Max Allowable Offer for each exit strategy — Google investor sheet format.

    Flip / Wholesale Assignment (Profit Calculator):
      MAO = ARV × (1 - closing% - holding% - target_profit%) - Repairs - WS_Fee
          ≈ ARV × 0.72 - Repairs - $15K   (with current defaults)

    Wholetail (Wholetail Calculator):
      MAO = ARV × (1 - wholetail_closing%) - Repairs - Buyer_Profit - WS_Fee
          = ARV × 0.90 - Repairs - $70K - $15K

    Rental Hold (Rental Calculator, 20% equity target):
      MAO = ARV × (1 - rental_closing% - rental_holding% - target_equity%) - Repairs - WS_Fee
          = ARV × 0.78 - Repairs - $5K
    """
    # Flip
    flip_mao = arv_mid * (1 - closing_pct - holding_pct - target_profit_pct) - rehab_cost - wholesale_fee

    # Wholetail
    wholetail_mao = arv_mid * (1 - wholetail_closing_pct) - rehab_cost - wholetail_buyer_profit - wholetail_ws_fee

    # Rental hold
    hold_mao = arv_mid * (1 - rental_closing_pct - rental_holding_pct - target_equity_pct) - rehab_cost - rental_ws_fee

    return MAOResult(
        flip_mao=max(0, round(flip_mao)),
        wholesale_mao=max(0, round(wholetail_mao)),  # naming kept for back-compat; this IS wholetail now
        hold_mao=max(0, round(hold_mao)),
        flip_rule_pct=round(1 - closing_pct - holding_pct - target_profit_pct, 4),
        wholesale_rule_pct=round(1 - wholetail_closing_pct, 4),
    )


def calculate_profit_breakdown(arv: float, purchase_price: float, rehab_cost: float,
                               closing_pct: float = DEFAULT_FLIP_CLOSING_PCT,
                               holding_pct: float = DEFAULT_FLIP_HOLDING_PCT,
                               wholesale_fee: float = DEFAULT_FLIP_WHOLESALE_FEE,
                               ) -> dict:
    """Return Profit Calculator line-item breakdown matching the Google sheet.

    Layout per the sheet (cumulative deductions from ARV):
      ARV                            $X
      - Closing Costs (7%)           $Y      running: $X-Y
      - Repairs                      $R      running: ...
      - Holding (5%)                 $H      running: All-In  ← key metric
      Contract Price to Seller      $P
      + Your Wholesale Fee           $W
      = Contract Price to Buyer      $P+W
      Potential Profit              All-In - (P+W)
      Profit % (FOR DISPO)           Potential / ARV
    """
    closing = arv * closing_pct
    holding = arv * holding_pct
    after_closing = arv - closing
    after_repairs = after_closing - rehab_cost
    all_in = after_repairs - holding
    contract_to_buyer = purchase_price + wholesale_fee
    potential_profit = all_in - contract_to_buyer
    profit_pct = potential_profit / arv if arv else 0
    return {
        "arv": round(arv),
        "closing_costs": round(closing),
        "closing_pct": closing_pct,
        "after_closing": round(after_closing),
        "repairs": round(rehab_cost),
        "after_repairs": round(after_repairs),
        "holding_costs": round(holding),
        "holding_pct": holding_pct,
        "all_in_number": round(all_in),
        "contract_to_seller": round(purchase_price),
        "wholesale_fee": round(wholesale_fee),
        "contract_to_buyer": round(contract_to_buyer),
        "potential_profit": round(potential_profit),
        "profit_pct": round(profit_pct, 4),
    }


def calculate_wholetail_breakdown(arv: float, purchase_price: float, rehab_cost: float,
                                  closing_pct: float = DEFAULT_WHOLETAIL_CLOSING_PCT,
                                  buyer_profit: float = DEFAULT_WHOLETAIL_BUYER_PROFIT,
                                  wholesale_fee: float = DEFAULT_WHOLETAIL_WS_FEE,
                                  ) -> dict:
    """Wholetail Calculator breakdown matching the sheet."""
    closing = arv * closing_pct
    mao = arv - closing - rehab_cost - buyer_profit - wholesale_fee
    your_profit = mao - purchase_price  # what YOU make if you bought at purchase_price
    return {
        "cmv": round(arv),
        "closing_costs": round(closing),
        "closing_pct": closing_pct,
        "repairs": round(rehab_cost),
        "buyer_profit": round(buyer_profit),
        "wholesale_fee": round(wholesale_fee),
        "mao": round(mao),
        "purchase_price": round(purchase_price),
        "your_profit": round(your_profit),
    }


def calculate_novation_breakdown(arv: float, rehab_cost: float,
                                  selling_pct: float = DEFAULT_NOVATION_SELLING_PCT,
                                  holding: float = DEFAULT_NOVATION_HOLDING,
                                  target_spread_pct: float = DEFAULT_NOVATION_TARGET_SPREAD,
                                  seller_bonus: float = DEFAULT_NOVATION_SELLER_BONUS,
                                  ) -> dict:
    """Novation Calculator breakdown (2026-06-12).

    Novation = listed retail under seller's name, deed transfers seller→buyer,
    we collect the spread. NO purchase closing costs (never take title).
    Light cosmetic rehab only — sets up for retail listing, not full renovation.

    Formula:
      MAO to seller = Retail ARV × (1 - selling%) - light rehab - holding
                      - (Retail ARV × target spread %) - seller bonus

    Note `rehab_cost` should be the LIGHT/WHOLETAIL rehab estimate, not the
    full-rehab number — novation properties get cosmetic-only refresh.
    """
    selling_costs = arv * selling_pct
    target_spread = arv * target_spread_pct
    mao = arv - selling_costs - rehab_cost - holding - target_spread - seller_bonus
    return {
        "retail_arv": round(arv),
        "selling_pct": selling_pct,
        "selling_costs": round(selling_costs),
        "light_rehab": round(rehab_cost),
        "holding": round(holding),
        "target_spread_pct": target_spread_pct,
        "target_spread": round(target_spread),
        "seller_bonus": round(seller_bonus),
        "mao": round(mao),
    }


def calculate_rental_breakdown(arv: float, purchase_price: float, rehab_cost: float,
                               monthly_rent: float = 0,
                               closing_pct: float = DEFAULT_RENTAL_CLOSING_PCT,
                               holding_pct: float = DEFAULT_RENTAL_HOLDING_PCT,
                               wholesale_fee: float = DEFAULT_RENTAL_WS_FEE,
                               market_tier: str = "B",
                               ) -> dict:
    """Rental Calculator breakdown matching the sheet (equity + rent% check)."""
    closing = arv * closing_pct
    holding = arv * holding_pct
    after_closing = arv - closing
    after_repairs = after_closing - rehab_cost
    all_in_number = after_repairs - holding
    contract_to_buyer = purchase_price + wholesale_fee
    equity = all_in_number - contract_to_buyer
    equity_pct = equity / arv if arv else 0
    # Rent % is computed against All-In (the sheet's "j2*0.8 - h12" / monthly rent formula)
    rent_pct = (monthly_rent / all_in_number) if (monthly_rent and all_in_number > 0) else 0
    tier_info = RENTAL_MARKET_TIERS.get(market_tier, RENTAL_MARKET_TIERS["B"])
    rent_pct_ok = tier_info["rent_pct_min"] <= rent_pct <= tier_info["rent_pct_max"] if rent_pct else None
    equity_ok = equity_pct >= DEFAULT_TARGET_EQUITY_PCT
    return {
        "arv": round(arv),
        "closing_costs": round(closing),
        "closing_pct": closing_pct,
        "after_closing": round(after_closing),
        "repairs": round(rehab_cost),
        "after_repairs": round(after_repairs),
        "holding_costs": round(holding),
        "holding_pct": holding_pct,
        "all_in_number": round(all_in_number),
        "contract_to_seller": round(purchase_price),
        "wholesale_fee": round(wholesale_fee),
        "contract_to_buyer": round(contract_to_buyer),
        "equity": round(equity),
        "equity_pct": round(equity_pct, 4),
        "equity_ok": equity_ok,
        "monthly_rent": round(monthly_rent),
        "rent_pct": round(rent_pct, 4),
        "market_tier": market_tier,
        "tier_min_rent_pct": tier_info["rent_pct_min"],
        "tier_max_rent_pct": tier_info["rent_pct_max"],
        "rent_pct_ok": rent_pct_ok,
    }


def calculate_holding_costs(purchase_price: float, rehab_months: float,
                            annual_tax: float = 0,
                            hard_money_rate: float = DEFAULT_HARD_MONEY_RATE) -> HoldingCosts:
    """Calculate holding costs during rehab period."""
    # Hard money interest-only payments
    mortgage_monthly = purchase_price * hard_money_rate / 12
    # Property taxes (use provided or estimate at 1% of purchase)
    taxes_monthly = (annual_tax / 12) if annual_tax else (purchase_price * 0.01 / 12)
    total_monthly = mortgage_monthly + taxes_monthly + DEFAULT_INSURANCE_MONTHLY + DEFAULT_UTILITIES_MONTHLY
    # Add 1 month for listing/closing after rehab
    total_months = rehab_months + 1

    return HoldingCosts(
        months=total_months,
        mortgage_monthly=round(mortgage_monthly),
        taxes_monthly=round(taxes_monthly),
        total_monthly=round(total_monthly),
        total=round(total_monthly * total_months),
    )


def calculate_selling_costs(sale_price: float, state: str = "") -> SellingCosts:
    """Calculate costs to sell the property.

    Empty state falls back to DEFAULT_PROPERTY_STATE (currently AL).
    SiftStack pipeline callers pass state explicitly from the record;
    the fallback is for REI Skill plugin callers (Co-Work sessions)
    where state may be omitted.
    """
    if not state:
        from state_resolver import DEFAULT_PROPERTY_STATE
        state = DEFAULT_PROPERTY_STATE
    commission = sale_price * DEFAULT_AGENT_COMMISSION
    closing = sale_price * DEFAULT_CLOSING_COSTS_PCT
    transfer = sale_price * transfer_tax_pct(state)

    return SellingCosts(
        agent_commission=round(commission),
        closing_costs=round(closing),
        transfer_tax=round(transfer),
        total=round(commission + closing + transfer),
    )


def calculate_flip(arv_mid: float, purchase_price: float, rehab_cost: float,
                   holding: HoldingCosts, selling: SellingCosts,
                   rehab_months: float) -> FlipProjection:
    """Calculate flip profit projection."""
    total_investment = purchase_price + rehab_cost + holding.total + selling.total
    net_profit = arv_mid - total_investment
    roi = (net_profit / (purchase_price + rehab_cost) * 100) if (purchase_price + rehab_cost) else 0

    return FlipProjection(
        arv=round(arv_mid),
        purchase_price=round(purchase_price),
        rehab_cost=round(rehab_cost),
        holding_costs=round(holding.total),
        selling_costs=round(selling.total),
        total_investment=round(total_investment),
        net_profit=round(net_profit),
        roi_pct=round(roi, 1),
        months_to_complete=round(rehab_months + 2, 1),  # +2 for listing/closing
    )


def calculate_wholesale(arv_mid: float, contract_price: float,
                        rehab_cost: float,
                        assignment_fee: float = DEFAULT_WHOLESALE_FEE) -> WholesaleProjection:
    """Calculate wholesale assignment projection."""
    buyer_total = contract_price + assignment_fee + rehab_cost
    buyer_profit = arv_mid - buyer_total - (arv_mid * (DEFAULT_AGENT_COMMISSION + DEFAULT_CLOSING_COSTS_PCT))

    return WholesaleProjection(
        arv=round(arv_mid),
        contract_price=round(contract_price),
        assignment_fee=round(assignment_fee),
        buyer_rehab_cost=round(rehab_cost),
        buyer_profit_estimate=round(buyer_profit),
    )


def calculate_hold(purchase_price: float, rehab_cost: float, arv_mid: float,
                   sqft: int, bedrooms: int, annual_tax: float = 0) -> HoldProjection:
    """Calculate buy-and-hold cash flow projection."""
    total_investment = purchase_price + rehab_cost
    rent = _estimate_monthly_rent(arv_mid, sqft, bedrooms)
    gross_annual = rent * 12
    vacancy = gross_annual * DEFAULT_VACANCY_RATE
    effective = gross_annual - vacancy

    # Annual expenses
    taxes = annual_tax if annual_tax else purchase_price * 0.01
    insurance = DEFAULT_INSURANCE_MONTHLY * 12
    maintenance = arv_mid * DEFAULT_MAINTENANCE_PCT
    management = effective * DEFAULT_PROP_MGMT_PCT
    expenses = taxes + insurance + maintenance + management

    noi = effective - expenses
    cap_rate = (noi / total_investment * 100) if total_investment else 0

    # Debt service (conventional financing)
    down = total_investment * DEFAULT_CONVENTIONAL_DOWN
    loan = total_investment - down
    monthly_pmt = _calc_monthly_payment(loan, DEFAULT_CONVENTIONAL_RATE, 360)
    debt_service = monthly_pmt * 12

    cash_flow = noi - debt_service
    cash_on_cash = (cash_flow / down * 100) if down else 0

    # 5-year equity buildup (3% annual appreciation)
    equity_5yr = arv_mid * (1.03 ** 5) - arv_mid + down

    return HoldProjection(
        purchase_price=round(purchase_price),
        rehab_cost=round(rehab_cost),
        total_investment=round(total_investment),
        estimated_rent_monthly=round(rent),
        gross_annual_income=round(gross_annual),
        vacancy_loss=round(vacancy),
        effective_income=round(effective),
        expenses_annual=round(expenses),
        noi=round(noi),
        cap_rate=round(cap_rate, 1),
        debt_service_annual=round(debt_service),
        cash_flow_annual=round(cash_flow),
        cash_on_cash=round(cash_on_cash, 1),
        equity_year5=round(equity_5yr),
    )


def calculate_financing(purchase_price: float, rehab_cost: float) -> list[FinancingScenario]:
    """Calculate financing comparison scenarios."""
    total_needed = purchase_price + rehab_cost
    scenarios = []

    # Cash purchase
    scenarios.append(FinancingScenario(
        name="All Cash",
        down_payment=total_needed,
        loan_amount=0,
        rate=0,
        monthly_payment=0,
        total_cost_of_money=0,
        points_cost=0,
    ))

    # Hard money
    hm_points = total_needed * (DEFAULT_HARD_MONEY_POINTS / 100)
    hm_monthly = total_needed * DEFAULT_HARD_MONEY_RATE / 12
    scenarios.append(FinancingScenario(
        name="Hard Money (12%, 2pts, 12mo)",
        down_payment=hm_points,
        loan_amount=total_needed,
        rate=DEFAULT_HARD_MONEY_RATE,
        monthly_payment=round(hm_monthly),
        total_cost_of_money=round(hm_monthly * DEFAULT_HARD_MONEY_TERM + hm_points),
        points_cost=round(hm_points),
    ))

    # Conventional
    conv_down = total_needed * DEFAULT_CONVENTIONAL_DOWN
    conv_loan = total_needed - conv_down
    conv_monthly = _calc_monthly_payment(conv_loan, DEFAULT_CONVENTIONAL_RATE, 360)
    scenarios.append(FinancingScenario(
        name=f"Conventional ({DEFAULT_CONVENTIONAL_DOWN:.0%} down, {DEFAULT_CONVENTIONAL_RATE:.0%})",
        down_payment=round(conv_down),
        loan_amount=round(conv_loan),
        rate=DEFAULT_CONVENTIONAL_RATE,
        monthly_payment=round(conv_monthly),
        total_cost_of_money=round(conv_monthly * 360 - conv_loan),
        points_cost=0,
    ))

    # Seller financing (hypothetical)
    sf_down = total_needed * 0.10
    sf_loan = total_needed - sf_down
    sf_rate = 0.08
    sf_monthly = _calc_monthly_payment(sf_loan, sf_rate, 240)  # 20-year term
    scenarios.append(FinancingScenario(
        name="Seller Financing (10% down, 8%, 20yr)",
        down_payment=round(sf_down),
        loan_amount=round(sf_loan),
        rate=sf_rate,
        monthly_payment=round(sf_monthly),
        total_cost_of_money=round(sf_monthly * 240 - sf_loan),
        points_cost=0,
    ))

    return scenarios


def _assess_risk(arv: ARVResult, flip: FlipProjection, subject: SubjectProperty) -> list[str]:
    """Assess risk factors for the deal."""
    risks = []

    if arv.confidence == "low":
        risks.append(f"LOW CONFIDENCE ARV — {arv.confidence_reason}")
    if arv.spread_pct > 15:
        risks.append(f"Wide comp spread ({arv.spread_pct:.0f}%) — ARV uncertainty")
    if arv.comp_count < 3:
        risks.append(f"Only {arv.comp_count} comps — limited market data")
    if flip.roi_pct < 15:
        risks.append(f"Thin flip margin ({flip.roi_pct:.0f}% ROI) — little room for error")
    if flip.net_profit < 20000:
        risks.append(f"Low profit (${flip.net_profit:,.0f}) — may not justify effort/risk")
    if subject.year_built and subject.year_built < 1960:
        risks.append(f"Pre-1960 construction — potential lead paint, asbestos, foundation issues")
    if subject.mls_status and "active" in subject.mls_status.lower():
        risks.append("Property currently listed — competing with retail buyers")

    if not risks:
        risks.append("No major risk factors identified")

    return risks


def _make_recommendation(flip: FlipProjection, wholesale: WholesaleProjection,
                         hold: HoldProjection, arv: ARVResult) -> str:
    """Generate a go/no-go recommendation."""
    if arv.confidence == "none":
        return "INSUFFICIENT DATA — Cannot recommend. Need more comparable sales."

    strategies = []
    if flip.roi_pct >= 25 and flip.net_profit >= 25000:
        strategies.append(("FLIP", flip.roi_pct, f"${flip.net_profit:,.0f} profit, {flip.roi_pct:.0f}% ROI"))
    if wholesale.assignment_fee >= 5000 and wholesale.buyer_profit_estimate > 20000:
        strategies.append(("WHOLESALE", 100, f"${wholesale.assignment_fee:,.0f} assignment fee, buyer profits ${wholesale.buyer_profit_estimate:,.0f}"))
    if hold.cash_on_cash >= 8:
        strategies.append(("BUY & HOLD", hold.cash_on_cash, f"{hold.cash_on_cash:.1f}% CoC, ${hold.cash_flow_annual:,.0f}/yr cash flow"))

    if not strategies:
        if flip.roi_pct >= 15:
            return f"MARGINAL FLIP — {flip.roi_pct:.0f}% ROI. Proceed with caution; negotiate harder on price."
        return "NO-GO — Numbers don't work at this price point. Counter-offer or pass."

    strategies.sort(key=lambda x: x[1], reverse=True)
    best = strategies[0]
    return f"GO — Best strategy: {best[0]} ({best[2]})"


# ── Excel report generation ──────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(name="Calibri", bold=True, color="006100")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_RED_FONT = Font(name="Calibri", bold=True, color="9C0006")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_MONEY_FMT = '#,##0'
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))

# ── Playbook design tokens (from Marketing Playbook Q3 2026, applied 2026-06-13) ──
# Reusable across MAO Calc, Wholetail, Novation (and future Profit/Rental beautify passes).
# Ink/typography/rule tokens from the artifact CSS :root — light-theme values.
_PB_INK        = "1B1D22"
_PB_INK_MUTE   = "55575E"
_PB_INK_FAINT  = "8A8B8F"
_PB_RULE       = "C0BCB2"
_PB_RULE_SOFT  = "E1DED4"
_PB_MONEY      = "0B6B47"   # financial results (MAO, etc.)
_PB_MONEY_TINT = "E4EFE8"
# Priority stripe palette — one per strategy, matches playbook campaign cards:
_PB_P1         = "A63232"   # red   — Flip (primary offer)
_PB_P1_TINT    = "F4E5E2"
_PB_P2         = "B98A2E"   # gold  — Wholetail (light cosmetic)
_PB_P2_TINT    = "F5EBD5"
_PB_P3         = "4A6D74"   # blue  — Rental Hold (cashflow)
_PB_P3_TINT    = "DDE7EA"
_PB_P4         = "3A6B5C"   # teal  — Novation (retail-listed, no title) — distinct from P3
_PB_P4_TINT    = "DEE7E3"

_PB_FILL_MONEY_TINT = PatternFill(start_color=_PB_MONEY_TINT, end_color=_PB_MONEY_TINT, fill_type="solid")
_PB_FILL_P1T = PatternFill(start_color=_PB_P1_TINT, end_color=_PB_P1_TINT, fill_type="solid")
_PB_FILL_P2T = PatternFill(start_color=_PB_P2_TINT, end_color=_PB_P2_TINT, fill_type="solid")
_PB_FILL_P3T = PatternFill(start_color=_PB_P3_TINT, end_color=_PB_P3_TINT, fill_type="solid")
_PB_FILL_P4T = PatternFill(start_color=_PB_P4_TINT, end_color=_PB_P4_TINT, fill_type="solid")

_PB_BORDER_BOTTOM_RULE = Border(bottom=Side(style="thin", color=_PB_RULE))
_PB_BORDER_TOP_RULE = Border(top=Side(style="thin", color=_PB_RULE))
_PB_BORDER_BOTTOM_SOFT = Border(bottom=Side(style="thin", color=_PB_RULE_SOFT))


def _pb_hero(ws, subject, section_label: str, hero_title: str):
    """Render the editorial hero block (rows 1-4) on any playbook-styled tab.

    Row 1 — uppercase eyebrow (section label, letter-spaced)
    Row 2 — big serif hero title (Georgia 28pt)
    Row 3 — subject subtitle (italic Georgia 12pt with address + specs)
    Row 4 — thin bottom rule (empty row with bordered cells)
    """
    c = ws.cell(row=1, column=1, value=section_label)
    c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
    c.alignment = Alignment(horizontal="left", vertical="bottom")
    ws.row_dimensions[1].height = 20

    c = ws.cell(row=2, column=1, value=hero_title)
    c.font = Font(name="Georgia", bold=True, size=28, color=_PB_INK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=5)
    ws.row_dimensions[2].height = 40

    subj_line = f"{subject.address}, {subject.city}, {subject.state} {subject.zip_code}"
    if subject.sqft: subj_line += f"  ·  {subject.sqft:,} sqft"
    if subject.bedrooms or subject.bathrooms: subj_line += f"  ·  {subject.bedrooms}bd/{subject.bathrooms}ba"
    c = ws.cell(row=3, column=1, value=subj_line)
    c.font = Font(name="Georgia", italic=True, size=12, color=_PB_INK_MUTE)
    ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=5)
    ws.row_dimensions[3].height = 22

    for col in range(1, 6):
        ws.cell(row=4, column=col).border = _PB_BORDER_BOTTOM_RULE
    ws.row_dimensions[4].height = 4


def _pb_section_header(ws, row, tag_text, section_title, thesis, tag_ink):
    """3-row playbook section header: tag chip → serif title → italic thesis."""
    tint_fill = {
        _PB_P1: _PB_FILL_P1T, _PB_P2: _PB_FILL_P2T,
        _PB_P3: _PB_FILL_P3T, _PB_P4: _PB_FILL_P4T,
    }.get(tag_ink, _PB_FILL_P1T)

    c = ws.cell(row=row, column=1, value=f"  {tag_text}")
    c.font = Font(name="Calibri", bold=True, size=9, color=tag_ink)
    c.fill = tint_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20

    c = ws.cell(row=row + 1, column=1, value=section_title)
    c.font = Font(name="Georgia", bold=True, size=18, color=_PB_INK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row + 1, end_row=row + 1, start_column=1, end_column=5)
    ws.row_dimensions[row + 1].height = 28

    c = ws.cell(row=row + 2, column=1, value=thesis)
    c.font = Font(name="Georgia", italic=True, size=10, color=_PB_INK_MUTE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row + 2, end_row=row + 2, start_column=1, end_column=5)
    ws.row_dimensions[row + 2].height = 18


def _pb_input(ws, row, col, value, tint_fill, fmt='"$"#,##0'):
    """Playbook-styled editable input cell: mono value, right-aligned, tinted."""
    c = ws.cell(row=row, column=col, value=value)
    c.number_format = fmt
    c.fill = tint_fill
    c.border = _PB_BORDER_BOTTOM_SOFT
    c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
    c.alignment = Alignment(horizontal="right", vertical="center")
    return c


def _pb_label(ws, row, text):
    """Playbook-styled row label — Calibri sans, indented, subtle bottom rule."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", size=11, color=_PB_INK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border = _PB_BORDER_BOTTOM_SOFT
    return c


def _pb_mao_row(ws, row, formula, caption="CONTRACT TO SELLER", big=False, value_col=2):
    """Big money-green MAO result row with 'CONTRACT TO SELLER' eyebrow caption."""
    lbl = ws.cell(row=row, column=1, value="MAO")
    lbl.font = Font(name="Calibri", bold=True, size=11, color=_PB_INK_MUTE)
    lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lbl.fill = _PB_FILL_MONEY_TINT

    c = ws.cell(row=row, column=value_col, value=formula)
    c.number_format = '"$"#,##0'
    c.fill = _PB_FILL_MONEY_TINT
    c.font = Font(name="Consolas", bold=True, size=18 if big else 15, color=_PB_MONEY)
    c.alignment = Alignment(horizontal="right", vertical="center")

    # Fill intermediate cells with money tint for visual continuity
    for col in range(2, value_col):
        ws.cell(row=row, column=col).fill = _PB_FILL_MONEY_TINT
    # Caption in next column after value
    cap_col = value_col + 1
    if cap_col <= 5:
        cap = ws.cell(row=row, column=cap_col, value=caption)
        cap.font = Font(name="Calibri", bold=True, size=9, color=_PB_MONEY)
        cap.fill = _PB_FILL_MONEY_TINT
        cap.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.row_dimensions[row].height = 36 if big else 30


def _pb_col_headers(ws, row, specs):
    """Small uppercase column labels (INPUT / $ EQUIVALENT / etc.).

    specs: list of (col_num, header_text) tuples.
    """
    for col, text in specs:
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Calibri", bold=True, size=8, color=_PB_INK_FAINT)
        c.alignment = Alignment(horizontal="right", vertical="bottom", indent=1)
        c.border = _PB_BORDER_BOTTOM_RULE
    ws.cell(row=row, column=1).border = _PB_BORDER_BOTTOM_RULE
    ws.row_dimensions[row].height = 18


def _pb_footer(ws, row, text):
    """Footer eyebrow with top rule."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Calibri", bold=True, size=8, color=_PB_INK_FAINT)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = _PB_BORDER_TOP_RULE
    ws.row_dimensions[row].height = 22


def _write_headers(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _write_calc_line(ws, row, label, percent, dollar, running, *,
                     bold=False, fill=None, border=True):
    """Write one line of a Profit/Wholetail/Rental Calculator table.

    Layout matches the Google sheet: A=label, B=blank, C=%, D=$ this step, E=running total.
    """
    cells = [(1, label), (3, percent), (4, dollar), (5, running)]
    for col, val in cells:
        c = ws.cell(row=row, column=col, value=val)
        if col == 1:
            c.font = Font(name="Calibri", bold=bold, size=12 if bold else 11,
                          color="2F5496" if bold else "222222")
        elif col == 3 and val is not None:
            c.value = val
            c.number_format = "0.00%"
            c.alignment = Alignment(horizontal="right")
        elif col in (4, 5) and val is not None:
            c.number_format = '"$"#,##0'
            c.alignment = Alignment(horizontal="right")
            if col == 5:  # running total — emphasize
                c.font = Font(name="Calibri", bold=True, size=12)
        if fill:
            c.fill = fill
        if border:
            c.border = _THIN_BORDER


def _build_profit_calc_tab(wb, pkg, breakdown, title, scope_note):
    """Build a Profit/Rental Calculator tab matching the Google sheet's layout."""
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    addr = f"{pkg.subject.address}, {pkg.subject.city}, {pkg.subject.state} {pkg.subject.zip_code}"
    ws.cell(row=2, column=1, value=addr).font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=scope_note).font = _LABEL_FONT
    return ws


def generate_deal_report(pkg: DealPackage, output_path: str = "") -> str:
    """Generate deal analysis Excel workbook (Google investor sheet format)."""
    wb = Workbook()

    # ── Tab 1: Profit Calculator — beautified 2026-06-13 (in-place restyle) ──
    # IN-PLACE: all cell coordinates preserved. Bidirectional refs from MAO Calc
    # to C29/C30/C31 remain intact. Snippet cross-refs to Comp Analysis / Rental
    # Calc also preserved. Only visual styling changes.
    ws0 = wb.active
    ws0.title = "Profit Calculator"
    # R1 — editorial eyebrow (uppercase, letter-spaced)
    c = ws0.cell(row=1, column=1, value="PROFIT CALCULATOR  ·  FLIP / WHOLESALE ASSIGNMENT")
    c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
    ws0.row_dimensions[1].height = 20

    addr = f"{pkg.subject.address}, {pkg.subject.city}, {pkg.subject.state} {pkg.subject.zip_code}"
    # R2 — Address (clickable hyperlink → Zillow subject page with photos), editorial serif treatment
    subj_zillow_url = (
        f"https://www.zillow.com/homes/{pkg.subject.address.replace(' ', '-')}-"
        f"{pkg.subject.city.replace(' ', '-')}-{pkg.subject.state}-{pkg.subject.zip_code}/"
    )
    addr_cell = ws0.cell(row=2, column=1, value=f"🏠 {addr}  📷 (click for pics on Zillow)")
    addr_cell.hyperlink = subj_zillow_url
    addr_cell.font = Font(name="Georgia", bold=True, size=22, color="0563C1", underline="single")
    ws0.row_dimensions[2].height = 32
    # R3 — Subject property detail line: sqft / beds-baths / year / lot / property type / zestimate
    s = pkg.subject
    subj_parts = []
    if s.sqft: subj_parts.append(f"{s.sqft:,} sqft")
    if s.bedrooms or s.bathrooms: subj_parts.append(f"{s.bedrooms}bd/{s.bathrooms}ba")
    if s.year_built: subj_parts.append(f"Built {s.year_built}")
    if s.lot_sqft: subj_parts.append(f"Lot {s.lot_sqft:,} sqft")
    if s.property_type: subj_parts.append(s.property_type)
    if s.zestimate: subj_parts.append(f"Zestimate ${s.zestimate:,.0f}")
    if s.last_sold_price and s.last_sold_date:
        subj_parts.append(f"Last sold {s.last_sold_date} ${s.last_sold_price:,.0f}")
    c = ws0.cell(row=3, column=1, value=" · ".join(subj_parts))
    c.font = Font(name="Georgia", italic=True, size=12, color=_PB_INK_MUTE)
    ws0.row_dimensions[3].height = 22
    # Generated date stamp + headline summary — pushed to a smaller line further down
    # (kept in cell metadata via the headline at row 4, no longer needs its own row)

    # Cell-coord constants for the Profit Calc tab formulas
    # INPUT cells (no formula — user edits): E6=ARV, C7=Closing%, D8=Repairs,
    #   C9=Holding%, E11=Purchase, E12=WS Fee
    # FORMULA cells: D7, E7, E8, D9, E9 (All-In), E13, E15 (Profit), E16 (Profit%),
    #   plus the Buyer Ceiling Lookup rows
    pb = pkg.profit_breakdown or {}
    INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
    HEADLINE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # bright green
    if pb:
        # ── R4 HEADLINE — MAO at editable target profit % (money-green in tint) ──
        c = ws0.cell(row=4, column=1, value="  RECOMMENDED MAX OFFER  ·  AT TARGET PROFIT %")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_MONEY)
        c.fill = _PB_FILL_MONEY_TINT
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 5):
            ws0.cell(row=4, column=col).fill = _PB_FILL_MONEY_TINT
        # E4 — MAO formula, references C12 (editable target %)
        headline_cell = ws0.cell(row=4, column=5, value="=E6*(1-C7-C9-C12)-D8-E15")
        headline_cell.number_format = '"$"#,##0'
        headline_cell.font = Font(name="Consolas", bold=True, size=18, color=_PB_MONEY)
        headline_cell.fill = _PB_FILL_MONEY_TINT
        headline_cell.alignment = Alignment(horizontal="right", vertical="center")
        ws0.row_dimensions[4].height = 32

        # ── R5 column labels ──
        _pb_col_headers(ws0, 5, [(3, "%"), (4, "AMOUNT"), (5, "RUNNING")])

        # ── R6 ARV / CMV (P1 red tint) ──
        _pb_label(ws0, 6, "ARV / CMV")
        c = ws0.cell(row=6, column=5, value=pb["arv"]); c.number_format = '"$"#,##0'
        c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
        c.fill = _PB_FILL_P1T; c.border = _PB_BORDER_BOTTOM_SOFT
        c.alignment = Alignment(horizontal="right", vertical="center")

        # ── R7 Closing Costs ──
        _pb_label(ws0, 7, "Less Closing Costs")
        _pb_input(ws0, 7, 3, pb["closing_pct"], _PB_FILL_P1T, "0.00%")
        for col_num, formula in [(4, "=E6*C7"), (5, "=E6-D7")]:
            c = ws0.cell(row=7, column=col_num, value=formula)
            c.number_format = '"$"#,##0'
            c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _PB_BORDER_BOTTOM_SOFT

        # ── R8 Repairs (RED input — at-a-glance rehab convention) ──
        _pb_label(ws0, 8, "Less Repairs")
        c = ws0.cell(row=8, column=4, value=pb["repairs"]); c.number_format = '"$"#,##0'
        c.fill = _RED_FILL; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c = ws0.cell(row=8, column=5, value="=E7-D8"); c.number_format = '"$"#,##0'
        c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _PB_BORDER_BOTTOM_SOFT

        # ── R9 Holding → All-In (highlighted running total in money-tint) ──
        _pb_label(ws0, 9, "Holding Costs (= All-In)")
        _pb_input(ws0, 9, 3, pb["holding_pct"], _PB_FILL_P1T, "0.00%")
        c = ws0.cell(row=9, column=4, value="=E6*C9"); c.number_format = '"$"#,##0'
        c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = _PB_BORDER_BOTTOM_SOFT
        c = ws0.cell(row=9, column=5, value="=E8-D9"); c.number_format = '"$"#,##0'
        c.fill = _PB_FILL_P1T; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", bold=True, size=13, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")

        # ── R11 Potential Profit (money-tint if positive, red if negative) ──
        profit_fill = _PB_FILL_MONEY_TINT if pb["potential_profit"] > 0 else _RED_FILL
        profit_color = _PB_MONEY if pb["potential_profit"] > 0 else "9C0006"
        _pb_label(ws0, 11, "Potential Profit")
        c = ws0.cell(row=11, column=5, value="=E9-E16"); c.number_format = '"$"#,##0;[Red]-"$"#,##0'
        c.font = Font(name="Consolas", bold=True, size=13, color=profit_color)
        c.fill = profit_fill; c.border = _PB_BORDER_BOTTOM_SOFT
        c.alignment = Alignment(horizontal="right", vertical="center")

        # ── R12 Profit % (editable target C12 + actual E12) ──
        _pb_label(ws0, 12, "Profit % (FOR DISPO)")
        target_cell = ws0.cell(row=12, column=3, value=DEFAULT_FLIP_TARGET_PROFIT_PCT)
        target_cell.number_format = "0.00%"
        target_cell.font = Font(name="Consolas", bold=True, size=11, color=_PB_MONEY)
        target_cell.fill = _PB_FILL_P1T
        target_cell.alignment = Alignment(horizontal="center", vertical="center")
        target_cell.border = _PB_BORDER_BOTTOM_SOFT
        c = ws0.cell(row=12, column=4, value="← target | actual →")
        c.font = Font(name="Calibri", italic=True, size=9, color=_PB_INK_FAINT)
        c.alignment = Alignment(horizontal="center")
        c = ws0.cell(row=12, column=5, value="=IFERROR(E11/E6,0)"); c.number_format = "0.00%"
        c.font = Font(name="Consolas", bold=True, size=13, color=profit_color)
        c.fill = profit_fill; c.border = _PB_BORDER_BOTTOM_SOFT
        c.alignment = Alignment(horizontal="right", vertical="center")

        # ── R14-R16 Contract block ──
        _pb_label(ws0, 14, "Contract Price to Seller")
        c = ws0.cell(row=14, column=5, value="=E4"); c.number_format = '"$"#,##0'
        c.fill = _PB_FILL_P1T; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        _pb_label(ws0, 15, "Your Wholesale Fee")
        _pb_input(ws0, 15, 5, pb["wholesale_fee"], _PB_FILL_P1T)
        _pb_label(ws0, 16, "Contract Price to Buyer")
        c = ws0.cell(row=16, column=5, value="=E14+E15"); c.number_format = '"$"#,##0'
        c.font = Font(name="Consolas", bold=True, size=12, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = _PB_BORDER_BOTTOM_SOFT

        # ── R18-R22 Buyer Ceiling Lookup (uppercase eyebrow section) ──
        c = ws0.cell(row=18, column=1, value="BUYER CEILING LOOKUP  ·  MAX CONTRACT PRICE AT EACH PROFIT TARGET")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        for col in range(1, 6):
            ws0.cell(row=18, column=col).border = _PB_BORDER_BOTTOM_RULE
        ws0.row_dimensions[18].height = 20
        c = ws0.cell(row=19, column=1, value="Edit ARV / Closing % / Repairs / Holding % / WS Fee above → these recalc live.")
        c.font = Font(name="Georgia", italic=True, size=10, color=_PB_INK_MUTE)

        for offset, (label, target_pct, tint) in enumerate([
            ("At 15% profit target (default / typical disp)", 0.15, _PB_FILL_MONEY_TINT),
            ("At 20% profit target (moderate / hard-money flipper)", 0.20, _PB_FILL_P1T),
            ("At 25% profit target (conservative / heavy rehab)", 0.25, None),
        ]):
            row = 20 + offset
            _pb_label(ws0, row, label)
            formula = f"=E6*(1-C7-C9-{target_pct})-D8-E15"
            c = ws0.cell(row=row, column=5, value=formula)
            c.number_format = '"$"#,##0'
            c.font = Font(name="Consolas", bold=True, size=12, color=_PB_INK)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _PB_BORDER_BOTTOM_SOFT
            if tint: c.fill = tint

        # ── R25 footnote (small uppercase eyebrow) ──
        c = ws0.cell(row=25, column=1, value="RED FILL (D8) = REHAB INPUT  ·  RED/GOLD-TINT CELLS ARE EDITABLE  ·  RUNNING TOTALS UPDATE LIVE")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_FAINT)

        # ── MAO Summary block (rows 27-31) — quick-glance cross-strategy comparison.
        # C29/C30/C31 are the CANONICAL source-of-truth inputs for target profit %,
        # buyer profit $, and target equity %. The MAO Calculation tab's B8/B17/B25
        # reference these cells, so edits here propagate to both tabs.
        # MAO formulas at E29/E30/E31 reference C29/C30/C31 directly (NOT via MAO
        # Calc tab — that would create a circular ref since MAO Calc B8/B17/B25
        # now reference back here). Other inputs (ARV/closing/holding/rehab/WS fee)
        # still pull from MAO Calc tab.
        c = ws0.cell(row=27, column=1, value="EXIT STRATEGY MAO SUMMARY  ·  EDIT C29/C30/C31 → MAO CALC UPDATES TOO")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        for col in range(1, 6):
            ws0.cell(row=27, column=col).border = _PB_BORDER_BOTTOM_RULE
        ws0.row_dimensions[27].height = 20
        c = ws0.cell(row=28, column=1, value="C29/C30/C31 are the canonical inputs. MAO Calc B13/B25/B36 mirror these cells.")
        c.font = Font(name="Georgia", italic=True, size=10, color=_PB_INK_MUTE)

        # (label, editable_default, input_fmt, input_label, mao_formula, fill)
        # mao_formula references local Cnn (NOT MAO Calc) to avoid circular refs.
        # MAO Calc cell coordinates updated 2026-06-13 after MAO tab beautification:
        #   Flip:      ARV=B10, closing=B11, holding=B12, target=B13(→C29), rehab=B14, WSfee=B15 → MAO=B16
        #   Wholetail: CMV=B22, closing=B23, rehab=B24, buyer=B25(→C30), WSfee=B26 → MAO=B27
        #   Rental:    ARV=B33, closing=B34, holding=B35, equity=B36(→C31), rehab=B37, WSfee=B38 → MAO=B39
        rows_def = [
            (
                "Flip MAO",
                DEFAULT_FLIP_TARGET_PROFIT_PCT, '0.00%', "← target profit %",
                "='MAO Calculation'!B10*(1-'MAO Calculation'!B11-'MAO Calculation'!B12-C29)-'MAO Calculation'!B14-'MAO Calculation'!B15",
                _GREEN_FILL,
            ),
            (
                "Wholetail MAO (cosmetic only)",
                DEFAULT_WHOLETAIL_BUYER_PROFIT, '"$"#,##0', "← buyer profit $",
                "='MAO Calculation'!B22*(1-'MAO Calculation'!B23)-'MAO Calculation'!B24-C30-'MAO Calculation'!B26",
                _YELLOW_FILL,
            ),
            (
                "Rental Hold MAO",
                DEFAULT_TARGET_EQUITY_PCT, '0.00%', "← target equity %",
                "='MAO Calculation'!B33*(1-'MAO Calculation'!B34-'MAO Calculation'!B35-C31)-'MAO Calculation'!B37-'MAO Calculation'!B38",
                None,
            ),
        ]
        for offset, (label, default_val, fmt, hint, mao_formula, fill) in enumerate(rows_def):
            row = 29 + offset
            ws0.cell(row=row, column=1, value=label).font = _LABEL_FONT
            # C — editable input (yellow)
            ci = ws0.cell(row=row, column=3, value=default_val)
            ci.number_format = fmt
            ci.fill = INPUT_FILL
            ci.border = _THIN_BORDER
            ci.font = Font(name="Calibri", bold=True, size=12, color="006100")
            ci.alignment = Alignment(horizontal="center", vertical="center")
            # D — input-purpose hint
            ws0.cell(row=row, column=4, value=hint).font = Font(name="Calibri", italic=True, size=9, color="555555")
            ws0.cell(row=row, column=4).alignment = Alignment(horizontal="center")
            # E — MAO formula (green/yellow result)
            ce = ws0.cell(row=row, column=5, value=mao_formula)
            ce.number_format = '"$"#,##0'
            ce.font = Font(name="Calibri", bold=True, size=12)
            if fill: ce.fill = fill
            ce.border = _THIN_BORDER

    ws0.column_dimensions["A"].width = 50
    for col in ["B", "C", "D", "E"]:
        ws0.column_dimensions[col].width = 16

    # ── Embedded Rehab Level Reference (columns I-L starting at row 1) ──
    # The per-line/subtotal/permits/contingency block was removed per user
    # request — the Profit Calculator's closing% + holding% already absorb
    # those padding costs, so D8 just uses the simple base-rehab number.
    rf = pkg.rehab_full
    subj_sqft = rf.sqft or 1
    bucket = _size_bucket(subj_sqft)
    region_mult = rf.regional_multiplier
    SC = 9  # start column I

    ws0.cell(row=1, column=SC, value=f"Rehab Levels — {subj_sqft:,} sqft (bucket: {bucket}, region ×{region_mult:.2f})").font = _TITLE_FONT
    ws0.cell(row=2, column=SC, value=(
        f"Chosen: Level {rf.tier} — {REHAB_LEVEL_NAMES.get(rf.tier, '?')}  |  D8 uses base rehab only (no permits/contingency)"
    )).font = Font(name="Calibri", bold=True, size=11, color="2F5496")

    # Header row at R5
    for off, h in enumerate(["Rehab Level", "$/sqft", f"× {subj_sqft:,} sqft", "× region mult"]):
        c = ws0.cell(row=5, column=SC + off, value=h)
        c.font = _HEADER_FONT; c.fill = _HEADER_FILL; c.alignment = _HEADER_ALIGN

    # 6 levels at R6-R11
    for offset, level in enumerate(range(1, 7)):
        row = 6 + offset
        per_sqft = PER_SQFT_BY_SIZE[bucket][level]
        base = per_sqft * subj_sqft
        base_w_region = base * region_mult
        is_chosen = (level == rf.tier)
        label_text = f"{level}. {REHAB_LEVEL_NAMES[level]}" + ("  ← CHOSEN (= D8)" if is_chosen else "")
        # CHOSEN row gets red fill + dark-red text (2026-06-10) so it visually pairs
        # with D8 (also red) for at-a-glance review of the active rehab assumption.
        ws0.cell(row=row, column=SC, value=label_text).font = (
            Font(name="Calibri", bold=True, size=11, color="9C0006") if is_chosen else _LABEL_FONT
        )
        c2 = ws0.cell(row=row, column=SC + 1, value=per_sqft); c2.number_format = '"$"#,##0'
        c3 = ws0.cell(row=row, column=SC + 2, value=base); c3.number_format = '"$"#,##0'
        c4 = ws0.cell(row=row, column=SC + 3, value=base_w_region); c4.number_format = '"$"#,##0'
        if is_chosen:
            for off in range(4):
                ws0.cell(row=row, column=SC + off).fill = _RED_FILL
        for off in range(4):
            ws0.cell(row=row, column=SC + off).border = _THIN_BORDER

    # Column widths for embedded block (I-L)
    ws0.column_dimensions["I"].width = 42
    for col in ["J", "K", "L"]:
        ws0.column_dimensions[col].width = 14

    # ── Embedded SOLD ARV COMP SNIPPET (I13:S24) — live from Comp Analysis tab ──
    # 11-row block = header + 10 comp rows. Map Comp Analysis cols B..K + M
    # (drop A=# and L=Used — green-fill on first 5 snippet rows already conveys
    # which comps drove ARV) → 11 cols I..S on Profit Calc.
    # ⚠️ FRAGILE: source-row mapping assumes Comp Analysis methodology block has
    # 9 lines (revised 2026-06-09 added 3 lines for PPSF tier methodology).
    # If you add/remove methodology lines, MUST update COMP_HDR_SRC_ROW.
    c = ws0.cell(row=13, column=SC, value="SOLD ARV COMP SNIPPET  ·  TOP 10 BY SIMILARITY  ·  TOP 5 GREEN = USED FOR ARV")
    c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
    for col in range(SC, SC + 11):
        ws0.cell(row=13, column=col).border = _PB_BORDER_BOTTOM_RULE
    ws0.row_dimensions[13].height = 20
    HYPERLINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
    USED_FILL_SNIPPET = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _comps_for_links = sorted(pkg.comps or [], key=lambda c: -c.similarity_score)[:10]
    # Comp Analysis source rows: header at row 32 (was 29 before adding PPSF tier
    # methodology block on 2026-06-09 — +3 methodology lines pushed everything down).
    COMP_HDR_SRC_ROW = 32
    # Comp Analysis column mapping: skip A (#), include B-K, skip L (Used), include M (Adj Price)
    src_cols = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 13]  # B,C,D,E,F,G,H,I,J,K,M
    for ridx in range(11):  # 1 header + 10 comp rows
        src_row = COMP_HDR_SRC_ROW + ridx
        target_row = 14 + ridx
        is_used_for_arv = (1 <= ridx <= 5)  # snippet rows 15-19 = top 5 by similarity
        for cidx, src_col_idx in enumerate(src_cols):
            src_col = get_column_letter(src_col_idx)
            target_col = SC + cidx
            cell = ws0.cell(row=target_row, column=target_col,
                            value=f"='Comp Analysis'!{src_col}{src_row}")
            cell.border = _THIN_BORDER
            if ridx == 0:  # header row
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGN
            else:
                cell.font = _LABEL_FONT
                if is_used_for_arv:
                    cell.fill = USED_FILL_SNIPPET
                # Format numeric columns. After dropping cols A and L:
                # cidx 0=Address, 1=Distance, 2=Sold Date, 3=Sold Price,
                # 4=Sqft, 5=Bd/Ba, 6=Year, 7=PPSF, 8=Similarity, 9=Source, 10=Adj Price
                if cidx in (3, 10):
                    cell.number_format = '"$"#,##0'
                elif cidx == 7:
                    cell.number_format = '"$"#,##0.00'
                elif cidx == 4:
                    cell.number_format = '#,##0'
        # Wire address hyperlink (column I = SC + 0)
        if ridx > 0:
            comp_idx = ridx - 1
            if comp_idx < len(_comps_for_links):
                comp = _comps_for_links[comp_idx]
                if getattr(comp, "detail_url", None):
                    addr_cell = ws0.cell(row=target_row, column=SC + 0)
                    addr_cell.hyperlink = comp.detail_url
                    addr_cell.font = HYPERLINK_FONT

    # Column widths for the comp snippet (I already wide for rehab labels above)
    ws0.column_dimensions["J"].width = 10  # Distance
    ws0.column_dimensions["K"].width = 11  # Sold Date
    ws0.column_dimensions["L"].width = 12  # Sold Price
    ws0.column_dimensions["M"].width = 8   # Sqft
    ws0.column_dimensions["N"].width = 8   # Bd/Ba
    ws0.column_dimensions["O"].width = 7   # Year
    ws0.column_dimensions["P"].width = 10  # PPSF
    ws0.column_dimensions["Q"].width = 11  # Similarity
    ws0.column_dimensions["R"].width = 7   # Source (Z / R / Z+R)
    ws0.column_dimensions["S"].width = 12  # Adj Price

    # ── AVERAGE row beneath the SOLD ARV COMP SNIPPET (row 25) ──
    # Two averages on the same row: top-5 (= ARV Mid) at L, all-10 at P for visual comparison.
    lbl = ws0.cell(row=25, column=SC, value="AVERAGE Sold price (top 5 used for ARV → green | all 10 → blue)")
    lbl.font = Font(name="Calibri", bold=True, size=11, color="2F5496")
    lbl.fill = _YELLOW_FILL
    lbl.border = _THIN_BORDER
    # Top-5 average — matches ARV Mid
    avg_top5 = ws0.cell(row=25, column=SC + 3, value="=AVERAGE(L15:L19)")  # L = SC+3 = Sold Price col
    avg_top5.number_format = '"$"#,##0'
    avg_top5.font = Font(name="Calibri", bold=True, size=13, color="006100")
    avg_top5.fill = _GREEN_FILL
    avg_top5.border = _THIN_BORDER
    # All-10 average — broader market sense check
    avg_all = ws0.cell(row=25, column=SC + 7, value="=AVERAGE(L15:L24)")  # P = SC+7 = PPSF col, reusing for visual offset
    avg_all.number_format = '"$"#,##0'
    avg_all.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    avg_all.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    avg_all.border = _THIN_BORDER

    # ── PENDING SALES SNIPPET (rows 27-36) — under-contract listings, directional only ──
    # Row 26 intentionally left blank as visual separator from AVERAGE row above.
    # List prices, not sold prices — these show where the market is heading next.
    # Written directly from pkg.pending_comps (no cross-sheet ref to keep this block
    # decoupled from Comp Analysis layout shifts).
    c = ws0.cell(row=27, column=SC, value="PENDING / CONTINGENT LISTINGS  ·  UNDER-CONTRACT  ·  REDFIN MLS DATA")
    c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
    for col in range(SC, SC + 9):
        ws0.cell(row=27, column=col).border = _PB_BORDER_BOTTOM_RULE
    ws0.row_dimensions[27].height = 20
    pending_headers = ["Address", "Distance", "List Price", "Sqft", "Bd/Ba", "Year", "PPSF", "DOM", "Status"]
    for col_off, h in enumerate(pending_headers):
        c = ws0.cell(row=28, column=SC + col_off, value=h)
        c.font = _HEADER_FONT; c.fill = _HEADER_FILL; c.alignment = _HEADER_ALIGN; c.border = _THIN_BORDER
    pendings_local = pkg.pending_comps or []
    pending_data_first = 29
    if pendings_local:
        for i, p in enumerate(pendings_local[:6]):
            tgt = 29 + i
            addr_str = f"{p.get('address','')}, {p.get('city','')} {p.get('zip_code','')}".strip(", ")
            ac = ws0.cell(row=tgt, column=SC, value=addr_str)
            if p.get("detail_url"):
                ac.hyperlink = p["detail_url"]
                ac.font = HYPERLINK_FONT
            else:
                ac.font = _LABEL_FONT
            ac.border = _THIN_BORDER
            c = ws0.cell(row=tgt, column=SC + 1, value=f"{p.get('distance_miles',0):.2f} mi"); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws0.cell(row=tgt, column=SC + 2, value=p.get("list_price", 0)); c.number_format = '"$"#,##0'; c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws0.cell(row=tgt, column=SC + 3, value=p.get("sqft", 0)); c.number_format = '#,##0'; c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws0.cell(row=tgt, column=SC + 4, value=f"{p.get('bedrooms',0)}/{p.get('bathrooms',0)}"); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws0.cell(row=tgt, column=SC + 5, value=p.get("year_built") or "—"); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws0.cell(row=tgt, column=SC + 6, value=p.get("ppsf", 0)); c.number_format = '"$"#,##0.00'; c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws0.cell(row=tgt, column=SC + 7, value=p.get("days_on_market", 0)); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws0.cell(row=tgt, column=SC + 8, value=p.get("status", "Pending")); c.font = _LABEL_FONT; c.border = _THIN_BORDER
        pending_data_last = 29 + len(pendings_local[:6]) - 1
        # AVERAGE list price row — quick read on where pendings are pricing vs the sold ARV
        avg_row = pending_data_last + 1
        lbl = ws0.cell(row=avg_row, column=SC, value="AVERAGE list price (pending/contingent)")
        lbl.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
        lbl.fill = _YELLOW_FILL; lbl.border = _THIN_BORDER
        list_col_letter = get_column_letter(SC + 2)
        ac = ws0.cell(row=avg_row, column=SC + 2,
                      value=f"=AVERAGE({list_col_letter}{pending_data_first}:{list_col_letter}{pending_data_last})")
        ac.number_format = '"$"#,##0'
        ac.font = Font(name="Calibri", bold=True, size=13, color="006100")
        ac.fill = _YELLOW_FILL; ac.border = _THIN_BORDER
    else:
        ws0.cell(row=29, column=SC, value="(No active listings within search radius — try widening with --radius 3.0)").font = _LABEL_FONT
        ws0.merge_cells(start_row=29, end_row=29, start_column=SC, end_column=SC + 8)

    # ── DEAL TYPE RECOMMENDATION (rows 32-38) — Fix-n-Flip vs Rental vs Both ──
    # Statistics-driven verdict. Inputs:
    #   • sold_count + rental_count → market depth signal
    #   • rent_pct vs tier_min_rent_pct → cashflow viability
    #   • ARV vs subject zestimate → equity-acquisition fit for flip
    sold_count = len(pkg.comps or [])
    rental_count = len(pkg.rental_comps or [])
    rb_local = pkg.rental_breakdown or {}
    arv_val = pkg.arv.arv_mid
    monthly_rent = rb_local.get("monthly_rent", 0) or 0
    rent_pct_actual = (monthly_rent / arv_val) if arv_val else 0
    tier_min = rb_local.get("tier_min_rent_pct", 0) or 0
    tier_max = rb_local.get("tier_max_rent_pct", 0) or 0
    market_tier = rb_local.get("market_tier", "B")

    rental_viable = (rental_count >= 3) and (rent_pct_actual >= tier_min * 0.9)
    flip_viable = (sold_count >= 4) and (pkg.flip.net_profit > 0)

    if rental_viable and flip_viable:
        verdict = "BOTH — flexible market"
        verdict_color = "006100"
        verdict_fill = _GREEN_FILL
        rationale = (
            f"Strong on both ends: {sold_count} recent sold comps prove a healthy retail-buyer demand "
            f"AND the area carries {rental_count} active rentals at a rent-to-ARV of "
            f"{rent_pct_actual*100:.2f}% (Market Tier {market_tier} target: {tier_min*100:.1f}%-{tier_max*100:.1f}%). "
            f"Run the flip math first (best margin), but a hold-for-cashflow exit is a real Plan B."
        )
    elif flip_viable and not rental_viable:
        verdict = "FIX-N-FLIP HEAVY"
        verdict_color = "9C0006"
        verdict_fill = _YELLOW_FILL
        rationale = (
            f"Sales-driven market: {sold_count} recent sold comps support a retail-exit flip, but the "
            f"rent-to-ARV ({rent_pct_actual*100:.2f}%) underperforms the Tier {market_tier} cashflow floor "
            f"({tier_min*100:.1f}%). Only {rental_count} active rentals nearby — thin renter demand. "
            f"Plan A = flip. Don't underwrite a long-term hold here without a thicker rental comp set."
        )
    elif rental_viable and not flip_viable:
        verdict = "RENTAL HEAVY"
        verdict_color = "2F5496"
        verdict_fill = _YELLOW_FILL
        rationale = (
            f"Rental-driven market: rent-to-ARV {rent_pct_actual*100:.2f}% clears the Tier {market_tier} "
            f"cashflow floor ({tier_min*100:.1f}%) and there are {rental_count} active rentals nearby. "
            f"Flip math is marginal ({sold_count} sold comps; net profit ${pkg.flip.net_profit:,.0f}). "
            f"Plan A = hold for cashflow / BRRRR. Wholesaling to a landlord buyer also fits."
        )
    else:
        verdict = "THIN — needs broader radius or pass"
        verdict_color = "9C0006"
        verdict_fill = _RED_FILL
        rationale = (
            f"Market signal weak on both axes: {sold_count} sold comps, {rental_count} rentals, "
            f"rent-to-ARV {rent_pct_actual*100:.2f}% vs Tier {market_tier} floor {tier_min*100:.1f}%. "
            f"Re-run with --radius 3.0 or --months 18 to widen the comp pool before underwriting, "
            f"OR pass on the deal if the comp pool truly is this sparse in the area."
        )

    # ── DEAL TYPE RECOMMENDATION (rows 33-39) — moved to column A under MAO Summary ──
    c = ws0.cell(row=33, column=1, value="DEAL TYPE RECOMMENDATION  ·  FIX-N-FLIP  ·  RENTAL  ·  BOTH")
    c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
    for col in range(1, 8):
        ws0.cell(row=33, column=col).border = _PB_BORDER_BOTTOM_RULE
    ws0.row_dimensions[33].height = 20
    vcell = ws0.cell(row=34, column=1, value=f"VERDICT: {verdict}")
    vcell.font = Font(name="Georgia", bold=True, size=16, color=verdict_color)
    vcell.fill = verdict_fill
    ws0.cell(row=35, column=1, value=(
        f"Sold comps: {sold_count}  ·  Rental comps: {rental_count}  ·  Pending: {len(pendings_local)}  ·  "
        f"Rent-to-ARV: {rent_pct_actual*100:.2f}%  ·  Tier {market_tier} target: "
        f"{tier_min*100:.1f}%-{tier_max*100:.1f}%"
    )).font = Font(name="Calibri", bold=True, size=11, color="2F5496")
    # Rationale paragraph — wrap across rows 36-39 in column A area (A:G)
    rcell = ws0.cell(row=36, column=1, value=rationale)
    rcell.font = _LABEL_FONT
    rcell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    ws0.merge_cells(start_row=36, end_row=39, start_column=1, end_column=7)
    for rr in range(36, 40):
        ws0.row_dimensions[rr].height = 30

    # ── Compact RENTAL SUMMARY (rows 41-43) — in column A, under Deal Type verdict ──
    # Just the headline rent + rent% from Rental Calculator tab. Full comp table
    # moved to the right-side I area at row 37 (under the pending block).
    if pkg.rental_breakdown:
        c = ws0.cell(row=41, column=1, value="RENTAL SNIPPET  ·  MONTHLY RENT + RENT %  ·  LIVE FROM RENTAL CALCULATOR")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        for col in range(1, 8):
            ws0.cell(row=41, column=col).border = _PB_BORDER_BOTTOM_RULE
        ws0.row_dimensions[41].height = 20

        # Row 42 — Monthly Rent (label A:C merged, value at E to align with deal-math layout)
        lbl = ws0.cell(row=42, column=1, value="='Rental Calculator'!A18")
        lbl.font = _LABEL_FONT; lbl.border = _THIN_BORDER
        lbl.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        ws0.merge_cells(start_row=42, end_row=42, start_column=1, end_column=4)
        v = ws0.cell(row=42, column=5, value="='Rental Calculator'!E18")
        v.number_format = '"$"#,##0'
        v.font = Font(name="Calibri", bold=True, size=12)
        v.border = _THIN_BORDER

        # Row 43 — Rent % (label A:C merged, value at E)
        lbl = ws0.cell(row=43, column=1, value="='Rental Calculator'!A19")
        lbl.font = _LABEL_FONT; lbl.border = _THIN_BORDER
        lbl.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        ws0.merge_cells(start_row=43, end_row=43, start_column=1, end_column=4)
        v = ws0.cell(row=43, column=5, value="='Rental Calculator'!E19")
        v.number_format = '0.00%'
        v.font = Font(name="Calibri", bold=True, size=12)
        v.border = _THIN_BORDER

        # ── RENTAL COMPS table in right-side I area (rows 37-43) ──
        # Pending block ends at row ~35 (title 27 + header 28 + 6 rows 29-34 + avg 35).
        # Row 36 left blank as visual separator. Rental comps table header at I37.
        # R37 — comp table header (Rental Calc R23, dropping the # col): map B:F → I:M
        for cidx in range(5):
            src_col = get_column_letter(2 + cidx)  # B..F
            cell = ws0.cell(row=37, column=SC + cidx, value=f"='Rental Calculator'!{src_col}23")
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        # R38-R43 — 6 rental comp rows (Rental Calc R24-R29), dropping # col
        rcs = pkg.rental_comps or []
        for comp_ridx in range(6):
            src_row = 24 + comp_ridx
            target_row = 38 + comp_ridx
            for cidx in range(5):
                src_col = get_column_letter(2 + cidx)  # B..F
                cell = ws0.cell(row=target_row, column=SC + cidx,
                                value=f"='Rental Calculator'!{src_col}{src_row}")
                cell.font = _LABEL_FONT
                cell.border = _THIN_BORDER
                if cidx == 1:    # Rent/mo
                    cell.number_format = '"$"#,##0'
                elif cidx == 2:  # Sqft
                    cell.number_format = '#,##0'
            # Hyperlink address (col I = SC + 0)
            if comp_ridx < len(rcs) and rcs[comp_ridx].get("detail_url"):
                a_cell = ws0.cell(row=target_row, column=SC)
                a_cell.hyperlink = rcs[comp_ridx]["detail_url"]
                a_cell.font = HYPERLINK_FONT

    # ── Tab 2: Wholetail Calculator — beautified 2026-06-13 with P2 gold aesthetic ──
    # Reuses playbook design helpers (_pb_hero, _pb_section_header, _pb_input, etc.)
    # New cell coordinates (was different — old B5=CMV, D7=closing$ etc.):
    #   Header rows 1-4 (hero) · Section 6-8 (P2 header) · Column labels R11
    #   R12 CMV=B12 · R13 Closing pct=C13 → $=D13 · R14 Repairs=B14
    #   R15 Buyer Profit=B15 · R16 WS Fee=B16 · R17 MAO (money-green)
    #   R19 At-purchase price · R20 Your Profit · R22+ Buyer Ceiling Lookup
    if pkg.wholetail_breakdown:
        wb_ = pkg.wholetail_breakdown
        ws_wt = wb.create_sheet("Wholetail Calculator")
        _pb_hero(ws_wt, pkg.subject, "WHOLETAIL CALCULATOR  ·  LIGHT COSMETIC RESALE", "Wholetail")
        _pb_section_header(
            ws_wt, 6, "P2 · LIGHT COSMETIC",
            "Retail-listed, cosmetic-only rehab",
            "Buy → light cosmetic touch-up → resell to retail buyer as-is. Buyer profit is downstream flipper's margin.",
            _PB_P2,
        )
        _pb_col_headers(ws_wt, 11, [(3, "%"), (4, "AMOUNT"), (5, "RESULT")])

        # R12 — CMV (INPUT, big)
        _pb_label(ws_wt, 12, "CMV / ARV")
        _pb_input(ws_wt, 12, 5, wb_["cmv"], _PB_FILL_P2T)

        # R13 — Closing Costs (INPUT % + FORMULA $)
        _pb_label(ws_wt, 13, "Closing Costs")
        _pb_input(ws_wt, 13, 3, wb_["closing_pct"], _PB_FILL_P2T, "0.00%")
        c = ws_wt.cell(row=13, column=4, value="=E12*C13")
        c.number_format = '"$"#,##0'
        c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill = _PB_FILL_P2T

        # R14 — Repairs (INPUT) — red fill for at-a-glance rehab review (matches Profit Calc D8 convention)
        _pb_label(ws_wt, 14, "Repairs (wholetail-grade cosmetic)")
        c = ws_wt.cell(row=14, column=4, value=wb_["repairs"])
        c.number_format = '"$"#,##0'
        c.fill = _RED_FILL
        c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")

        # R15 — Buyer Profit (INPUT)
        _pb_label(ws_wt, 15, "Buyer Profit (what downstream flipper makes)")
        _pb_input(ws_wt, 15, 4, wb_["buyer_profit"], _PB_FILL_P2T)

        # R16 — WS Fee (INPUT)
        _pb_label(ws_wt, 16, "Your Wholesale Fee")
        _pb_input(ws_wt, 16, 4, wb_["wholesale_fee"], _PB_FILL_P2T)

        # R17 — MAO (money-green headline)
        _pb_mao_row(ws_wt, 17, "=E12-D13-D14-D15-D16", big=True, value_col=5)

        # R19 — At-purchase comparison (INPUT)
        _pb_label(ws_wt, 19, "At purchase price")
        _pb_input(ws_wt, 19, 5, wb_["purchase_price"] or 0, _PB_FILL_P2T)

        # R20 — Your Profit (FORMULA — green/red conditional)
        _pb_label(ws_wt, 20, "Your Profit if bought at this price")
        c = ws_wt.cell(row=20, column=5, value="=E17-E19")
        c.number_format = '"$"#,##0;[Red]-"$"#,##0'
        c.font = Font(name="Consolas", bold=True, size=14, color=_PB_MONEY if (wb_.get("your_profit") or 0) >= 0 else "9C0006")
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _PB_BORDER_BOTTOM_SOFT
        c.fill = _PB_FILL_MONEY_TINT if (wb_.get("your_profit") or 0) >= 0 else _RED_FILL

        # R22+ — Buyer Ceiling Lookup (uppercase section eyebrow + 3 scenario rows)
        c = ws_wt.cell(row=22, column=1, value="BUYER CEILING LOOKUP  ·  MAX CONTRACT AT DIFFERENT BUYER-PROFIT ASSUMPTIONS")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws_wt.cell(row=22, column=col).border = _PB_BORDER_BOTTOM_RULE
        ws_wt.row_dimensions[22].height = 20

        for offset, (label, buyer_profit, tint) in enumerate([
            ("At $50K buyer profit (lean target)", 50000, _PB_FILL_P2T),
            ("At $70K buyer profit (sheet default)", 70000, _PB_FILL_MONEY_TINT),
            ("At $100K buyer profit (premium markets)", 100000, None),
        ]):
            row = 23 + offset
            _pb_label(ws_wt, row, label)
            formula = f"=E12-D13-D14-{buyer_profit}-D16"
            c = ws_wt.cell(row=row, column=5, value=formula)
            c.number_format = '"$"#,##0'
            c.font = Font(name="Consolas", bold=True, size=12, color=_PB_INK)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _PB_BORDER_BOTTOM_SOFT
            if tint: c.fill = tint

        # Footer
        _pb_footer(ws_wt, 27, "REPAIRS (D14) IS RED  ·  EDIT ANY YELLOW/GOLD CELL TO DIAL SCENARIO")

        ws_wt.column_dimensions["A"].width = 52
        ws_wt.column_dimensions["B"].width = 4
        ws_wt.column_dimensions["C"].width = 12
        ws_wt.column_dimensions["D"].width = 16
        ws_wt.column_dimensions["E"].width = 22

    # ── Tab 3: Novation Calculator (2026-06-12) ───────────────────────
    # Novation = control listing without title. We never close on the purchase;
    # the seller's deed transfers directly to the end buyer. We collect the
    # spread between retail sale and the seller's contracted price.
    # KEY MATH DIFFERENCES from Wholetail:
    #   - NO purchase closing costs (never take title)
    #   - Selling % is the FULL retail close (buyer comm + listing comm + title + transfer)
    #   - "Target Spread %" replaces Wholetail's "Buyer Profit $" — it's OUR margin
    #   - "Seller Bonus" optional — extra above seller's net to incentivize agreement
    # Novation Calculator — beautified 2026-06-13 with P4 teal aesthetic ─────
    # Distinct from Rental's P3 blue-gray so operators visually distinguish
    # the two hold-adjacent strategies at a glance in the tab bar.
    # New cell coordinates:
    #   Hero rows 1-4 · Section 6-8 (P4 header) · Col labels R11
    #   R12 Retail ARV=E12 · R13 Selling %=C13 → $=D13
    #   R14 Light Rehab=D14 (RED) · R15 Holding=D15 · R16 Target Spread=C16 → $=D16
    #   R17 Seller Bonus=D17 · R18 MAO (money-green, big)
    #   R20+ Spread Lookup · R25+ Operator notes
    if pkg.novation_breakdown:
        nb = pkg.novation_breakdown
        ws_n = wb.create_sheet("Novation Calculator")
        _pb_hero(ws_n, pkg.subject, "NOVATION CALCULATOR  ·  RETAIL-LISTED, NO TITLE", "Novation")
        _pb_section_header(
            ws_n, 6, "P4 · RETAIL SPREAD",
            "Control listing, capture spread",
            "Seller signs novation agreement; you list under their name; deed goes seller → end buyer. You keep the spread minus costs.",
            _PB_P4,
        )
        _pb_col_headers(ws_n, 11, [(3, "%"), (4, "AMOUNT"), (5, "RESULT")])

        # R12 — Retail ARV
        _pb_label(ws_n, 12, "Retail ARV (MLS-listed price)")
        _pb_input(ws_n, 12, 5, nb["retail_arv"], _PB_FILL_P4T)

        # R13 — Selling %
        _pb_label(ws_n, 13, "Selling Costs (buyer comm + listing comm + title + transfer)")
        _pb_input(ws_n, 13, 3, nb["selling_pct"], _PB_FILL_P4T, "0.00%")
        c = ws_n.cell(row=13, column=4, value="=E12*C13")
        c.number_format = '"$"#,##0'; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill = _PB_FILL_P4T

        # R14 — Light Rehab (RED input — same convention as Profit Calc D8)
        _pb_label(ws_n, 14, "Light Rehab (cosmetic only — Level 1 wholetail-grade)")
        c = ws_n.cell(row=14, column=4, value=nb["light_rehab"])
        c.number_format = '"$"#,##0'; c.fill = _RED_FILL; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")

        # R15 — Holding (flat $, no mortgage)
        _pb_label(ws_n, 15, "Holding (utilities, insurance, lawn — ~60d listing period)")
        _pb_input(ws_n, 15, 4, nb["holding"], _PB_FILL_P4T)

        # R16 — Target Spread % + computed $
        _pb_label(ws_n, 16, "Target Spread % (YOUR profit margin on retail sale)")
        _pb_input(ws_n, 16, 3, nb["target_spread_pct"], _PB_FILL_P4T, "0.00%")
        c = ws_n.cell(row=16, column=4, value="=E12*C16")
        c.number_format = '"$"#,##0'; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill = _PB_FILL_P4T

        # R17 — Seller Bonus
        _pb_label(ws_n, 17, "Seller Bonus (extra above seller's net — incentivize agreement)")
        _pb_input(ws_n, 17, 4, nb["seller_bonus"], _PB_FILL_P4T)

        # R18 — MAO to Seller (headline)
        _pb_mao_row(ws_n, 18, "=E12-D13-D14-D15-D16-D17", big=True, value_col=5)

        # R20 — SPREAD LOOKUP eyebrow
        c = ws_n.cell(row=20, column=1, value="SPREAD LOOKUP  ·  MAX CONTRACT AT DIFFERENT TARGET-SPREAD ASSUMPTIONS")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws_n.cell(row=20, column=col).border = _PB_BORDER_BOTTOM_RULE
        ws_n.row_dimensions[20].height = 20

        for offset, (label, spread_pct, tint) in enumerate([
            ("At 15% target spread (lean — competitive market)", 0.15, _PB_FILL_P4T),
            ("At 20% target spread (default — balanced)", 0.20, _PB_FILL_MONEY_TINT),
            ("At 25% target spread (conservative — risky market)", 0.25, None),
        ]):
            row = 21 + offset
            _pb_label(ws_n, row, label)
            formula = f"=E12-D13-D14-D15-(E12*{spread_pct})-D17"
            c = ws_n.cell(row=row, column=5, value=formula)
            c.number_format = '"$"#,##0'
            c.font = Font(name="Consolas", bold=True, size=12, color=_PB_INK)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _PB_BORDER_BOTTOM_SOFT
            if tint: c.fill = tint

        # R25 — Operator notes eyebrow
        c = ws_n.cell(row=25, column=1, value="HOW NOVATION DIFFERS FROM WHOLETAIL")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 6):
            ws_n.cell(row=25, column=col).border = _PB_BORDER_BOTTOM_RULE
        ws_n.row_dimensions[25].height = 20

        notes = [
            "• You NEVER take title — no purchase closing costs (would be ~2-3% in Wholetail)",
            "• Seller signs a novation agreement allowing YOU to list under their name",
            "• Deed transfers directly seller → end buyer at retail closing",
            "• Light cosmetic rehab only — paint, flooring, curb appeal, staging (not full renovation)",
            "• Your profit = retail sale price − seller's contracted price − everything in this calc",
            "• REQUIRES: attorney-reviewed novation contract + seller cooperation throughout listing",
            "• HIGHER RISK than wholesale (longer time to close), LOWER RISK than flip (no title)",
        ]
        for i, note in enumerate(notes):
            c = ws_n.cell(row=26 + i, column=1, value=note)
            c.font = Font(name="Georgia", size=11, color=_PB_INK)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws_n.merge_cells(start_row=26 + i, end_row=26 + i, start_column=1, end_column=5)
            ws_n.row_dimensions[26 + i].height = 18

        _pb_footer(ws_n, 34, "REHAB (D14) IS RED  ·  EDIT ANY TEAL CELL TO DIAL SCENARIO")

        ws_n.column_dimensions["A"].width = 60
        ws_n.column_dimensions["B"].width = 4
        ws_n.column_dimensions["C"].width = 12
        ws_n.column_dimensions["D"].width = 16
        ws_n.column_dimensions["E"].width = 22

    # ── Tab 5: Rental Calculator — beautified 2026-06-13 with P3 blue-gray aesthetic ──
    # IN-PLACE restyle: all cell coordinates preserved (E4 headline, E6 ARV, C7 closing%,
    # C9 holding%, D8 repairs, C12 target equity%, E15 WS fee, E18 rent, E19 rent%,
    # A21/A22 comp section headers, B23:F29 comp table). The Profit Calc rental snippet
    # cross-refs these exact cells — moving them would break the snippet.
    if pkg.rental_breakdown:
        rb = pkg.rental_breakdown
        ws_r = wb.create_sheet("Rental Calculator")

        # ── R1-3 Editorial hero (Georgia serif, uppercase eyebrow) ──
        c = ws_r.cell(row=1, column=1, value="RENTAL CALCULATOR  ·  HOLD FOR CASHFLOW")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        ws_r.row_dimensions[1].height = 20
        c = ws_r.cell(row=2, column=1, value="Rental Hold")
        c.font = Font(name="Georgia", bold=True, size=28, color=_PB_INK)
        ws_r.merge_cells(start_row=2, end_row=2, start_column=1, end_column=6)
        ws_r.row_dimensions[2].height = 40
        subj_line = f"{pkg.subject.address}, {pkg.subject.city}, {pkg.subject.state} {pkg.subject.zip_code}  ·  Market tier {rb['market_tier']}  ·  Rent % target {rb['tier_min_rent_pct']*100:.1f}%-{rb['tier_max_rent_pct']*100:.1f}%"
        c = ws_r.cell(row=3, column=1, value=subj_line)
        c.font = Font(name="Georgia", italic=True, size=12, color=_PB_INK_MUTE)
        ws_r.merge_cells(start_row=3, end_row=3, start_column=1, end_column=6)
        ws_r.row_dimensions[3].height = 22

        # ── R4 HEADLINE — MAO in money-tint, mono value ──
        c = ws_r.cell(row=4, column=1, value="  RECOMMENDED MAX OFFER  ·  AT TARGET EQUITY %")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_MONEY)
        c.fill = _PB_FILL_MONEY_TINT
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 5):
            ws_r.cell(row=4, column=col).fill = _PB_FILL_MONEY_TINT
        headline = ws_r.cell(row=4, column=5, value="=E6*(1-C7-C9-C12)-D8-E15")
        headline.number_format = '"$"#,##0'
        headline.font = Font(name="Consolas", bold=True, size=18, color=_PB_MONEY)
        headline.fill = _PB_FILL_MONEY_TINT
        headline.alignment = Alignment(horizontal="right", vertical="center")
        ws_r.row_dimensions[4].height = 32

        # ── R5 column labels (uppercase small) ──
        _pb_col_headers(ws_r, 5, [(3, "%"), (4, "AMOUNT"), (5, "RUNNING")])

        # ── R6 ARV / CMV (P3 tint) ──
        _pb_label(ws_r, 6, "ARV / CMV")
        c = ws_r.cell(row=6, column=5, value=rb["arv"]); c.number_format = '"$"#,##0'
        c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
        c.fill = _PB_FILL_P3T; c.border = _PB_BORDER_BOTTOM_SOFT
        c.alignment = Alignment(horizontal="right", vertical="center")

        # ── R7 Closing Costs (P3 tint) ──
        _pb_label(ws_r, 7, "Less Closing Costs")
        _pb_input(ws_r, 7, 3, rb["closing_pct"], _PB_FILL_P3T, "0.00%")
        for col_num, formula in [(4, "=E6*C7"), (5, "=E6-D7")]:
            c = ws_r.cell(row=7, column=col_num, value=formula)
            c.number_format = '"$"#,##0'
            c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _PB_BORDER_BOTTOM_SOFT

        # ── R8 Repairs (RED fill — matches Profit Calc D8) ──
        _pb_label(ws_r, 8, "Less Repairs")
        c = ws_r.cell(row=8, column=4, value=rb["repairs"]); c.number_format = '"$"#,##0'
        c.fill = _RED_FILL; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c = ws_r.cell(row=8, column=5, value="=E7-D8"); c.number_format = '"$"#,##0'
        c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.border = _PB_BORDER_BOTTOM_SOFT

        # ── R9 Holding → All-In (highlighted running total) ──
        _pb_label(ws_r, 9, "Holding Costs (= All-In)")
        _pb_input(ws_r, 9, 3, rb["holding_pct"], _PB_FILL_P3T, "0.00%")
        c = ws_r.cell(row=9, column=4, value="=E6*C9"); c.number_format = '"$"#,##0'
        c.font = Font(name="Consolas", size=11, color=_PB_INK_FAINT, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = _PB_BORDER_BOTTOM_SOFT
        c = ws_r.cell(row=9, column=5, value="=E8-D9"); c.number_format = '"$"#,##0'
        c.fill = _PB_FILL_P3T; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", bold=True, size=13, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")

        # ── R11-R12 Equity block (money-tint if in target, red if not) ──
        eq_fill = _PB_FILL_MONEY_TINT if rb["equity_ok"] else _RED_FILL
        eq_color = _PB_MONEY if rb["equity_ok"] else "9C0006"
        _pb_label(ws_r, 11, "Equity (All-In minus Contract+WS)")
        c = ws_r.cell(row=11, column=5, value="=E9-E16"); c.number_format = '"$"#,##0;[Red]-"$"#,##0'
        c.font = Font(name="Consolas", bold=True, size=13, color=eq_color)
        c.fill = eq_fill; c.border = _PB_BORDER_BOTTOM_SOFT
        c.alignment = Alignment(horizontal="right", vertical="center")

        _pb_label(ws_r, 12, "Equity Percentage")
        target_eq = ws_r.cell(row=12, column=3, value=DEFAULT_TARGET_EQUITY_PCT)
        target_eq.number_format = "0.00%"
        target_eq.font = Font(name="Consolas", bold=True, size=11, color=_PB_MONEY)
        target_eq.fill = _PB_FILL_P3T; target_eq.border = _PB_BORDER_BOTTOM_SOFT
        target_eq.alignment = Alignment(horizontal="center", vertical="center")
        c = ws_r.cell(row=12, column=4, value="← target | actual →")
        c.font = Font(name="Calibri", italic=True, size=9, color=_PB_INK_FAINT)
        c.alignment = Alignment(horizontal="center")
        c = ws_r.cell(row=12, column=5, value="=IFERROR(E11/E6,0)"); c.number_format = "0.00%"
        c.font = Font(name="Consolas", bold=True, size=13, color=eq_color)
        c.fill = eq_fill; c.border = _PB_BORDER_BOTTOM_SOFT
        c.alignment = Alignment(horizontal="right", vertical="center")

        # ── R14-R16 Contract block ──
        _pb_label(ws_r, 14, "Contract Price to Seller")
        c = ws_r.cell(row=14, column=5, value="=E4"); c.number_format = '"$"#,##0'
        c.fill = _PB_FILL_P3T; c.border = _PB_BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", bold=True, size=11, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        _pb_label(ws_r, 15, "Wholesale Fee")
        _pb_input(ws_r, 15, 5, rb["wholesale_fee"], _PB_FILL_P3T)
        _pb_label(ws_r, 16, "Contract Price to Buyer")
        c = ws_r.cell(row=16, column=5, value="=E14+E15"); c.number_format = '"$"#,##0'
        c.font = Font(name="Consolas", bold=True, size=12, color=_PB_INK)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = _PB_BORDER_BOTTOM_SOFT

        # ── R18 Monthly Rent + R19 Rent % (cross-sheet refs land here) ──
        _pb_label(ws_r, 18, "Monthly Rent (Zillow est — validate via comps below)")
        _pb_input(ws_r, 18, 5, rb["monthly_rent"] or 0, _PB_FILL_P3T)
        rent_fill = _PB_FILL_MONEY_TINT if rb.get("rent_pct_ok") else _RED_FILL
        rent_color = _PB_MONEY if rb.get("rent_pct_ok") else "9C0006"
        _pb_label(ws_r, 19, f"Rent % of All-In (Market {rb['market_tier']} target {rb['tier_min_rent_pct']*100:.1f}%+)")
        c = ws_r.cell(row=19, column=5, value="=IFERROR(E18/E9,0)"); c.number_format = "0.00%"
        c.font = Font(name="Consolas", bold=True, size=13, color=rent_color)
        c.fill = rent_fill; c.border = _PB_BORDER_BOTTOM_SOFT
        c.alignment = Alignment(horizontal="right", vertical="center")

        # ── R21-R22 Rental comps section eyebrow + subhead ──
        # NOTE: A21/A22 are referenced by Profit Calc rental snippet — preserve literal text.
        c = ws_r.cell(row=21, column=1, value="── NEARBY RENTAL COMPS (validates Monthly Rent input at E18) ──")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 7):
            ws_r.cell(row=21, column=col).border = _PB_BORDER_BOTTOM_RULE
        ws_r.row_dimensions[21].height = 20
        c = ws_r.cell(row=22, column=1, value="Active FOR_RENT listings within 1 mi — click address for pics + listing detail")
        c.font = Font(name="Georgia", italic=True, size=10, color=_PB_INK_MUTE)

        # ── R23 comp table header (B23:F23 also referenced by Profit Calc snippet) ──
        rc_hdr_row = 23
        for off, h in enumerate(["#", "Address (click for Zillow)", "Rent/mo", "Sqft", "Bd/Ba", "Distance"]):
            c = ws_r.cell(row=rc_hdr_row, column=1 + off, value=h)
            c.font = Font(name="Calibri", bold=True, size=8, color=_PB_INK_FAINT)
            c.alignment = Alignment(horizontal="right" if off > 1 else "left", vertical="bottom", indent=1)
            c.border = _PB_BORDER_BOTTOM_RULE

        HYPERLINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
        rental_comps = pkg.rental_comps or []
        for i, rc in enumerate(rental_comps[:6], 1):
            rc_row = rc_hdr_row + i
            c = ws_r.cell(row=rc_row, column=1, value=i)
            c.font = Font(name="Calibri", size=10, color=_PB_INK_FAINT); c.alignment = Alignment(horizontal="left", indent=1)
            addr_cell = ws_r.cell(row=rc_row, column=2, value=rc.get("address", ""))
            if rc.get("detail_url"):
                addr_cell.hyperlink = rc["detail_url"]
                addr_cell.font = HYPERLINK_FONT
            else:
                addr_cell.font = Font(name="Calibri", size=11, color=_PB_INK)
            c = ws_r.cell(row=rc_row, column=3, value=rc.get("rent", 0)); c.number_format = '"$"#,##0'
            c.font = Font(name="Consolas", size=11, color=_PB_INK); c.alignment = Alignment(horizontal="right")
            c = ws_r.cell(row=rc_row, column=4, value=rc.get("sqft", 0)); c.number_format = '#,##0'
            c.font = Font(name="Consolas", size=11, color=_PB_INK); c.alignment = Alignment(horizontal="right")
            c = ws_r.cell(row=rc_row, column=5, value=f"{rc.get('bedrooms', 0)}/{rc.get('bathrooms', 0)}")
            c.font = Font(name="Consolas", size=11, color=_PB_INK); c.alignment = Alignment(horizontal="right")
            c = ws_r.cell(row=rc_row, column=6, value=f"{rc.get('distance_miles', 0):.2f} mi")
            c.font = Font(name="Consolas", size=11, color=_PB_INK); c.alignment = Alignment(horizontal="right")
            for col in range(1, 7):
                ws_r.cell(row=rc_row, column=col).border = _PB_BORDER_BOTTOM_SOFT
        if not rental_comps:
            c = ws_r.cell(row=rc_hdr_row + 1, column=1, value="(No active rental listings returned from Zillow for this ZIP within 1 mi.)")
            c.font = Font(name="Georgia", italic=True, size=10, color=_PB_INK_MUTE)

        # ── R31+ RENTAL MAO LOOKUP ──
        mao_start_row = rc_hdr_row + 8
        c = ws_r.cell(row=mao_start_row, column=1, value="RENTAL MAO LOOKUP  ·  MAX CONTRACT AT DIFFERENT EQUITY TARGETS")
        c.font = Font(name="Calibri", bold=True, size=9, color=_PB_INK_MUTE)
        for col in range(1, 7):
            ws_r.cell(row=mao_start_row, column=col).border = _PB_BORDER_BOTTOM_RULE
        ws_r.row_dimensions[mao_start_row].height = 20

        for offset, (label, target_eq_val, tint) in enumerate([
            ("At 10% target equity (aggressive)", 0.10, _PB_FILL_P3T),
            ("At 15% target equity (moderate)", 0.15, _PB_FILL_MONEY_TINT),
            ("At 20% target equity (conservative — default)", 0.20, None),
        ]):
            row = mao_start_row + 1 + offset
            _pb_label(ws_r, row, label)
            formula = f"=E6*(1-C7-C9-{target_eq_val})-D8-E15"
            c = ws_r.cell(row=row, column=5, value=formula)
            c.number_format = '"$"#,##0'
            c.font = Font(name="Consolas", bold=True, size=12, color=_PB_INK)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = _PB_BORDER_BOTTOM_SOFT
            if tint: c.fill = tint

        _pb_footer(ws_r, mao_start_row + 5, "REPAIRS (D8) IS RED  ·  EQUITY/RENT % TURN GREEN WHEN IN TARGET RANGE")

        ws_r.column_dimensions["A"].width = 52
        ws_r.column_dimensions["B"].width = 42
        for col in ["C", "D", "E", "F"]:
            ws_r.column_dimensions[col].width = 16

    # ── Tab 4: Comp Analysis (ChatARV-style ARV derivation) ──────
    ws_c = wb.create_sheet("Comp Analysis")
    ws_c.cell(row=1, column=1, value="Comp Analysis — How the ARV was derived").font = _TITLE_FONT
    ws_c.cell(row=2, column=1, value=addr).font = _SUBTITLE_FONT
    s = pkg.subject
    a = pkg.arv
    ws_c.cell(row=3, column=1,
        value=f"Subject: {s.sqft:,} sqft · {s.bedrooms}bd/{s.bathrooms}ba · "
              + (f"built {s.year_built} · " if s.year_built else "")
              + f"Zestimate ${s.zestimate:,.0f} · {len(pkg.comps)} comps used"
    ).font = _LABEL_FONT
    r = 5
    # ARV verdict block
    ws_c.cell(row=r, column=1, value="── ARV CONCLUSION ──").font = Font(
        name="Calibri", bold=True, size=13, color="2F5496")
    r += 1
    arv_rows = [
        ("ARV — Low",  a.arv_low,  _RED_FILL if a.confidence == "low" else None),
        ("ARV — Mid",  a.arv_mid,  _GREEN_FILL),
        ("ARV — High", a.arv_high, None),
        ("Confidence", a.confidence.upper(), _YELLOW_FILL if a.confidence != "high" else _GREEN_FILL),
        ("Why",        a.confidence_reason, None),
        ("Comp spread", f"{a.spread_pct:.1f}%", None),
        ("Avg PPSF",   f"${a.ppsf_avg:.2f}", None),
        ("PPSF range", f"${a.ppsf_range[0]:.2f} — ${a.ppsf_range[1]:.2f}", None),
        ("Bucket A (off-market) / Bucket B (MLS)",
            f"{a.bucket_a_count} / {a.bucket_b_count}", None),
    ]
    for label, val, fill in arv_rows:
        ws_c.cell(row=r, column=1, value=label).font = _LABEL_FONT
        c = ws_c.cell(row=r, column=2, value=val)
        c.font = _VALUE_FONT
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            c.number_format = '"$"#,##0'
        if fill:
            c.fill = fill
        r += 1

    # ── METHODOLOGY — top-5-by-similarity SOLD prices (revised 2026-05-31) ──
    r += 2
    ws_c.cell(row=r, column=1, value="── HOW ARV MID WAS DERIVED (mean of top 5 SOLD prices by similarity) ──").font = Font(
        name="Calibri", bold=True, size=12, color="2F5496")
    r += 1
    # Top-10 by similarity for display; top-5 for ARV
    comps_sorted = sorted(pkg.comps, key=lambda c: -c.similarity_score)[:10]
    top_for_arv = comps_sorted[:5]  # the 5 most-similar comps drive ARV
    used_set = {id(c) for c in top_for_arv}
    avg_used = sum(c.sold_price for c in top_for_arv) / len(top_for_arv) if top_for_arv else 0

    # Source + tier-distribution counts for transparency
    z_only = sum(1 for c in comps_sorted if c.source == "zillow")
    r_only = sum(1 for c in comps_sorted if c.source == "redfin")
    both = sum(1 for c in comps_sorted if c.source == "zillow+redfin")
    renovated_count = sum(1 for c in comps_sorted if c.tier == "renovated")
    distressed_count = sum(1 for c in comps_sorted if c.tier == "distressed")

    method_lines = [
        f"Step 1: Pull SOLD comps from BOTH Zillow + Redfin within 1mi / 6mo. Dedupe by address.",
        f"        → Pool: {len(pkg.comps)} unique sales ({z_only} Zillow-only, {r_only} Redfin-only, {both} cross-confirmed).",
        f"Step 2: Rank by similarity. Priorities: Sold Date > Proximity > PPSF Tier > Sqft > Year Built > Bd/Ba.",
        f"        → PPSF Tier (revised 2026-06-09): renovated comps (PPSF > pool median × 1.25) get +0.15 bonus,",
        f"          distressed (PPSF < median × 0.75) get -0.10 penalty. Year-built ±10yr = no penalty.",
        f"        → Top 10 displayed: {renovated_count} 🔨 renovated, {distressed_count} 💰 distressed, {len(comps_sorted) - renovated_count - distressed_count} standard.",
        f"Step 3: Average the top {len(top_for_arv)} most-similar comps' SOLD prices = ${avg_used:,.0f}  ← ARV Mid",
        f"Step 4: Confidence bands (driven by spread {a.spread_pct:.1f}% within top-5, '{a.confidence}'):"
        f"  ARV Low = ${a.arv_low:,.0f}  |  ARV High = ${a.arv_high:,.0f}",
        f"(Adj Price column shows what each comp would normalize to vs subject — kept for reference, NOT used in ARV.)",
    ]
    for line in method_lines:
        ws_c.cell(row=r, column=1, value=line).font = _LABEL_FONT
        r += 1

    r += 1
    ws_c.cell(row=r, column=1, value="── COMPS USED (sorted by similarity — top 5 USED for ARV are green) ──").font = Font(
        name="Calibri", bold=True, size=13, color="2F5496")
    r += 1
    ws_c.cell(row=r, column=1,
        value="Dual-source pull: Zillow (OpenWeb Ninja) + Redfin (MLS-direct via Webshare proxy). Top 10 ranked by similarity, top 5 used for ARV."
    ).font = _LABEL_FONT
    r += 1
    ws_c.cell(row=r, column=1,
        value="Similarity priority (heaviest → lightest): 1) Sold Date · 2) Proximity · 3) PPSF Tier 🔨/💰 · 4) Sqft · 5) Year Built (±10yr ok) · then beds/baths/type"
    ).font = Font(name="Calibri", italic=True, size=10, color="555555")
    r += 2
    headers = ["#", "Address", "Distance", "Sold Date", "Sold Price",
               "Sqft", "Bd/Ba", "Year", "PPSF", "Similarity", "Source", "Used", "Adj Price"]
    for col, h in enumerate(headers, 1):
        c = ws_c.cell(row=r, column=col, value=h)
        c.font = _HEADER_FONT; c.fill = _HEADER_FILL; c.alignment = _HEADER_ALIGN
    r += 1
    HYPERLINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
    USED_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    REF_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    SOURCE_LABEL = {"zillow": "Z", "redfin": "R", "zillow+redfin": "Z+R"}
    TIER_EMOJI = {"renovated": "🔨", "distressed": "💰", "standard": ""}
    comps_data_first_row = r  # capture for AVERAGE formula below
    for i, comp in enumerate(comps_sorted, 1):
        addr_text = f"{comp.address}, {comp.city} {comp.zip_code}"
        is_used = id(comp) in used_set
        used_label = "✅ ARV" if is_used else "reference"
        row_fill = USED_FILL if is_used else REF_FILL
        src_label = SOURCE_LABEL.get(comp.source, comp.source)
        tier_emoji = TIER_EMOJI.get(comp.tier, "")
        # Show tier emoji alongside similarity % so renovated/distressed comps
        # are visually identifiable without adding a new column.
        sim_display = f"{tier_emoji}{comp.similarity_score:.0%}" if tier_emoji else f"{comp.similarity_score:.0%}"
        row_vals = [
            i, addr_text,
            f"{comp.distance_miles:.2f} mi",
            comp.sold_date or "—",
            comp.sold_price, comp.sqft,
            f"{comp.bedrooms}/{comp.bathrooms}",
            comp.year_built or "—",
            comp.ppsf,
            sim_display,
            src_label,
            used_label,
            comp.adjusted_price,
        ]
        for col, val in enumerate(row_vals, 1):
            c = ws_c.cell(row=r, column=col, value=val)
            c.font = _LABEL_FONT
            c.fill = row_fill
            c.border = _THIN_BORDER
            if col in (5, 13):  # Sold Price + Adj Price → money
                c.number_format = '"$"#,##0'
            elif col == 9:  # PPSF
                c.number_format = '"$"#,##0.00'
            elif col == 6:  # sqft
                c.number_format = '#,##0'
        # Make the Address cell (col 2) a clickable hyperlink to the source listing
        if comp.detail_url:
            addr_cell = ws_c.cell(row=r, column=2)
            addr_cell.hyperlink = comp.detail_url
            addr_cell.font = HYPERLINK_FONT
        r += 1
    comps_data_last_row = r - 1

    # ── AVERAGE Sold Price row beneath COMPS USED ──
    # Top-5 average matches ARV Mid. All-10 average gives broader market sense.
    top5_last = min(comps_data_first_row + 4, comps_data_last_row)  # rows 1-5 of comp data
    lbl = ws_c.cell(row=r, column=2, value="AVERAGE Sold price  (top 5 = ARV  |  all 10 = market check)")
    lbl.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    lbl.fill = _YELLOW_FILL
    lbl.border = _THIN_BORDER
    # Top-5 average → column E (Sold Price col)
    avg5 = ws_c.cell(row=r, column=5,
                     value=f"=AVERAGE(E{comps_data_first_row}:E{top5_last})")
    avg5.number_format = '"$"#,##0'
    avg5.font = Font(name="Calibri", bold=True, size=13, color="006100")
    avg5.fill = _GREEN_FILL
    avg5.border = _THIN_BORDER
    # All-10 average → column I (PPSF col, used as visual offset for the second average)
    avg_all = ws_c.cell(row=r, column=9,
                        value=f"=AVERAGE(E{comps_data_first_row}:E{comps_data_last_row})")
    avg_all.number_format = '"$"#,##0'
    avg_all.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    avg_all.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    avg_all.border = _THIN_BORDER
    r += 1
    r += 2

    # Adjustments detail (one row per comp showing what was applied)
    ws_c.cell(row=r, column=1, value="── ADJUSTMENTS PER COMP (how each comp was normalized to subject) ──").font = Font(
        name="Calibri", bold=True, size=13, color="2F5496")
    r += 1
    ws_c.cell(row=r, column=1, value="Positive = comp added value to reach subject; negative = comp subtracted to reach subject.").font = _LABEL_FONT
    r += 2
    adj_types = ["sqft", "bedrooms", "bathrooms", "year_built", "lot_size", "garage", "market_conditions"]
    adj_headers = ["#", "Address"] + [t.replace("_", " ").title() for t in adj_types] + ["Total Adj", "Sold→Adj"]
    for col, h in enumerate(adj_headers, 1):
        c = ws_c.cell(row=r, column=col, value=h)
        c.font = _HEADER_FONT; c.fill = _HEADER_FILL; c.alignment = _HEADER_ALIGN
    r += 1
    adj_first_row = r
    for i, comp in enumerate(comps_sorted[:10], 1):
        ws_c.cell(row=r, column=1, value=i).border = _THIN_BORDER
        ws_c.cell(row=r, column=2, value=comp.address).font = _LABEL_FONT
        ws_c.cell(row=r, column=2).border = _THIN_BORDER
        total_adj = 0
        for col, atype in enumerate(adj_types, 3):
            v = comp.adjustments.get(atype, 0) or 0
            total_adj += v
            c = ws_c.cell(row=r, column=col, value=v)
            c.number_format = '"$"#,##0;[Red]-"$"#,##0'
            c.border = _THIN_BORDER
        c = ws_c.cell(row=r, column=3 + len(adj_types), value=total_adj)
        c.font = _VALUE_FONT; c.number_format = '"$"#,##0;[Red]-"$"#,##0'; c.border = _THIN_BORDER
        c = ws_c.cell(row=r, column=4 + len(adj_types), value=comp.adjusted_price)
        c.font = _VALUE_FONT; c.number_format = '"$"#,##0'; c.border = _THIN_BORDER
        r += 1

    # Average row — average of all 6 adjusted comp prices (and total adjustments)
    adj_last_row = r - 1
    total_adj_col = 3 + len(adj_types)  # 10 = J
    adj_price_col = 4 + len(adj_types)  # 11 = K
    label_cell = ws_c.cell(row=r, column=2, value="AVERAGE (all displayed adjusted comps)")
    label_cell.font = Font(name="Calibri", bold=True, size=12, color="2F5496")
    label_cell.border = _THIN_BORDER
    label_cell.fill = _YELLOW_FILL
    avg_total_col_letter = get_column_letter(total_adj_col)
    c = ws_c.cell(row=r, column=total_adj_col,
                  value=f"=AVERAGE({avg_total_col_letter}{adj_first_row}:{avg_total_col_letter}{adj_last_row})")
    c.font = Font(name="Calibri", bold=True, size=12)
    c.number_format = '"$"#,##0;[Red]-"$"#,##0'
    c.border = _THIN_BORDER
    c.fill = _YELLOW_FILL
    avg_price_col_letter = get_column_letter(adj_price_col)
    c = ws_c.cell(row=r, column=adj_price_col,
                  value=f"=AVERAGE({avg_price_col_letter}{adj_first_row}:{avg_price_col_letter}{adj_last_row})")
    c.font = Font(name="Calibri", bold=True, size=13, color="006100")
    c.number_format = '"$"#,##0'
    c.border = _THIN_BORDER
    c.fill = _YELLOW_FILL
    r += 1

    # ── Pending Sales block (under-contract listings — directional ARV signal) ──
    r += 2
    pending_section_first_row = r  # captured so the Profit Calc snippet knows where to reference
    ws_c.cell(row=r, column=1, value="── PENDING / CONTINGENT LISTINGS (under-contract — front-runs sold by 30-60d) ──").font = Font(
        name="Calibri", bold=True, size=13, color="2F5496")
    r += 1
    ws_c.cell(row=r, column=1,
        value="Pulled from Redfin's MLS feed via Webshare US residential proxy. List prices only — directional ARV signal, NOT used in ARV calc."
    ).font = _LABEL_FONT
    r += 2
    pending_headers = ["#", "Address", "Distance", "List Price", "Sqft", "Bd/Ba", "Year", "PPSF", "DOM", "Status"]
    pending_header_row = r
    for col, h in enumerate(pending_headers, 1):
        c = ws_c.cell(row=r, column=col, value=h)
        c.font = _HEADER_FONT; c.fill = _HEADER_FILL; c.alignment = _HEADER_ALIGN
    r += 1
    pending_data_first_row = r
    pendings = pkg.pending_comps or []
    if pendings:
        for i, p in enumerate(pendings[:6], 1):
            ws_c.cell(row=r, column=1, value=i).border = _THIN_BORDER
            addr_str = f"{p.get('address','')}, {p.get('city','')} {p.get('zip_code','')}".strip(", ")
            addr_cell = ws_c.cell(row=r, column=2, value=addr_str)
            if p.get("detail_url"):
                addr_cell.hyperlink = p["detail_url"]
                addr_cell.font = HYPERLINK_FONT
            else:
                addr_cell.font = _LABEL_FONT
            addr_cell.border = _THIN_BORDER
            c = ws_c.cell(row=r, column=3, value=f"{p.get('distance_miles',0):.2f} mi"); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws_c.cell(row=r, column=4, value=p.get("list_price", 0)); c.number_format = '"$"#,##0'; c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws_c.cell(row=r, column=5, value=p.get("sqft", 0)); c.number_format = '#,##0'; c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws_c.cell(row=r, column=6, value=f"{p.get('bedrooms',0)}/{p.get('bathrooms',0)}"); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws_c.cell(row=r, column=7, value=p.get("year_built") or "—"); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws_c.cell(row=r, column=8, value=p.get("ppsf", 0)); c.number_format = '"$"#,##0.00'; c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws_c.cell(row=r, column=9, value=p.get("days_on_market", 0)); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            c = ws_c.cell(row=r, column=10, value=p.get("status", "Pending")); c.font = _LABEL_FONT; c.border = _THIN_BORDER
            r += 1
        # Average list price row — helps see if pendings are trending above/below the sold ARV
        avg_row = r
        ws_c.cell(row=r, column=2, value="AVERAGE list price").font = Font(name="Calibri", bold=True, size=12, color="2F5496")
        ws_c.cell(row=r, column=2).fill = _YELLOW_FILL; ws_c.cell(row=r, column=2).border = _THIN_BORDER
        ac = ws_c.cell(row=r, column=4, value=f"=AVERAGE(D{pending_data_first_row}:D{r-1})")
        ac.number_format = '"$"#,##0'; ac.font = Font(name="Calibri", bold=True, size=13, color="006100")
        ac.fill = _YELLOW_FILL; ac.border = _THIN_BORDER
        r += 1
    else:
        ws_c.cell(row=r, column=1, value="(No active listings within search radius — try widening with --radius 3.0)").font = _LABEL_FONT
        r += 1
    # How many rows the Profit Calc snippet should pull (header + up to 6 + avg row OR no-data line)
    pending_section_last_row = r - 1

    # Column widths — now 13 cols on COMPS USED table (added Source + Used)
    ws_c.column_dimensions["A"].width = 5
    ws_c.column_dimensions["B"].width = 40
    for col in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
        ws_c.column_dimensions[col].width = 11
    ws_c.column_dimensions["E"].width = 13  # Sold Price
    ws_c.column_dimensions["K"].width = 7   # Source (Z / R / Z+R)
    ws_c.column_dimensions["L"].width = 10  # Used (ARV / reference)
    ws_c.column_dimensions["M"].width = 13  # Adj Price

    # ── Tab 5: Deal Summary (existing) ────────────────────────────
    ws = wb.create_sheet("Deal Summary")
    ws.cell(row=1, column=1, value="Deal Analysis Report").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=addr).font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _LABEL_FONT

    # ── HEADLINE — Seller Offer at 15% target profit ──
    # The number you tell the seller. Big, green, top of Deal Summary.
    ws.cell(row=5, column=1, value="🎯 SELLER OFFER (MAO at 15% target profit)").font = Font(
        name="Calibri", bold=True, size=14, color="006100")
    ws.cell(row=5, column=1).fill = _GREEN_FILL
    seller_offer = ws.cell(row=5, column=2, value=pkg.mao.flip_mao)
    seller_offer.number_format = '"$"#,##0'
    seller_offer.font = Font(name="Calibri", bold=True, size=16, color="006100")
    seller_offer.fill = _GREEN_FILL
    ws.row_dimensions[5].height = 28

    # Recommendation
    rec_cell = ws.cell(row=7, column=1, value=pkg.recommendation)
    rec_cell.font = Font(name="Calibri", bold=True, size=12)
    if "GO —" in pkg.recommendation and "NO-GO" not in pkg.recommendation:
        rec_cell.fill = _GREEN_FILL
        rec_cell.font = Font(name="Calibri", bold=True, size=12, color="006100")
    elif "NO-GO" in pkg.recommendation:
        rec_cell.fill = _RED_FILL
        rec_cell.font = Font(name="Calibri", bold=True, size=12, color="9C0006")
    else:
        rec_cell.fill = _YELLOW_FILL

    row = 9
    data = [
        ("ARV (Recommended)", _fmt_money(pkg.arv.arv_mid)),
        ("ARV Confidence", pkg.arv.confidence.upper()),
        ("", ""),
        ("MAO — Flip (15% target profit)", _fmt_money(pkg.mao.flip_mao)),
        ("MAO — Wholetail ($70K buyer profit)", _fmt_money(pkg.mao.wholesale_mao)),
        ("MAO — Rental Hold (20% target equity)", _fmt_money(pkg.mao.hold_mao)),
        ("", ""),
        ("Full Rehab Cost", _fmt_money(pkg.rehab_full.grand_total)),
        ("Wholetail Cost", _fmt_money(pkg.rehab_wholetail.grand_total)),
        ("", ""),
        ("Flip Profit", _fmt_money(pkg.flip.net_profit)),
        ("Flip ROI", f"{pkg.flip.roi_pct:.1f}%"),
        ("Flip Timeline", f"{pkg.flip.months_to_complete:.0f} months"),
        ("", ""),
        ("Wholesale Fee", _fmt_money(pkg.wholesale.assignment_fee)),
        ("", ""),
        ("Hold Cash Flow (Annual)", _fmt_money(pkg.hold.cash_flow_annual)),
        ("Hold Cash-on-Cash", f"{pkg.hold.cash_on_cash:.1f}%"),
        ("Hold Cap Rate", f"{pkg.hold.cap_rate:.1f}%"),
    ]
    for label, value in data:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _VALUE_FONT
        if "Profit" in label and pkg.flip.net_profit > 0:
            cell.fill = _GREEN_FILL
            cell.font = _GREEN_FONT
        elif "Profit" in label and pkg.flip.net_profit <= 0:
            cell.fill = _RED_FILL
            cell.font = _RED_FONT
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 25

    # ── Tab 2: ARV Analysis ───────────────────────────────────────
    ws2 = wb.create_sheet("ARV Analysis")
    ws2.cell(row=1, column=1, value="ARV Analysis (from Comp Report)").font = _TITLE_FONT
    arv_data = [
        ("ARV Low", _fmt_money(pkg.arv.arv_low)),
        ("ARV Mid (Recommended)", _fmt_money(pkg.arv.arv_mid)),
        ("ARV High", _fmt_money(pkg.arv.arv_high)),
        ("Confidence", pkg.arv.confidence.upper()),
        ("Reason", pkg.arv.confidence_reason),
        ("Comps Used", str(pkg.arv.comp_count)),
        ("Avg PPSF", f"${pkg.arv.ppsf_avg:,.2f}"),
        ("PPSF Range", f"${pkg.arv.ppsf_range[0]:,.2f} — ${pkg.arv.ppsf_range[1]:,.2f}"),
        ("Avg Adjustment", _fmt_money(pkg.arv.avg_adjustment)),
        ("Spread", f"{pkg.arv.spread_pct:.1f}%"),
    ]
    for i, (label, value) in enumerate(arv_data, 3):
        ws2.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws2.cell(row=i, column=2, value=value).font = _VALUE_FONT
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 30

    # ── Tab 3: Rehab Budget (with per-sqft breakdown + all-level reference) ──
    ws3 = wb.create_sheet("Rehab Budget")
    rf = pkg.rehab_full
    rw = pkg.rehab_wholetail
    subj_sqft = rf.sqft or 1
    bucket = _size_bucket(subj_sqft)
    region_mult = rf.regional_multiplier

    ws3.cell(row=1, column=1, value="Rehab Budget — Per-sqft Breakdown").font = _TITLE_FONT
    ws3.cell(row=2, column=1, value=(
        f"Subject: {subj_sqft:,} sqft  ·  Size bucket: {bucket}  ·  "
        f"Region: {rf.region.title()} (×{region_mult:.2f})"
    )).font = _LABEL_FONT
    ws3.cell(row=3, column=1, value=(
        f"Chosen Rehab Level: {rf.tier} — {REHAB_LEVEL_NAMES.get(rf.tier, '?')}"
    )).font = Font(name="Calibri", bold=True, size=12, color="2F5496")

    # Per-line breakdown with $/sqft column
    headers = ["Category", "Full Rehab $", "Full $/sqft", "Wholetail $", "Wholetail $/sqft"]
    _write_headers(ws3, 5, headers)

    full_dict = {r.category: r.total for r in rf.rooms}
    wt_dict = {r.category: r.total for r in rw.rooms}
    all_cats = list(dict.fromkeys([r.category for r in rf.rooms] +
                                  [r.category for r in rw.rooms]))
    wt_sqft = rw.sqft or subj_sqft
    for i, cat in enumerate(all_cats, 6):
        full_val = full_dict.get(cat, 0)
        wt_val = wt_dict.get(cat, 0)
        ws3.cell(row=i, column=1, value=cat).font = _LABEL_FONT
        c2 = ws3.cell(row=i, column=2, value=full_val); c2.number_format = '"$"#,##0'
        c3 = ws3.cell(row=i, column=3, value=full_val / subj_sqft if full_val else 0); c3.number_format = '"$"#,##0.00'
        c4 = ws3.cell(row=i, column=4, value=wt_val); c4.number_format = '"$"#,##0'
        c5 = ws3.cell(row=i, column=5, value=wt_val / wt_sqft if wt_val else 0); c5.number_format = '"$"#,##0.00'
        for col in range(1, 6):
            ws3.cell(row=i, column=col).border = _THIN_BORDER

    # Subtotal (before permits + contingency, which are calculated separately)
    sub_row = 6 + len(all_cats)
    ws3.cell(row=sub_row, column=1, value="Subtotal (line items)").font = _LABEL_FONT
    sub_full = sum(full_dict.values())
    sub_wt = sum(wt_dict.values())
    for col, val, sf in [(2, sub_full, subj_sqft), (3, sub_full / subj_sqft if sub_full else 0, None),
                          (4, sub_wt, wt_sqft),    (5, sub_wt / wt_sqft if sub_wt else 0, None)]:
        c = ws3.cell(row=sub_row, column=col, value=val)
        c.font = _LABEL_FONT
        c.number_format = '"$"#,##0' if col in (2, 4) else '"$"#,##0.00'
        c.border = _THIN_BORDER

    # Permits + contingency lines (separate — they're % adders, not line items)
    perm_row = sub_row + 1
    ws3.cell(row=perm_row, column=1, value="Permits (3%)").font = _LABEL_FONT
    ws3.cell(row=perm_row, column=2, value=rf.permits_cost).number_format = '"$"#,##0'
    ws3.cell(row=perm_row, column=3, value=rf.permits_cost / subj_sqft).number_format = '"$"#,##0.00'
    ws3.cell(row=perm_row, column=4, value=rw.permits_cost).number_format = '"$"#,##0'
    ws3.cell(row=perm_row, column=5, value=rw.permits_cost / wt_sqft).number_format = '"$"#,##0.00'

    cont_row = perm_row + 1
    cont_label = f"Contingency ({rf.contingency_pct*100:.0f}%)"
    ws3.cell(row=cont_row, column=1, value=cont_label).font = _LABEL_FONT
    ws3.cell(row=cont_row, column=2, value=rf.contingency_cost).number_format = '"$"#,##0'
    ws3.cell(row=cont_row, column=3, value=rf.contingency_cost / subj_sqft).number_format = '"$"#,##0.00'
    ws3.cell(row=cont_row, column=4, value=rw.contingency_cost).number_format = '"$"#,##0'
    ws3.cell(row=cont_row, column=5, value=rw.contingency_cost / wt_sqft).number_format = '"$"#,##0.00'
    for r_ in (perm_row, cont_row):
        for col in range(1, 6):
            ws3.cell(row=r_, column=col).border = _THIN_BORDER

    # Grand total (bold, green highlight)
    tot_row = cont_row + 1
    ws3.cell(row=tot_row, column=1, value="GRAND TOTAL").font = Font(name="Calibri", bold=True, size=13, color="006100")
    tot_cells = [
        (2, rf.grand_total, '"$"#,##0'),
        (3, rf.grand_total / subj_sqft, '"$"#,##0.00'),
        (4, rw.grand_total, '"$"#,##0'),
        (5, rw.grand_total / wt_sqft, '"$"#,##0.00'),
    ]
    for col, val, fmt in tot_cells:
        c = ws3.cell(row=tot_row, column=col, value=val)
        c.number_format = fmt
        c.font = Font(name="Calibri", bold=True, size=13, color="006100")
        c.fill = _GREEN_FILL
        c.border = _THIN_BORDER

    # ── Reference: all 6 Rehab Levels for THIS subject's size + region ──
    ref_row = tot_row + 3
    ws3.cell(row=ref_row, column=1, value=f"── REFERENCE: What other Rehab Levels would cost ({subj_sqft:,} sqft × birmingham bucket '{bucket}') ──").font = Font(
        name="Calibri", bold=True, size=12, color="2F5496")
    ref_row += 1
    ws3.cell(row=ref_row, column=1, value="Level base only (excludes permits + age contingency). Add ~13% for permits+contingency at this property's age.").font = _LABEL_FONT
    ref_row += 2
    _write_headers(ws3, ref_row, ["Rehab Level", "$/sqft", f"× {subj_sqft:,} sqft", "× region mult"])
    ref_row += 1
    for level in range(1, 7):
        per_sqft = PER_SQFT_BY_SIZE[bucket][level]
        base = per_sqft * subj_sqft
        base_w_region = base * region_mult
        is_chosen = (level == rf.tier)
        label_text = f"{level}. {REHAB_LEVEL_NAMES[level]}" + ("  ← CHOSEN" if is_chosen else "")
        ws3.cell(row=ref_row, column=1, value=label_text).font = (
            Font(name="Calibri", bold=True, size=11, color="006100") if is_chosen else _LABEL_FONT
        )
        c2 = ws3.cell(row=ref_row, column=2, value=per_sqft); c2.number_format = '"$"#,##0'
        c3 = ws3.cell(row=ref_row, column=3, value=base); c3.number_format = '"$"#,##0'
        c4 = ws3.cell(row=ref_row, column=4, value=base_w_region); c4.number_format = '"$"#,##0'
        if is_chosen:
            for col in range(1, 5):
                ws3.cell(row=ref_row, column=col).fill = _GREEN_FILL
        for col in range(1, 5):
            ws3.cell(row=ref_row, column=col).border = _THIN_BORDER
        ref_row += 1

    ws3.column_dimensions["A"].width = 50
    for col in ["B", "C", "D", "E"]:
        ws3.column_dimensions[col].width = 16

    # ── Tab 4: MAO Calculation — beautified 2026-06-13 with playbook aesthetic ──
    # Design tokens borrowed from the SiftStack Marketing Playbook Q3 2026:
    #   Ink #1B1D22, Ink-mute #55575E, Ink-faint #8A8B8F
    #   Rule #C0BCB2 (borders), Money #0B6B47 (financial results), Money-tint #E4EFE8
    #   Priority stripes: P1 red #A63232 (Flip), P2 gold #B98A2E (Wholetail),
    #                     P3 blue-gray #4A6D74 (Rental)
    # Typography: Georgia serif for editorial titles; Consolas mono for numbers;
    # Calibri sans for labels + uppercase tags.
    ws4 = wb.create_sheet("MAO Calculation")

    # Playbook color tokens (light-theme values from the artifact)
    _INK       = "1B1D22"
    _INK_MUTE  = "55575E"
    _INK_FAINT = "8A8B8F"
    _RULE      = "C0BCB2"
    _RULE_SOFT = "E1DED4"
    _MONEY     = "0B6B47"
    _MONEY_TINT = "E4EFE8"
    _P1        = "A63232"   # Flip
    _P1_TINT   = "F4E5E2"
    _P2        = "B98A2E"   # Wholetail
    _P2_TINT   = "F5EBD5"
    _P3        = "4A6D74"   # Rental
    _P3_TINT   = "DDE7EA"

    _FILL_MONEY_TINT = PatternFill(start_color=_MONEY_TINT, end_color=_MONEY_TINT, fill_type="solid")
    _FILL_P1  = PatternFill(start_color=_P1, end_color=_P1, fill_type="solid")
    _FILL_P2  = PatternFill(start_color=_P2, end_color=_P2, fill_type="solid")
    _FILL_P3  = PatternFill(start_color=_P3, end_color=_P3, fill_type="solid")
    _FILL_P1T = PatternFill(start_color=_P1_TINT, end_color=_P1_TINT, fill_type="solid")
    _FILL_P2T = PatternFill(start_color=_P2_TINT, end_color=_P2_TINT, fill_type="solid")
    _FILL_P3T = PatternFill(start_color=_P3_TINT, end_color=_P3_TINT, fill_type="solid")
    _FILL_RULE_SOFT = PatternFill(start_color=_RULE_SOFT, end_color=_RULE_SOFT, fill_type="solid")

    _BORDER_BOTTOM_RULE = Border(bottom=Side(style="thin", color=_RULE))
    _BORDER_TOP_RULE = Border(top=Side(style="thin", color=_RULE))
    _BORDER_BOTTOM_SOFT = Border(bottom=Side(style="thin", color=_RULE_SOFT))

    # ── Editorial hero (rows 1-4) ──
    # eyebrow (uppercase letter-spaced) + big serif title + subject subtitle + rule
    c = ws4.cell(row=1, column=1, value="MAO CALCULATION  ·  BY EXIT STRATEGY")
    c.font = Font(name="Calibri", bold=True, size=9, color=_INK_MUTE)
    c.alignment = Alignment(horizontal="left", vertical="bottom")
    ws4.row_dimensions[1].height = 20

    c = ws4.cell(row=2, column=1, value="Maximum Allowable Offer")
    c.font = Font(name="Georgia", bold=True, size=28, color=_INK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws4.merge_cells(start_row=2, end_row=2, start_column=1, end_column=3)
    ws4.row_dimensions[2].height = 40

    subj = pkg.subject
    subj_line = f"{subj.address}, {subj.city}, {subj.state} {subj.zip_code}"
    if subj.sqft: subj_line += f"  ·  {subj.sqft:,} sqft"
    if subj.bedrooms or subj.bathrooms: subj_line += f"  ·  {subj.bedrooms}bd/{subj.bathrooms}ba"
    c = ws4.cell(row=3, column=1, value=subj_line)
    c.font = Font(name="Georgia", italic=True, size=12, color=_INK_MUTE)
    ws4.merge_cells(start_row=3, end_row=3, start_column=1, end_column=3)
    ws4.row_dimensions[3].height = 22

    # Row 4 — rule (thin bottom border on empty row across A:C)
    for col in range(1, 4):
        ws4.cell(row=4, column=col).border = _BORDER_BOTTOM_RULE
    ws4.row_dimensions[4].height = 4

    # ── Helper: input value (light tint per strategy) ──
    def _input(row, col, value, tint_fill, fmt='"$"#,##0'):
        c = ws4.cell(row=row, column=col, value=value)
        c.number_format = fmt
        c.fill = tint_fill
        c.border = _BORDER_BOTTOM_SOFT
        # Consolas mono, tabular-aligned numbers per playbook aesthetic
        c.font = Font(name="Consolas", bold=True, size=11, color=_INK)
        c.alignment = Alignment(horizontal="right", vertical="center")
        return c

    def _derived(row, formula, tint_fill):
        c = ws4.cell(row=row, column=3, value=formula)
        c.number_format = '"$"#,##0'
        c.fill = tint_fill
        c.border = _BORDER_BOTTOM_SOFT
        c.font = Font(name="Consolas", size=11, color=_INK_FAINT, italic=True)
        c.alignment = Alignment(horizontal="right", vertical="center")
        return c

    def _label(row, text):
        c = ws4.cell(row=row, column=1, value=text)
        c.font = Font(name="Calibri", size=11, color=_INK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = _BORDER_BOTTOM_SOFT
        return c

    def _mao_row(row, formula, stripe_fill, big=False):
        """Big money-green MAO result — the headline number for each section."""
        # Column A: "MAO" label with strategy color
        lbl = ws4.cell(row=row, column=1, value="MAO")
        lbl.font = Font(name="Calibri", bold=True, size=11, color=_INK_MUTE)
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lbl.fill = _FILL_MONEY_TINT
        # Column B: the big MAO number in money green
        c = ws4.cell(row=row, column=2, value=formula)
        c.number_format = '"$"#,##0'
        c.fill = _FILL_MONEY_TINT
        c.font = Font(name="Consolas", bold=True, size=18 if big else 15, color=_MONEY)
        c.alignment = Alignment(horizontal="right", vertical="center")
        # Column C: caption (uppercase eyebrow "Contract to seller")
        cap = ws4.cell(row=row, column=3, value="CONTRACT TO SELLER")
        cap.font = Font(name="Calibri", bold=True, size=9, color=_MONEY)
        cap.fill = _FILL_MONEY_TINT
        cap.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws4.row_dimensions[row].height = 36 if big else 30

    def _section_header(row, tag_text, section_title, thesis, stripe_fill, tint_fill, tag_ink):
        """Playbook-style section header: colored strip + tag chip + serif title + thesis."""
        # Left color stripe (1-col wide, spans header row height)
        # Approximate with a colored left border on column A
        # Row structure: 2 rows tall — [tag chip] [serif title / thesis]
        # Row `row`: uppercase tag chip in column A (with tint fill)
        c = ws4.cell(row=row, column=1, value=f"  {tag_text}")
        c.font = Font(name="Calibri", bold=True, size=9, color=tag_ink)
        c.fill = tint_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = Border(left=Side(style="thick", color=stripe_fill.start_color.value if hasattr(stripe_fill.start_color, 'value') else stripe_fill.start_color.rgb[-6:]))
        ws4.row_dimensions[row].height = 20

        # Row row+1: big serif section title
        c = ws4.cell(row=row + 1, column=1, value=section_title)
        c.font = Font(name="Georgia", bold=True, size=18, color=_INK)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=0)
        ws4.merge_cells(start_row=row + 1, end_row=row + 1, start_column=1, end_column=3)
        ws4.row_dimensions[row + 1].height = 28

        # Row row+2: italic thesis
        c = ws4.cell(row=row + 2, column=1, value=thesis)
        c.font = Font(name="Georgia", italic=True, size=10, color=_INK_MUTE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws4.merge_cells(start_row=row + 2, end_row=row + 2, start_column=1, end_column=3)
        ws4.row_dimensions[row + 2].height = 18

    def _col_header(row):
        """Small uppercase column labels ala playbook table headers."""
        specs = [(2, "INPUT"), (3, "$ EQUIVALENT")]
        for col, text in specs:
            c = ws4.cell(row=row, column=col, value=text)
            c.font = Font(name="Calibri", bold=True, size=8, color=_INK_FAINT)
            c.alignment = Alignment(horizontal="right", vertical="bottom", indent=1)
            c.border = _BORDER_BOTTOM_RULE
        # Also add rule to column A for continuity
        ws4.cell(row=row, column=1).border = _BORDER_BOTTOM_RULE
        ws4.row_dimensions[row].height = 18

    arv = pkg.arv.arv_mid
    rehab_full_base = pkg.rehab_full.rooms[0].total if pkg.rehab_full.rooms else pkg.rehab_full.grand_total
    rehab_wt_base = pkg.rehab_wholetail.rooms[0].total if pkg.rehab_wholetail.rooms else pkg.rehab_wholetail.grand_total

    # ── FLIP section (rows 6-14) — P1 red stripe ──
    _section_header(6, "P1 · PRIMARY OFFER", "Flip", "Target profit % on retail sale. The primary MAO you present to seller.", _FILL_P1, _FILL_P1T, _P1)
    _col_header(9)
    _label(10, "ARV / CMV");                                       _input(10, 2, arv, _FILL_P1T)
    _label(11, "Closing % (sale-side)");                           _input(11, 2, DEFAULT_FLIP_CLOSING_PCT, _FILL_P1T, '0.00%'); _derived(11, "=B10*B11", _FILL_P1T)
    _label(12, "Holding %");                                       _input(12, 2, DEFAULT_FLIP_HOLDING_PCT, _FILL_P1T, '0.00%'); _derived(12, "=B10*B12", _FILL_P1T)
    _label(13, "Target Profit % ← Profit Calc C29");               _input(13, 2, "='Profit Calculator'!C29", _FILL_P1T, '0.00%'); _derived(13, "=B10*B13", _FILL_P1T)
    _label(14, "Rehab (base, no permits/contingency)");            _input(14, 2, rehab_full_base, _FILL_P1T)
    _label(15, "Wholesale Fee");                                   _input(15, 2, DEFAULT_FLIP_WHOLESALE_FEE, _FILL_P1T)
    _mao_row(16, "=B10*(1-B11-B12-B13)-B14-B15", _FILL_P1, big=True)

    # ── WHOLETAIL section (rows 18-25) — P2 gold stripe ──
    _section_header(18, "P2 · LIGHT COSMETIC", "Wholetail", "Cosmetic-only, sell as-is to retail buyer. Buyer profit is downstream flipper's margin.", _FILL_P2, _FILL_P2T, _P2)
    _col_header(21)
    _label(22, "CMV (defaults to flip ARV)")
    c = ws4.cell(row=22, column=2, value="=B10"); c.number_format='"$"#,##0'; c.fill=_FILL_P2T; c.border=_BORDER_BOTTOM_SOFT
    c.font = Font(name="Consolas", bold=True, size=11, color=_INK); c.alignment=Alignment(horizontal="right", vertical="center")
    _label(23, "Closing %");                                       _input(23, 2, DEFAULT_WHOLETAIL_CLOSING_PCT, _FILL_P2T, '0.00%'); _derived(23, "=B22*B23", _FILL_P2T)
    _label(24, "Wholetail Rehab (base)");                          _input(24, 2, rehab_wt_base, _FILL_P2T)
    _label(25, "Buyer Profit ← Profit Calc C30");                  _input(25, 2, "='Profit Calculator'!C30", _FILL_P2T, '"$"#,##0')
    _label(26, "Wholesale Fee");                                   _input(26, 2, DEFAULT_WHOLETAIL_WS_FEE, _FILL_P2T)
    _mao_row(27, "=B22*(1-B23)-B24-B25-B26", _FILL_P2)

    # ── RENTAL section (rows 29-36) — P3 blue-gray stripe ──
    _section_header(29, "P3 · HOLD FOR CASHFLOW", "Rental Hold", "Buy-and-hold at target equity %. MAO reflects long-term cashflow acquisition price.", _FILL_P3, _FILL_P3T, _P3)
    _col_header(32)
    _label(33, "ARV (defaults to flip ARV)")
    c = ws4.cell(row=33, column=2, value="=B10"); c.number_format='"$"#,##0'; c.fill=_FILL_P3T; c.border=_BORDER_BOTTOM_SOFT
    c.font = Font(name="Consolas", bold=True, size=11, color=_INK); c.alignment=Alignment(horizontal="right", vertical="center")
    _label(34, "Closing %");                                       _input(34, 2, DEFAULT_RENTAL_CLOSING_PCT, _FILL_P3T, '0.00%'); _derived(34, "=B33*B34", _FILL_P3T)
    _label(35, "Holding %");                                       _input(35, 2, DEFAULT_RENTAL_HOLDING_PCT, _FILL_P3T, '0.00%'); _derived(35, "=B33*B35", _FILL_P3T)
    _label(36, "Target Equity % ← Profit Calc C31");               _input(36, 2, "='Profit Calculator'!C31", _FILL_P3T, '0.00%'); _derived(36, "=B33*B36", _FILL_P3T)
    _label(37, "Rehab (base)");                                    _input(37, 2, rehab_full_base, _FILL_P3T)
    _label(38, "Wholesale Fee");                                   _input(38, 2, DEFAULT_RENTAL_WS_FEE, _FILL_P3T)
    _mao_row(39, "=B33*(1-B34-B35-B36)-B37-B38", _FILL_P3)

    # ── Footer eyebrow (row 41) ──
    c = ws4.cell(row=41, column=1, value="EDIT LINKED CELLS ON PROFIT CALCULATOR TAB  ·  MAO RESULTS RECALC LIVE")
    c.font = Font(name="Calibri", bold=True, size=8, color=_INK_FAINT)
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(1, 4):
        ws4.cell(row=41, column=col).border = _BORDER_TOP_RULE
    ws4.row_dimensions[41].height = 22

    # Column widths tuned for editorial breathing room
    ws4.column_dimensions["A"].width = 52
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 24

    # ── Tab 5: Flip Projection ────────────────────────────────────
    ws5 = wb.create_sheet("Flip Projection")
    ws5.cell(row=1, column=1, value="Flip Profit Projection").font = _TITLE_FONT
    flip_data = [
        ("Sale Price (ARV)", pkg.flip.arv),
        ("", 0),
        ("Less: Purchase Price", -pkg.flip.purchase_price),
        ("Less: Rehab Cost", -pkg.flip.rehab_cost),
        ("Less: Holding Costs", -pkg.flip.holding_costs),
        ("Less: Selling Costs", -pkg.flip.selling_costs),
        ("", 0),
        ("NET PROFIT", pkg.flip.net_profit),
        ("ROI", None),
        ("Timeline", None),
    ]
    for i, (label, value) in enumerate(flip_data, 3):
        ws5.cell(row=i, column=1, value=label).font = _LABEL_FONT
        if value is not None:
            ws5.cell(row=i, column=2, value=value).number_format = _MONEY_FMT
        if label == "NET PROFIT":
            ws5.cell(row=i, column=2).font = Font(name="Calibri", bold=True, size=14, color="006100" if value > 0 else "9C0006")
        elif label == "ROI":
            ws5.cell(row=i, column=2, value=f"{pkg.flip.roi_pct:.1f}%").font = _VALUE_FONT
        elif label == "Timeline":
            ws5.cell(row=i, column=2, value=f"{pkg.flip.months_to_complete:.0f} months").font = _VALUE_FONT
    ws5.column_dimensions["A"].width = 25
    ws5.column_dimensions["B"].width = 20

    # ── Tab 6: Hold Projection ────────────────────────────────────
    ws6 = wb.create_sheet("Hold Projection")
    ws6.cell(row=1, column=1, value="Buy & Hold Cash Flow Analysis").font = _TITLE_FONT
    hold_data = [
        ("Total Investment", _fmt_money(pkg.hold.total_investment)),
        ("Est. Monthly Rent", _fmt_money(pkg.hold.estimated_rent_monthly)),
        ("", ""),
        ("Gross Annual Income", _fmt_money(pkg.hold.gross_annual_income)),
        ("Less: Vacancy (8%)", f"- {_fmt_money(pkg.hold.vacancy_loss)}"),
        ("Effective Income", _fmt_money(pkg.hold.effective_income)),
        ("Less: Expenses", f"- {_fmt_money(pkg.hold.expenses_annual)}"),
        ("NOI", _fmt_money(pkg.hold.noi)),
        ("Less: Debt Service", f"- {_fmt_money(pkg.hold.debt_service_annual)}"),
        ("", ""),
        ("ANNUAL CASH FLOW", _fmt_money(pkg.hold.cash_flow_annual)),
        ("Cap Rate", f"{pkg.hold.cap_rate:.1f}%"),
        ("Cash-on-Cash Return", f"{pkg.hold.cash_on_cash:.1f}%"),
        ("5-Year Equity Buildup", _fmt_money(pkg.hold.equity_year5)),
    ]
    for i, (label, value) in enumerate(hold_data, 3):
        ws6.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws6.cell(row=i, column=2, value=value).font = _VALUE_FONT
    ws6.column_dimensions["A"].width = 25
    ws6.column_dimensions["B"].width = 20

    # ── Tab 7: Financing ──────────────────────────────────────────
    ws7 = wb.create_sheet("Financing")
    ws7.cell(row=1, column=1, value="Financing Comparison").font = _TITLE_FONT
    _write_headers(ws7, 3, ["Scenario", "Down Payment", "Loan Amount", "Rate",
                             "Monthly Payment", "Total Cost of Money"])
    for i, fin in enumerate(pkg.financing, 4):
        ws7.cell(row=i, column=1, value=fin.name)
        ws7.cell(row=i, column=2, value=fin.down_payment).number_format = _MONEY_FMT
        ws7.cell(row=i, column=3, value=fin.loan_amount).number_format = _MONEY_FMT
        ws7.cell(row=i, column=4, value=f"{fin.rate:.1%}" if fin.rate else "N/A")
        ws7.cell(row=i, column=5, value=fin.monthly_payment).number_format = _MONEY_FMT
        ws7.cell(row=i, column=6, value=fin.total_cost_of_money).number_format = _MONEY_FMT
        for c in range(1, 7):
            ws7.cell(row=i, column=c).border = _THIN_BORDER
    for col, w in enumerate([35, 15, 15, 10, 18, 22], 1):
        ws7.column_dimensions[chr(64 + col)].width = w

    # ── Tab 8: Risk Factors ───────────────────────────────────────
    ws8 = wb.create_sheet("Risk Factors")
    ws8.cell(row=1, column=1, value="Risk Assessment").font = _TITLE_FONT
    for i, risk in enumerate(pkg.risk_factors, 3):
        cell = ws8.cell(row=i, column=1, value=f"• {risk}")
        cell.font = _LABEL_FONT
        if "LOW" in risk or "Wide" in risk or "Only" in risk or "Thin" in risk:
            cell.font = Font(name="Calibri", color="9C0006")
    ws8.column_dimensions["A"].width = 70

    # Save — filename = address + timestamp (no "deal_analysis_" prefix so
    # folder sorts alphabetically by address). Timestamp retained so repeat
    # runs on the same property don't overwrite each other.
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_addr = "".join(c if c.isalnum() or c == "-" else "_" for c in pkg.subject.address)[:40]
        output_path = str(config.DEALS_DIR / f"{safe_addr}_{timestamp}.xlsx")

    wb.save(output_path)
    logger.info("Deal analysis saved to %s", output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

def run_deal_analysis(address: str, city: str = "", state: str = "",
                      zip_code: str = "", purchase_price: float = 0,
                      rehab_tier: int = 2, exit_strategy: str = "flip",
                      region: str = DEFAULT_REGION,
                      radius: float = DEFAULT_RADIUS_MILES,
                      months: int = DEFAULT_MONTHS_BACK,
                      output_path: str = "") -> dict:
    """Run complete deal analysis: comp → rehab → MAO → projections → report.

    Returns dict with the DealPackage and report path. Empty state
    falls back to DEFAULT_PROPERTY_STATE (currently AL — single knob
    to flip when scaling to new states).
    """
    if not state:
        from state_resolver import DEFAULT_PROPERTY_STATE
        state = DEFAULT_PROPERTY_STATE
    logger.info("Starting deal analysis for: %s", address)

    # Step 1: Fetch subject property
    subject = fetch_subject_property(address, city, state, zip_code)
    if not subject:
        return {"error": "Could not fetch property data"}

    # Step 2: Fetch comps and calculate ARV
    comps = fetch_comparable_sales(subject, radius, months)
    arv = calculate_arv(subject, comps)
    # Also fetch rental comps for the Rental Calc tab (active FOR_RENT listings)
    rental_comps = fetch_rental_comps(subject, radius_miles=1.0, n=6)
    # Pending sales — directional ARV signal (under-contract listings front-run sold by 30-60d)
    pending_comps = fetch_pending_sales(subject, radius_miles=radius)

    if arv.confidence == "none":
        logger.warning("No ARV could be calculated — insufficient comp data")

    # Auto-resolve regional cost multiplier from subject location unless caller
    # passed an explicit non-"auto" override.
    from rehab_estimator import resolve_region
    resolved_region = resolve_region(subject.city, subject.state, region)
    if resolved_region != region:
        logger.info("Region resolver: %s → %s (from %s, %s)",
                    region, resolved_region, subject.city, subject.state)

    # Step 3: Rehab estimates (pass city/state so resolver fires consistently downstream)
    rehab_full = estimate_rehab(
        address=subject.address, sqft=subject.sqft, bedrooms=subject.bedrooms,
        bathrooms=subject.bathrooms, year_built=subject.year_built,
        tier=rehab_tier, scope="full", region=resolved_region,
        city=subject.city, state=subject.state,
    )
    rehab_wholetail = estimate_wholetail(
        address=subject.address, sqft=subject.sqft, bedrooms=subject.bedrooms,
        bathrooms=subject.bathrooms, year_built=subject.year_built,
        region=resolved_region, city=subject.city, state=subject.state,
    )

    # Per 2026-05-31 user request: the Profit/Wholetail/Rental calculators already
    # pad via closing % + holding % at the deal level, so the Rehab number in those
    # calcs should be the BASE rehab only (no permits, no contingency padding).
    # Standalone Rehab Budget tab keeps the full breakdown (permits + contingency)
    # since that's what you actually budget for the contractor.
    base_rehab_full = rehab_full.rooms[0].total if rehab_full.rooms else rehab_full.grand_total
    base_rehab_wholetail = rehab_wholetail.rooms[0].total if rehab_wholetail.rooms else rehab_wholetail.grand_total

    # Step 4: MAO (calculate first so we can use it as the default purchase price)
    mao = calculate_mao(arv.arv_mid, base_rehab_full)

    # Use purchase price or default to flip MAO (matches headline = consistent across tabs)
    if not purchase_price:
        purchase_price = max(0, mao.flip_mao)
        logger.info("No purchase price given — defaulting to flip MAO (15%% target): %s", _fmt_money(purchase_price))

    # Step 5: Projections (Flip/Wholesale/Hold projections use grand_total since they
    # model the full cost-of-money including contractor's permit/contingency budget)
    rehab_months = rehab_full.total_weeks / 4.0  # weeks to months
    holding = calculate_holding_costs(purchase_price, rehab_months)
    selling = calculate_selling_costs(arv.arv_mid, subject.state)

    flip = calculate_flip(arv.arv_mid, purchase_price, rehab_full.grand_total,
                          holding, selling, rehab_months)
    wholesale = calculate_wholesale(arv.arv_mid, purchase_price, rehab_full.grand_total)
    hold = calculate_hold(purchase_price, rehab_full.grand_total, arv.arv_mid,
                          subject.sqft, subject.bedrooms)
    financing = calculate_financing(purchase_price, rehab_full.grand_total)

    # Step 6: Profit Calculator breakdowns — use BASE rehab (no permits + contingency)
    market_tier = classify_market_tier(arv.arv_mid, subject.city, subject.state)
    profit_breakdown = calculate_profit_breakdown(
        arv.arv_mid, purchase_price, base_rehab_full)
    wholetail_breakdown = calculate_wholetail_breakdown(
        arv.arv_mid, purchase_price, base_rehab_wholetail)
    # Novation uses the LIGHT (wholetail-grade) rehab — cosmetic only,
    # listing-ready, not full renovation.
    novation_breakdown = calculate_novation_breakdown(
        arv.arv_mid, base_rehab_wholetail)
    rental_breakdown = calculate_rental_breakdown(
        arv.arv_mid, purchase_price, base_rehab_full,
        monthly_rent=hold.estimated_rent_monthly, market_tier=market_tier)

    # Step 7: Risk assessment and recommendation
    risk_factors = _assess_risk(arv, flip, subject)
    recommendation = _make_recommendation(flip, wholesale, hold, arv)

    # Assemble package
    pkg = DealPackage(
        subject=subject,
        arv=arv,
        rehab_full=rehab_full,
        rehab_wholetail=rehab_wholetail,
        mao=mao,
        flip=flip,
        wholesale=wholesale,
        hold=hold,
        holding_costs=holding,
        selling_costs=selling,
        financing=financing,
        recommendation=recommendation,
        risk_factors=risk_factors,
        profit_breakdown=profit_breakdown,
        wholetail_breakdown=wholetail_breakdown,
        novation_breakdown=novation_breakdown,
        rental_breakdown=rental_breakdown,
        market_tier=market_tier,
        comps=comps,
        rental_comps=rental_comps,
        pending_comps=pending_comps,
    )

    # Step 7: Generate report
    report_path = generate_deal_report(pkg, output_path)

    logger.info("Deal analysis complete: %s | ARV %s | Flip profit %s (%s%% ROI) | %s",
                subject.address, _fmt_money(arv.arv_mid),
                _fmt_money(flip.net_profit), f"{flip.roi_pct:.0f}",
                recommendation.split("—")[0].strip())

    return {
        "package": pkg,
        "report_path": report_path,
    }
