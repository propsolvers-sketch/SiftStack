"""Cash buyer list building and investor portfolio analysis.

Identifies active real estate investors from property transaction data,
classifies them by investment strategy, and generates contact-ready buyer lists.

Buyer Types:
  - Wholesale: rapid assignment <72 days, multiple quick transactions
  - Wholetail: cosmetic + MLS list within 90 days
  - Flip: full rehab, 90-day to 18-month hold
  - Buy-and-Hold: 18+ month hold, portfolio builders

Data sources:
  - Our own scraped/enriched CSV records (transaction patterns)
  - Entity research (existing entity_researcher.py)
  - Skip trace (existing tracerfy_skip_tracer.py)

Usage:
  python src/main.py buyer-prospect --counties Knox,Blount --months-back 18
"""

import csv
import logging
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── Buyer classification thresholds ───────────────────────────────────
WHOLESALE_MAX_HOLD_DAYS = 72
WHOLETAIL_MAX_HOLD_DAYS = 90
FLIP_MAX_HOLD_DAYS = 540     # 18 months
# > 540 days = buy-and-hold

MIN_TRANSACTIONS = 2         # Minimum to be considered an investor
LOOKBACK_MONTHS = 18         # Default lookback period

# ── Scoring weights ───────────────────────────────────────────────────
WEIGHT_RECENCY = 0.30         # More recent activity = better
WEIGHT_FREQUENCY = 0.25       # More transactions = better
WEIGHT_DEAL_SIZE = 0.20       # Larger avg deals = more capital
WEIGHT_GEOGRAPHIC = 0.15      # Focused in target area = better
WEIGHT_CONSISTENCY = 0.10     # Regular cadence = better


@dataclass
class InvestorProfile:
    """Profile for an identified investor."""
    name: str = ""
    entity_name: str = ""         # LLC/Corp name if applicable
    person_behind: str = ""       # Person identified via entity research
    buyer_type: str = ""          # wholesale, wholetail, flip, hold
    transaction_count: int = 0
    avg_purchase_price: float = 0.0
    total_invested: float = 0.0
    avg_hold_days: float = 0.0
    zip_codes: list = field(default_factory=list)
    primary_zip: str = ""
    property_types: list = field(default_factory=list)
    first_transaction: str = ""
    last_transaction: str = ""
    score: float = 0.0
    rank: int = 0
    # Contact info (from skip trace)
    phone: str = ""
    email: str = ""
    mailing_address: str = ""


@dataclass
class BuyerReport:
    """Complete buyer prospecting report."""
    county: str = ""
    analysis_date: str = ""
    lookback_months: int = LOOKBACK_MONTHS
    total_investors: int = 0
    investors: list = field(default_factory=list)
    by_type: dict = field(default_factory=dict)
    top_zips: list = field(default_factory=list)


# ── Data loading ──────────────────────────────────────────────────────

def _load_transaction_data(counties: list[str] | None = None,
                           months_back: int = LOOKBACK_MONTHS) -> list[dict]:
    """Load transaction data from our enriched CSV files."""
    records = []
    cutoff = (datetime.now() - timedelta(days=months_back * 30)).strftime("%Y-%m-%d")

    for csv_path in config.OUTPUT_DIR.glob("*.csv"):
        try:
            with open(csv_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    county = (row.get("county") or "").strip().lower()
                    if counties and county not in {c.lower() for c in counties}:
                        continue

                    date_added = row.get("date_added") or row.get("Date Added") or ""
                    if date_added and date_added < cutoff:
                        continue

                    records.append(row)
        except Exception as e:
            logger.debug("Error reading %s: %s", csv_path, e)

    logger.info("Loaded %d records for buyer analysis", len(records))
    return records


# ── Investor identification ───────────────────────────────────────────

def _identify_investors(records: list[dict],
                        min_transactions: int = MIN_TRANSACTIONS) -> list[InvestorProfile]:
    """Identify investors from transaction patterns.

    An investor is someone who appears as an owner across multiple properties
    or has entity ownership patterns (LLC, Corp, Trust).
    """
    # Group by owner name
    owner_records = defaultdict(list)
    for row in records:
        owner = (row.get("owner_name") or row.get("Owner Name") or
                 row.get("full_name") or "").strip()
        if not owner or len(owner) < 3:
            continue
        owner_records[owner.upper()].append(row)

    investors = []
    for name, recs in owner_records.items():
        # Check if this looks like an investor
        is_entity = any(kw in name.upper() for kw in
                        ["LLC", "CORP", "INC", "TRUST", "LP", "LLP", "LTD",
                         "PROPERTIES", "HOLDINGS", "INVESTMENTS", "CAPITAL",
                         "REALTY", "REAL ESTATE", "VENTURES"])

        if len(recs) < min_transactions and not is_entity:
            continue

        # Analyze transaction patterns
        prices = []
        zips = []
        types = []
        dates = []

        for r in recs:
            price = r.get("mls_last_sold_price") or r.get("estimated_value") or ""
            if price:
                try:
                    prices.append(float(str(price).replace(",", "").replace("$", "")))
                except ValueError:
                    pass

            z = (r.get("zip") or "")[:5]
            if z:
                zips.append(z)

            pt = r.get("property_type") or ""
            if pt:
                types.append(pt)

            d = r.get("date_added") or r.get("mls_last_sold_date") or ""
            if d:
                dates.append(d)

        # Classify buyer type based on hold patterns
        avg_price = sum(prices) / len(prices) if prices else 0

        # Default classification (without hold time data, use heuristics)
        if len(recs) >= 5 and is_entity:
            buyer_type = "hold"       # High volume entity = portfolio builder
        elif len(recs) >= 3 and avg_price < 150000:
            buyer_type = "wholesale"  # Multiple cheap properties
        elif is_entity:
            buyer_type = "flip"       # Entity with moderate activity
        else:
            buyer_type = "flip"       # Default

        zip_counter = Counter(zips)
        primary_zip = zip_counter.most_common(1)[0][0] if zip_counter else ""

        profile = InvestorProfile(
            name=name.title(),
            entity_name=name.title() if is_entity else "",
            buyer_type=buyer_type,
            transaction_count=len(recs),
            avg_purchase_price=round(avg_price),
            total_invested=round(sum(prices)),
            zip_codes=list(set(zips)),
            primary_zip=primary_zip,
            property_types=list(set(types)),
            first_transaction=min(dates) if dates else "",
            last_transaction=max(dates) if dates else "",
        )
        investors.append(profile)

    logger.info("Identified %d investors from %d unique owners",
                len(investors), len(owner_records))
    return investors


# ── Scoring ───────────────────────────────────────────────────────────

def _score_investors(investors: list[InvestorProfile]) -> list[InvestorProfile]:
    """Score investors for buyer quality."""
    if not investors:
        return []

    now = datetime.now()

    for inv in investors:
        recency_score = 50.0
        if inv.last_transaction:
            try:
                last = datetime.strptime(inv.last_transaction[:10], "%Y-%m-%d")
                days_ago = (now - last).days
                recency_score = max(0, 100 - days_ago * 0.5)
            except ValueError:
                pass

        freq_score = min(100, inv.transaction_count * 20)
        size_score = min(100, inv.avg_purchase_price / 3000) if inv.avg_purchase_price else 0
        geo_score = max(0, 100 - len(inv.zip_codes) * 10)  # More focused = better
        consistency_score = 50.0  # Default without cadence data

        inv.score = (
            recency_score * WEIGHT_RECENCY +
            freq_score * WEIGHT_FREQUENCY +
            size_score * WEIGHT_DEAL_SIZE +
            geo_score * WEIGHT_GEOGRAPHIC +
            consistency_score * WEIGHT_CONSISTENCY
        )

    investors.sort(key=lambda x: x.score, reverse=True)
    for i, inv in enumerate(investors):
        inv.rank = i + 1

    return investors


# ── Excel report ──────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
_MONEY_FMT = '#,##0'
_TYPE_COLORS = {
    "wholesale": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "wholetail": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
    "flip": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "hold": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
}


def _write_headers(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_widths(ws, min_w=12, max_w=30):
    for col in ws.columns:
        mx = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max(mx + 2, min_w), max_w)


def generate_buyer_report(report: BuyerReport, output_path: str = "") -> str:
    """Generate buyer prospecting Excel workbook."""
    wb = Workbook()

    # ── Tab 1: Buyer Scorecard ────────────────────────────────────
    ws = wb.active
    ws.title = "Buyer Scorecard"
    ws.cell(row=1, column=1, value="Cash Buyer Prospecting Report").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"County: {report.county} | {report.analysis_date}").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Lookback: {report.lookback_months} months | {report.total_investors} investors found").font = _LABEL_FONT

    headers = ["Rank", "Name", "Buyer Type", "Score", "# Transactions",
               "Avg Price", "Total Invested", "Primary ZIP", "Last Activity"]
    _write_headers(ws, 5, headers)
    for i, inv in enumerate(report.investors[:50], 6):
        ws.cell(row=i, column=1, value=inv.rank)
        ws.cell(row=i, column=2, value=inv.name)
        type_cell = ws.cell(row=i, column=3, value=inv.buyer_type.title())
        type_cell.fill = _TYPE_COLORS.get(inv.buyer_type, PatternFill())
        ws.cell(row=i, column=4, value=round(inv.score, 1))
        ws.cell(row=i, column=5, value=inv.transaction_count)
        ws.cell(row=i, column=6, value=inv.avg_purchase_price).number_format = _MONEY_FMT
        ws.cell(row=i, column=7, value=inv.total_invested).number_format = _MONEY_FMT
        ws.cell(row=i, column=8, value=inv.primary_zip)
        ws.cell(row=i, column=9, value=inv.last_transaction)
        for c in range(1, 10):
            ws.cell(row=i, column=c).border = _THIN_BORDER
    _auto_widths(ws)

    # ── Tab 2: Portfolio Detail ───────────────────────────────────
    ws2 = wb.create_sheet("Portfolio Detail")
    ws2.cell(row=1, column=1, value="Investor Portfolio Breakdown").font = _TITLE_FONT
    _write_headers(ws2, 3, ["Name", "Type", "Transactions", "Avg Price",
                             "Zip Codes", "Property Types", "First Deal", "Last Deal"])
    for i, inv in enumerate(report.investors[:50], 4):
        ws2.cell(row=i, column=1, value=inv.name)
        ws2.cell(row=i, column=2, value=inv.buyer_type.title())
        ws2.cell(row=i, column=3, value=inv.transaction_count)
        ws2.cell(row=i, column=4, value=inv.avg_purchase_price).number_format = _MONEY_FMT
        ws2.cell(row=i, column=5, value=", ".join(inv.zip_codes[:5]))
        ws2.cell(row=i, column=6, value=", ".join(inv.property_types[:3]))
        ws2.cell(row=i, column=7, value=inv.first_transaction)
        ws2.cell(row=i, column=8, value=inv.last_transaction)
        for c in range(1, 9):
            ws2.cell(row=i, column=c).border = _THIN_BORDER
    _auto_widths(ws2)

    # ── Tab 3: Geographic ─────────────────────────────────────────
    ws3 = wb.create_sheet("Geographic Heat Map")
    ws3.cell(row=1, column=1, value="Investor Concentration by ZIP").font = _TITLE_FONT
    zip_counter = Counter()
    for inv in report.investors:
        for z in inv.zip_codes:
            zip_counter[z] += 1
    _write_headers(ws3, 3, ["ZIP", "Investor Count"])
    for i, (z, count) in enumerate(zip_counter.most_common(30), 4):
        ws3.cell(row=i, column=1, value=z)
        ws3.cell(row=i, column=2, value=count)
    _auto_widths(ws3)

    # ── Tab 4: Contact List ───────────────────────────────────────
    ws4 = wb.create_sheet("Contact List")
    ws4.cell(row=1, column=1, value="Buyer Contact List (DataSift-Ready)").font = _TITLE_FONT
    ws4.cell(row=2, column=1,
             value="Run skip trace to populate phone/email columns").font = _LABEL_FONT
    _write_headers(ws4, 4, ["Rank", "Name", "Entity Name", "Buyer Type",
                             "Phone", "Email", "Mailing Address", "Score"])
    for i, inv in enumerate(report.investors[:100], 5):
        ws4.cell(row=i, column=1, value=inv.rank)
        ws4.cell(row=i, column=2, value=inv.person_behind or inv.name)
        ws4.cell(row=i, column=3, value=inv.entity_name)
        ws4.cell(row=i, column=4, value=inv.buyer_type.title())
        ws4.cell(row=i, column=5, value=inv.phone)
        ws4.cell(row=i, column=6, value=inv.email)
        ws4.cell(row=i, column=7, value=inv.mailing_address)
        ws4.cell(row=i, column=8, value=round(inv.score, 1))
        for c in range(1, 9):
            ws4.cell(row=i, column=c).border = _THIN_BORDER
    _auto_widths(ws4)

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(config.OUTPUT_DIR / f"buyer_prospecting_{report.county}_{timestamp}.xlsx")

    wb.save(output_path)
    logger.info("Buyer report saved to %s", output_path)
    return output_path


# ── CSV export for DataSift ───────────────────────────────────────────

def export_buyers_csv(investors: list[InvestorProfile], output_path: str = "") -> str:
    """Export top buyers as DataSift-ready CSV."""
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(config.OUTPUT_DIR / f"buyers_datasift_{timestamp}.csv")

    headers = ["owner_name", "address", "city", "state", "zip", "tags", "lists", "notes"]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        # Buyer state defaults to the active primary state — buyers in
        # this list invest in that state's properties (primary_zip is
        # in-market). DEFAULT_PROPERTY_STATE is the single knob for
        # scaling to new states.
        from state_resolver import DEFAULT_PROPERTY_STATE
        buyer_state = DEFAULT_PROPERTY_STATE
        writer.writeheader()
        for inv in investors[:100]:
            writer.writerow({
                "owner_name": inv.person_behind or inv.name,
                "address": "",
                "city": "",
                "state": buyer_state,
                "zip": inv.primary_zip,
                "tags": f"buyer,{inv.buyer_type},buyer_score_{round(inv.score)}",
                "lists": "Cash Buyers",
                "notes": f"Type: {inv.buyer_type.title()}, {inv.transaction_count} deals, "
                         f"Avg ${inv.avg_purchase_price:,.0f}",
            })

    logger.info("Exported %d buyers to %s", min(len(investors), 100), output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

def run_buyer_prospecting(counties: list[str] | None = None,
                          months_back: int = LOOKBACK_MONTHS,
                          min_transactions: int = MIN_TRANSACTIONS,
                          output_path: str = "") -> dict:
    """Run buyer prospecting analysis.

    Returns dict with report and output paths.
    """
    counties = counties or ["Knox", "Blount"]
    county_str = ", ".join(counties)
    logger.info("Starting buyer prospecting for: %s (last %d months)", county_str, months_back)

    # Load transaction data
    records = _load_transaction_data(counties, months_back)
    if not records:
        return {"error": "No transaction data found"}

    # Identify and classify investors
    investors = _identify_investors(records, min_transactions)
    if not investors:
        return {"error": "No investors identified"}

    # Score and rank
    scored = _score_investors(investors)

    # Type breakdown
    by_type = Counter(inv.buyer_type for inv in scored)

    report = BuyerReport(
        county=county_str,
        analysis_date=datetime.now().strftime("%Y-%m-%d"),
        lookback_months=months_back,
        total_investors=len(scored),
        investors=scored,
        by_type=dict(by_type),
    )

    # Generate reports
    report_path = generate_buyer_report(report, output_path)
    csv_path = export_buyers_csv(scored)

    logger.info("Buyer prospecting complete: %d investors, %s",
                len(scored), ", ".join(f"{t}: {c}" for t, c in by_type.items()))

    return {
        "report": report,
        "report_path": report_path,
        "csv_path": csv_path,
    }
