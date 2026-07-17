"""Comparable sales analysis with Two-Bucket ARV methodology.

Generates appraiser-grade property valuations for real estate investment
analysis. Fetches comparable sales from the Zillow API, applies property-
specific adjustments, and produces a 7-tab Excel workbook.

Tennessee is a non-disclosure state — MLS/Zillow data is the primary
source, not public deed records.

Usage:
  python src/main.py comp --address "123 Main St, Knoxville, TN 37918"
  python src/main.py comp --address "123 Main St" --city Knoxville --zip 37918 --radius 0.5 --months 6
"""

import logging
import math
import random
import re
import time
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config

logger = logging.getLogger(__name__)

# ── API Configuration ─────────────────────────────────────────────────
API_BASE = "https://api.openwebninja.com/realtime-zillow-data"
PROPERTY_ENDPOINT = f"{API_BASE}/property-details-address"
SEARCH_ENDPOINT = f"{API_BASE}/search"
# Sold comps come from `/search?location=<zip>&home_status=RECENTLY_SOLD` — the same
# query Zillow's web UI uses for its "Sold" filter. Returns the real sold price
# (`unformattedPrice`) and sold date (`dateSold`, epoch-ms).
# `nearbyHomes` inside `/property-details-address` is a fallback only — it's Zillow's
# curated "you might also like" panel and rarely contains SOLD records.
REQUEST_DELAY_MIN = 1.0
REQUEST_DELAY_MAX = 2.0
REQUEST_TIMEOUT = 30
MAX_RETRIES = 4  # bumped from 2: OpenWeb Ninja returns 200/OK with empty `data` ~25% of calls

# ── Comp selection defaults (revised 2026-05-31 for dual-source pull) ─
# Strategy: pull from BOTH Zillow + Redfin, dedupe, rank by similarity.
# Start with tight 1mi/6mo window. If combined pool < TARGET, expand radius
# then expand timeframe. Display top 10 by similarity; ARV = average of top 5.
DEFAULT_RADIUS_MILES = 1.0      # Primary search radius — proximity is #1 weight
MAX_RADIUS_MILES = 3.0          # Final fallback if combined pool still too thin
DEFAULT_MONTHS_BACK = 6         # Primary lookback — recency matters for fast-moving markets
MAX_MONTHS_BACK = 18            # Final fallback if combined pool still too thin
MIN_COMPS = 3                   # Below this, ARV confidence drops to "low"
TARGET_COMPS = 10               # Display target — 10 ranked-by-similarity comps shown on report
MAX_COMPS = 10                  # Hard cap on comps in candidate pool after dedup
ARV_AVG_TOP_N = 5               # 2026-05-31: ARV = mean of top N comps by similarity score
                                # (replaces middle-3-by-price trimmed mean)

# ── Adjustment values (Knoxville regional calibration) ────────────────
# These are per-unit adjustment amounts used when a comp differs from subject
ADJ_PER_SQFT = 85.0            # $ per sqft difference
ADJ_PER_BEDROOM = 5000.0       # $ per bedroom difference
ADJ_PER_BATHROOM = 7500.0      # $ per bathroom difference
ADJ_PER_YEAR_BUILT = 500.0     # $ per year of age difference
ADJ_PER_LOT_SQFT = 2.0         # $ per sqft of lot size difference
ADJ_LOT_MAX = 15000.0          # Cap on lot size adjustment
ADJ_PER_GARAGE = 8000.0        # $ per garage stall difference
# Market condition adjustment: % per month of age (appreciating market)
MARKET_CONDITION_PCT_PER_MONTH = 0.003  # 0.3% per month ≈ 3.6% annual

# ── Data structures ───────────────────────────────────────────────────


@dataclass
class SubjectProperty:
    """The property being analyzed."""
    address: str = ""
    city: str = ""
    # Empty default — populated by fetch_subject_property() from the
    # property record or CLI args. Was "TN" historically; emptied so
    # missing-state isn't silently TN-stamped on AL properties.
    state: str = ""
    zip_code: str = ""
    latitude: float = 0.0
    longitude: float = 0.0
    sqft: int = 0
    bedrooms: int = 0
    bathrooms: float = 0.0
    year_built: int = 0
    lot_sqft: int = 0
    property_type: str = ""
    zestimate: float = 0.0
    mls_status: str = ""
    last_sold_date: str = ""
    last_sold_price: float = 0.0
    garage_spaces: int = 0
    description: str = ""


@dataclass
class CompProperty:
    """A comparable property with sale data."""
    address: str = ""
    city: str = ""
    # Empty default — populated by the API response (OpenWeb Ninja
    # returns the actual state per comparable). Was "TN" historically.
    state: str = ""
    zip_code: str = ""
    latitude: float = 0.0
    longitude: float = 0.0
    distance_miles: float = 0.0
    sqft: int = 0
    bedrooms: int = 0
    bathrooms: float = 0.0
    year_built: int = 0
    lot_sqft: int = 0
    property_type: str = ""
    detail_url: str = ""           # Zillow listing URL for clickable comps in report
    sold_price: float = 0.0
    sold_date: str = ""
    days_on_market: int = 0
    garage_spaces: int = 0
    # Calculated fields
    similarity_score: float = 0.0
    adjusted_price: float = 0.0
    ppsf: float = 0.0
    bucket: str = ""  # "A" (non-disclosure baseline) or "B" (disclosure/adjusted)
    adjustments: dict = field(default_factory=dict)
    # Source tracking (2026-05-31: dual-source comp pulls from Zillow + Redfin)
    source: str = "zillow"  # "zillow", "redfin", or "zillow+redfin" if same comp confirmed by both
    mls_number: str = ""
    # Price tier (2026-06-09: bias ARV toward renovated comps in fix-n-flip markets)
    # Set in calculate_arv() based on pool median PPSF: "renovated" / "standard" / "distressed"
    tier: str = "standard"


@dataclass
class ARVResult:
    """Final ARV calculation result."""
    arv_low: float = 0.0
    arv_mid: float = 0.0
    arv_high: float = 0.0
    confidence: str = ""  # "high", "medium", "low"
    confidence_reason: str = ""
    ppsf_avg: float = 0.0
    ppsf_range: tuple = (0.0, 0.0)
    comp_count: int = 0
    bucket_a_count: int = 0
    bucket_b_count: int = 0
    avg_adjustment: float = 0.0
    spread_pct: float = 0.0


# ── Distance calculation ──────────────────────────────────────────────

def _haversine_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate distance between two lat/lon points in miles."""
    R = 3958.8  # Earth radius in miles
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return R * 2 * math.asin(math.sqrt(a))


# ── API calls ─────────────────────────────────────────────────────────

def _api_get(endpoint: str, params: dict, api_key: str) -> dict | None:
    """Make an authenticated GET request to OpenWeb Ninja API."""
    headers = {"x-api-key": api_key}
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(endpoint, headers=headers, params=params, timeout=REQUEST_TIMEOUT)
            if resp.status_code == 404:
                return None
            if resp.status_code == 429:
                logger.warning("Rate limit hit — waiting 10s (attempt %d)", attempt)
                time.sleep(10)
                continue
            resp.raise_for_status()
            body = resp.json()
            if body.get("status") == "OK" and body.get("data"):
                return body["data"]
            if isinstance(body, list):
                return body
            # 200 OK but empty `data` — OpenWeb Ninja flakes here intermittently.
            # Brief backoff and retry rather than treat as "not found".
            logger.warning("Empty data field (attempt %d/%d) — retrying", attempt, MAX_RETRIES)
            time.sleep(1.5)
            continue
        except requests.Timeout:
            logger.warning("Timeout (attempt %d/%d)", attempt, MAX_RETRIES)
        except requests.RequestException as e:
            logger.warning("API error: %s (attempt %d/%d)", e, attempt, MAX_RETRIES)
    return None


def fetch_subject_property(address: str, city: str = "", state: str = "",
                           zip_code: str = "", api_key: str = "") -> SubjectProperty | None:
    """Fetch full property details for the subject property.

    Empty state falls back to DEFAULT_PROPERTY_STATE (currently AL).
    """
    if not state:
        from state_resolver import DEFAULT_PROPERTY_STATE
        state = DEFAULT_PROPERTY_STATE
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key:
        logger.error("No OpenWeb Ninja API key configured")
        return None

    parts = [p for p in [address, city, state, zip_code] if p]
    full_address = " ".join(parts)

    data = _api_get(PROPERTY_ENDPOINT, {"address": full_address}, api_key)
    if not data:
        logger.warning("No property data found for '%s'", full_address)
        return None

    # Parse price history for last sold
    last_sold_date, last_sold_price = "", 0.0
    for entry in (data.get("priceHistory") or []):
        event = (entry.get("event") or "").lower()
        if event in ("sold", "listed (sold)"):
            last_sold_date = str(entry.get("date", ""))[:10]
            last_sold_price = float(entry.get("price") or 0)
            break

    # Parse lot size
    lot_sqft = 0
    lot_val = data.get("lotAreaValue")
    lot_units = (data.get("lotAreaUnits") or data.get("lotAreaUnit") or "").lower()
    if lot_val:
        lot_sqft = int(float(lot_val) * 43560) if "acre" in lot_units else int(float(lot_val))

    return SubjectProperty(
        address=data.get("streetAddress") or address,
        city=data.get("city") or city,
        state=data.get("state") or state,
        zip_code=str(data.get("zipcode") or zip_code),
        latitude=float(data.get("latitude") or 0),
        longitude=float(data.get("longitude") or 0),
        sqft=int(data.get("livingArea") or 0),
        bedrooms=int(data.get("bedrooms") or 0),
        bathrooms=float(data.get("bathrooms") or 0),
        year_built=int(data.get("yearBuilt") or 0),
        lot_sqft=lot_sqft,
        property_type=data.get("homeType") or "",
        zestimate=float(data.get("zestimate") or 0),
        mls_status=data.get("homeStatus") or "",
        last_sold_date=last_sold_date,
        last_sold_price=last_sold_price,
        garage_spaces=int(data.get("garageSpaces") or 0),
        description=data.get("description") or "",
    )


_SOLD_STATUSES = {"SOLD", "RECENTLY_SOLD"}


def _normalize_comp_item(item: dict, subject: SubjectProperty) -> dict | None:
    """Normalize a /search or nearbyHomes record into a flat field dict.

    The two endpoints return overlapping but differently-shaped records:
    /search has top-level scalars (`address` as string, `dateSold` as epoch-ms,
    `unformattedPrice` as int); nearbyHomes nests address in a sub-dict and
    rarely populates `lastSoldDate`/`lastSoldPrice`. Returns None if the record
    isn't a sold comp.
    """
    if not isinstance(item, dict):
        return None

    status = (item.get("homeStatus") or "").upper()
    if status not in _SOLD_STATUSES:
        return None

    # Address can be a dict (nearbyHomes) or string (search)
    addr = item.get("address")
    if isinstance(addr, dict):
        street = addr.get("streetAddress") or ""
        city = addr.get("city") or ""
        state = addr.get("state") or subject.state
        zipc = str(addr.get("zipcode") or "")
    else:
        street = item.get("streetAddress") or item.get("addressStreet") or ""
        city = item.get("city") or item.get("addressCity") or ""
        state = item.get("state") or item.get("addressState") or subject.state
        zipc = str(item.get("zipcode") or item.get("addressZipcode") or "")

    # Sold price: search puts the real sold price in `unformattedPrice`/`price`;
    # nearbyHomes occasionally has `lastSoldPrice` but usually `price` is the only signal.
    sold_price = float(
        item.get("unformattedPrice")
        or item.get("lastSoldPrice")
        or item.get("price")
        or 0
    )

    # Sold date: search returns epoch-ms in `dateSold`; nearbyHomes returns a
    # date string in `lastSoldDate` (or None).
    sold_date = ""
    ds = item.get("dateSold") or item.get("lastSoldDate")
    if isinstance(ds, (int, float)):
        try:
            sold_date = datetime.fromtimestamp(int(ds) / 1000).strftime("%Y-%m-%d")
        except (ValueError, OSError):
            pass
    elif isinstance(ds, str):
        sold_date = ds[:10]

    lot_sqft = 0
    lot_val = item.get("lotAreaValue") or item.get("lotSize")
    lot_units = (item.get("lotAreaUnits") or item.get("lotAreaUnit") or "").lower()
    if lot_val:
        try:
            lot_sqft = int(float(lot_val) * 43560) if "acre" in lot_units else int(float(lot_val))
        except (ValueError, TypeError):
            pass

    return {
        "street": street, "city": city, "state": state, "zipc": zipc,
        "lat": float(item.get("latitude") or 0),
        "lon": float(item.get("longitude") or 0),
        "sqft": int(item.get("livingArea") or item.get("area") or 0),
        "bedrooms": int(item.get("bedrooms") or item.get("beds") or 0),
        "bathrooms": float(item.get("bathrooms") or item.get("baths") or 0),
        "year_built": int(item.get("yearBuilt") or 0),
        "lot_sqft": lot_sqft,
        "property_type": item.get("homeType") or item.get("propertyTypeDimension") or "",
        "sold_price": sold_price,
        "sold_date": sold_date,
        "days_on_market": int(item.get("daysOnZillow") or 0),
        "garage_spaces": int(item.get("garageSpaces") or 0),
        "zpid": str(item.get("zpid") or item.get("id") or ""),
        "detail_url": item.get("detailUrl") or item.get("hdpUrl") or "",
    }


def _fetch_zillow_sold_pool(subject: SubjectProperty, radius_miles: float,
                            months_back: int, api_key: str) -> list[CompProperty]:
    """Inner: pull RECENTLY_SOLD from Zillow via OpenWeb Ninja at the given window."""
    if not subject.zip_code:
        return []
    search_data = _api_get(
        SEARCH_ENDPOINT,
        {"location": subject.zip_code, "home_status": "RECENTLY_SOLD"},
        api_key,
    )
    items = search_data if isinstance(search_data, list) else []
    logger.info("Zillow /search returned %d RECENTLY_SOLD records for ZIP %s",
                len(items), subject.zip_code)
    if not items:
        full_address = f"{subject.address} {subject.city} {subject.state} {subject.zip_code}"
        details = _api_get(PROPERTY_ENDPOINT, {"address": full_address}, api_key)
        items = (details or {}).get("nearbyHomes") or []
        logger.info("Zillow /search yielded 0 — fallback nearbyHomes (%d items)", len(items))
    time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))

    cutoff_date = (datetime.now() - timedelta(days=months_back * 30)).strftime("%Y-%m-%d")
    comps = []
    for item in items:
        n = _normalize_comp_item(item, subject)
        if not n:
            continue
        if n["street"] and n["street"].lower().strip() == subject.address.lower().strip():
            continue
        if n["sold_price"] < 10000:
            continue
        if n["sold_date"] and n["sold_date"] < cutoff_date:
            continue
        dist = 0.0
        if subject.latitude and subject.longitude and n["lat"] and n["lon"]:
            dist = _haversine_miles(subject.latitude, subject.longitude, n["lat"], n["lon"])
            if dist > radius_miles:
                continue
        detail_url = n.get("detail_url") or ""
        if not detail_url and n.get("zpid"):
            detail_url = f"https://www.zillow.com/homedetails/{n['zpid']}_zpid/"
        comp = CompProperty(
            address=n["street"], city=n["city"], state=n["state"], zip_code=n["zipc"],
            latitude=n["lat"], longitude=n["lon"], distance_miles=round(dist, 2),
            sqft=n["sqft"], bedrooms=n["bedrooms"], bathrooms=n["bathrooms"],
            year_built=n["year_built"], lot_sqft=n["lot_sqft"],
            property_type=n["property_type"],
            detail_url=detail_url,
            sold_price=n["sold_price"], sold_date=n["sold_date"],
            days_on_market=n["days_on_market"], garage_spaces=n["garage_spaces"],
            source="zillow",
        )
        comp.ppsf = round(comp.sold_price / comp.sqft, 2) if comp.sqft else 0.0
        if not comp.sqft or comp.ppsf < 30:
            continue
        comps.append(comp)
    return comps


def _fetch_redfin_sold_pool(subject: SubjectProperty, radius_miles: float,
                            months_back: int) -> list[CompProperty]:
    """Inner: pull SOLD listings from Redfin via the CSV API."""
    try:
        from redfin_pending_scraper import fetch_sold_comps_via_redfin
    except ImportError:
        return []
    raw = fetch_sold_comps_via_redfin(
        zip_code=subject.zip_code,
        subject_lat=subject.latitude,
        subject_lon=subject.longitude,
        radius_miles=radius_miles,
        months_back=months_back,
        max_results=50,
    )
    comps = []
    for r in raw:
        if r["address"].lower().strip() == subject.address.lower().strip():
            continue  # subject itself
        comp = CompProperty(
            address=r["address"], city=r["city"], state=r.get("state", subject.state),
            zip_code=r["zip_code"],
            latitude=r["latitude"], longitude=r["longitude"],
            distance_miles=r["distance_miles"],
            sqft=r["sqft"], bedrooms=r["bedrooms"], bathrooms=r["bathrooms"],
            year_built=r["year_built"], lot_sqft=0,
            property_type=r.get("property_type", ""),
            detail_url=r["detail_url"],
            sold_price=r["sold_price"], sold_date=r["sold_date"],
            days_on_market=0, garage_spaces=0,
            source="redfin",
            mls_number=r.get("mls_number", ""),
        )
        comp.ppsf = r["ppsf"] or (round(r["sold_price"] / r["sqft"], 2) if r["sqft"] else 0.0)
        if not comp.sqft or comp.ppsf < 30:
            continue
        comps.append(comp)
    return comps


def _normalize_address_key(addr: str) -> str:
    """Stable dedup key — strips suffix variants, casing, punctuation."""
    s = (addr or "").lower().strip()
    # Drop common suffix variations
    for long, short in [
        (" street", " st"), (" drive", " dr"), (" road", " rd"),
        (" avenue", " ave"), (" boulevard", " blvd"), (" lane", " ln"),
        (" circle", " cir"), (" court", " ct"), (" place", " pl"),
        (" terrace", " ter"), (" parkway", " pkwy"), (" highway", " hwy"),
    ]:
        s = s.replace(long, short)
    # Strip punctuation + collapse whitespace
    s = re.sub(r"[.,#]", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _merge_dedupe_comps(zillow_comps: list[CompProperty],
                        redfin_comps: list[CompProperty]) -> list[CompProperty]:
    """Combine + dedupe Zillow and Redfin sold pools.

    Match rule: normalized street address. When the same property exists in
    both sources, prefer the Redfin record (MLS-direct = more reliable sold
    price) but tag source as "zillow+redfin" so the report shows cross-
    confirmation.
    """
    by_key: dict[str, CompProperty] = {}
    for c in zillow_comps:
        key = _normalize_address_key(c.address)
        if key:
            by_key[key] = c
    for c in redfin_comps:
        key = _normalize_address_key(c.address)
        if not key:
            continue
        if key in by_key:
            # Cross-confirmed comp — prefer Redfin's MLS price but flag both sources
            existing = by_key[key]
            c.source = "zillow+redfin"
            # Preserve Zillow's detail_url if Redfin's is missing (Zillow URLs are reliable)
            if not c.detail_url and existing.detail_url:
                c.detail_url = existing.detail_url
            # Preserve year_built if one source has it and the other doesn't
            if not c.year_built and existing.year_built:
                c.year_built = existing.year_built
            by_key[key] = c
        else:
            by_key[key] = c
    return list(by_key.values())


def fetch_comparable_sales(subject: SubjectProperty, radius_miles: float = DEFAULT_RADIUS_MILES,
                           months_back: int = DEFAULT_MONTHS_BACK,
                           api_key: str = "") -> list[CompProperty]:
    """Fetch comparable sold properties from BOTH Zillow + Redfin with tiered expansion.

    Strategy (revised 2026-05-31):
      1. Pull RECENTLY_SOLD from Zillow (OpenWeb Ninja) at requested radius/window
      2. Pull SOLD from Redfin CSV API (via Webshare US residential proxy) at same window
      3. Dedupe by normalized address → cross-confirmed comps tagged as both sources
      4. If combined pool < MIN_COMPS, expand radius (1mi → 2mi → 3mi)
      5. If still thin, expand timeframe (6mo → 12mo → 18mo)
      6. Return all comps (similarity ranking + ARV-top-5 selection happen downstream)

    Proximity is the heaviest single weight in similarity scoring, so starting
    with a tight 1mi/6mo window and expanding only when necessary keeps the
    candidate pool tightly clustered around the subject.
    """
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key:
        return []

    # Tier 1: requested radius / months (default 1mi / 6mo)
    pool = _dual_source_fetch(subject, radius_miles, months_back, api_key)
    if len(pool) >= TARGET_COMPS:
        return pool

    # Tier 2: expand radius (cap at MAX_RADIUS_MILES), same timeframe
    if radius_miles < MAX_RADIUS_MILES:
        expanded_radius = min(radius_miles * 2.0, MAX_RADIUS_MILES)
        logger.info("Pool size %d < %d — expanding radius %.1f → %.1f mi",
                    len(pool), TARGET_COMPS, radius_miles, expanded_radius)
        pool = _dual_source_fetch(subject, expanded_radius, months_back, api_key)
        if len(pool) >= TARGET_COMPS:
            return pool
        radius_miles = expanded_radius

    # Tier 3: extend timeframe (cap at MAX_MONTHS_BACK)
    if months_back < MAX_MONTHS_BACK:
        expanded_months = min(months_back * 2, MAX_MONTHS_BACK)
        logger.info("Pool size %d still < %d — expanding lookback %d → %d months",
                    len(pool), TARGET_COMPS, months_back, expanded_months)
        pool = _dual_source_fetch(subject, radius_miles, expanded_months, api_key)

    return pool


def _dual_source_fetch(subject: SubjectProperty, radius_miles: float,
                       months_back: int, api_key: str) -> list[CompProperty]:
    """Pull from both Zillow + Redfin at a given radius/window and dedupe."""
    zillow_comps = _fetch_zillow_sold_pool(subject, radius_miles, months_back, api_key)
    redfin_comps = _fetch_redfin_sold_pool(subject, radius_miles, months_back)
    merged = _merge_dedupe_comps(zillow_comps, redfin_comps)
    z_only = sum(1 for c in merged if c.source == "zillow")
    r_only = sum(1 for c in merged if c.source == "redfin")
    both = sum(1 for c in merged if c.source == "zillow+redfin")
    logger.info("Dual-source pool @ %.1fmi/%dmo: %d total (%d Zillow-only, %d Redfin-only, %d cross-confirmed)",
                radius_miles, months_back, len(merged), z_only, r_only, both)
    return merged


def fetch_pending_sales(subject: SubjectProperty, radius_miles: float = DEFAULT_RADIUS_MILES,
                        api_key: str = "") -> list[dict]:
    """Fetch under-contract / pending / contingent listings near subject.

    Three-tier source strategy:
      1. **Preferred — Redfin CSV API** (via Webshare US residential proxy).
         Returns true Pending + Contingent listings with MLS-direct status.
         Most reliable: Redfin's anti-bot is light enough that residential
         proxy traffic flows through with no CAPTCHA. Requires
         `ZILLOW_PROXY_URL` in .env.
      2. **Tier 2 — Zillow scrape** (Playwright + Webshare). Same proxy. Often
         blocked by PerimeterX even with residential IPs, so rarely succeeds
         in practice — kept for the rare case it works.
      3. **Tier 3 — OpenWeb Ninja /search FOR_SALE sorted by DOM ASC**.
         Approximation only — fresh active listings, NOT true pending. Used
         when both scrapes fail or no proxy is configured.

    Returns dicts (not CompProperty — list price isn't a sold price) with:
        address, list_price, sqft, bedrooms, bathrooms, year_built,
        distance_miles, days_on_market, detail_url, status
    """
    # ── Tier 1: Redfin CSV API via Webshare residential proxy ──
    if subject.zip_code and config.ZILLOW_PROXY_URL:
        try:
            from redfin_pending_scraper import fetch_pending_via_redfin
            scraped = fetch_pending_via_redfin(
                zip_code=subject.zip_code,
                subject_lat=subject.latitude,
                subject_lon=subject.longitude,
                radius_miles=max(radius_miles, 2.0),
                max_results=6,
            )
            if scraped:
                logger.info("Got %d Pending/Contingent listings via Redfin for ZIP %s",
                            len(scraped), subject.zip_code)
                return scraped
            logger.info("Redfin returned 0 pendings for ZIP %s — falling back to Zillow scrape",
                        subject.zip_code)
        except Exception as e:
            logger.warning("Redfin scrape failed (%s) — falling back to Zillow scrape", e)

    # ── Tier 2: Zillow scrape via Playwright (often blocked by PerimeterX) ──
    if subject.zip_code and config.ZILLOW_PROXY_URL:
        try:
            from zillow_pending_scraper import fetch_pending_via_zillow
            scraped = fetch_pending_via_zillow(
                zip_code=subject.zip_code,
                subject_lat=subject.latitude,
                subject_lon=subject.longitude,
                radius_miles=max(radius_miles, 2.0),
                max_results=6,
            )
            if scraped:
                logger.info("Got %d listings via Zillow scrape for ZIP %s",
                            len(scraped), subject.zip_code)
                return scraped
            logger.info("Zillow scrape returned 0 — falling back to FOR_SALE fresh-listings proxy")
        except Exception as e:
            logger.warning("Zillow pending scrape failed (%s) — falling back to FOR_SALE proxy", e)

    # ── Tier 3: FOR_SALE fresh-listings proxy via OpenWeb Ninja ──
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key or not subject.zip_code:
        return []
    data = _api_get(
        SEARCH_ENDPOINT,
        {"location": subject.zip_code, "home_status": "FOR_SALE"},
        api_key,
    )
    items = data if isinstance(data, list) else []
    logger.info("/search returned %d FOR_SALE records for ZIP %s (fallback proxy for pendings)",
                len(items), subject.zip_code)

    def _parse(item):
        if not isinstance(item, dict):
            return None
        price = item.get("unformattedPrice") or item.get("price") or 0
        sqft = item.get("livingArea") or item.get("area") or 0
        if not price or not sqft or price > 50_000_000:
            return None
        addr_obj = item.get("address") or {}
        if isinstance(addr_obj, dict):
            street = addr_obj.get("streetAddress") or ""
            city = addr_obj.get("city") or ""
            zipc = addr_obj.get("zipcode") or ""
        else:
            street = item.get("streetAddress") or ""
            city = ""
            zipc = ""
        lat = float(item.get("latitude") or 0)
        lon = float(item.get("longitude") or 0)
        dist = 0.0
        if subject.latitude and subject.longitude and lat and lon:
            dist = _haversine_miles(subject.latitude, subject.longitude, lat, lon)
            if dist > radius_miles:
                return None
        if street and street.lower().strip() == subject.address.lower().strip():
            return None
        # Convert daysOnZillow from ms-since-epoch or int days (API mixes formats)
        dom_raw = item.get("daysOnZillow")
        if isinstance(dom_raw, (int, float)):
            dom = int(dom_raw) if dom_raw < 100000 else int(dom_raw / 86_400_000)
        else:
            dom = 0
        return {
            "address": street,
            "city": city,
            "zip_code": zipc,
            "list_price": int(price),
            "sqft": int(sqft),
            "bedrooms": item.get("bedrooms") or item.get("beds") or 0,
            "bathrooms": item.get("bathrooms") or item.get("baths") or 0,
            "year_built": int(item.get("yearBuilt") or 0),
            "ppsf": round(price / sqft, 2) if sqft else 0.0,
            "distance_miles": round(dist, 2),
            "days_on_market": dom,
            "detail_url": item.get("detailUrl") or "",
            "status": item.get("statusText") or "For sale",
        }

    pendings = [p for p in (_parse(item) for item in items) if p]
    # Auto-expand if sparse — fresh listings are rarer than mature listings
    if len(pendings) < 3 and radius_miles < 2.0:
        for try_radius in (1.0, 2.0):
            if try_radius <= radius_miles:
                continue
            wider = []
            for item in items:
                p = _parse(item)
                if p and p["distance_miles"] <= try_radius:
                    wider.append(p)
            if len(wider) >= 3 or try_radius >= 2.0:
                pendings = wider
                radius_miles = try_radius
                break
    # Sort by days-on-market ASC — freshest listings first (highest probability of
    # going pending soon). Then by proximity within the same DOM bucket.
    pendings.sort(key=lambda p: (p["days_on_market"], p["distance_miles"]))
    time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))
    logger.info("Kept %d active listings within %.1f mi (fresh-listing proxy for pending)",
                len(pendings), radius_miles)
    return pendings[:6]


def _enrich_comp_year_built(comps: list[CompProperty], api_key: str = "") -> None:
    """Backfill comp.year_built via property-details-address (one call per comp).

    /search doesn't return yearBuilt, but /property-details-address does. Mutates
    comps in place. Silently leaves year_built at 0 if a call fails (so the rest
    of the analysis still works). Brief delay between calls to avoid rate-limit.
    """
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key:
        return
    for comp in comps:
        if comp.year_built or not comp.address:
            continue  # already have it (or no address to look up)
        full_addr = f"{comp.address} {comp.city} {comp.state} {comp.zip_code}".strip()
        data = _api_get(PROPERTY_ENDPOINT, {"address": full_addr}, api_key)
        if isinstance(data, dict) and data.get("yearBuilt"):
            try:
                comp.year_built = int(data["yearBuilt"])
            except (ValueError, TypeError):
                pass
        time.sleep(random.uniform(REQUEST_DELAY_MIN, REQUEST_DELAY_MAX))


def fetch_rental_comps(subject: SubjectProperty, radius_miles: float = 1.0,
                       n: int = 6, api_key: str = "") -> list[dict]:
    """Fetch active rental listings near subject for Monthly Rent validation.

    Auto-expands radius (1mi → 3mi → 5mi) if fewer than 3 comps found at the
    starting radius, since rentals are sparser than sales in suburban markets.

    Returns a list of dicts (not CompProperty — different shape) with:
        address, rent, sqft, bedrooms, bathrooms, distance_miles, detail_url

    Sorted by closeness to subject's sqft (most similar size first). Filters
    null/zero rent + records without sqft (often apartment-complex aggregates).
    """
    api_key = api_key or config.OPENWEBNINJA_API_KEY
    if not api_key or not subject.zip_code:
        return []
    data = _api_get(
        SEARCH_ENDPOINT,
        {"location": subject.zip_code, "home_status": "FOR_RENT"},
        api_key,
    )
    items = data if isinstance(data, list) else []

    def _parse(item, max_radius):
        if not isinstance(item, dict):
            return None
        rent = item.get("unformattedPrice") or item.get("price") or 0
        sqft = item.get("livingArea") or item.get("area") or 0
        if not rent or not sqft:
            return None
        if rent > 50000:  # sale price leak
            return None
        addr = item.get("address") or {}
        if isinstance(addr, dict):
            street = addr.get("streetAddress") or ""
        else:
            street = item.get("streetAddress") or str(addr) or ""
        lat = float(item.get("latitude") or 0)
        lon = float(item.get("longitude") or 0)
        dist = 0.0
        if subject.latitude and subject.longitude and lat and lon:
            dist = _haversine_miles(subject.latitude, subject.longitude, lat, lon)
            if dist > max_radius:
                return None
        return {
            "address": street,
            "rent": rent,
            "sqft": sqft,
            "bedrooms": item.get("bedrooms") or item.get("beds") or 0,
            "bathrooms": item.get("bathrooms") or item.get("baths") or 0,
            "distance_miles": round(dist, 2),
            "detail_url": item.get("detailUrl") or "",
        }

    # Auto-expand if first attempt yields too few
    rentals = []
    for try_radius in [radius_miles, 3.0, 5.0]:
        if try_radius < radius_miles:
            continue
        rentals = [r for r in (_parse(item, try_radius) for item in items) if r]
        if len(rentals) >= 3 or try_radius >= 5.0:
            radius_miles = try_radius
            break

    if subject.sqft:
        rentals.sort(key=lambda r: abs(r["sqft"] - subject.sqft))
    logger.info("Fetched %d rental comps within %.1f mi (subject ZIP %s)",
                len(rentals), radius_miles, subject.zip_code)
    return rentals[:n]


# ── Similarity scoring ────────────────────────────────────────────────

def _score_similarity(subject: SubjectProperty, comp: CompProperty,
                       pool_median_ppsf: float = 0.0,
                       as_is: bool = False) -> float:
    """Score comp similarity (0.0-1.0+, higher = more similar).

    Priority order (revised 2026-06-09 for fix-n-flip ARV accuracy):
      1. SOLD DATE   (most important — up to -0.30 deduction)
      2. PROXIMITY   (up to -0.25)
      3. PPSF TIER   (-0.10 to +0.15) — NEW: bias toward renovated comps
                     (high-PPSF) since they reflect post-rehab sale price in
                     fix-n-flip markets. Distressed (low-PPSF) penalized.
      4. SQFT        (up to -0.15)
      5. YEAR BUILT  (up to -0.10) — SOFTENED: ±10yr no penalty (typical
                     market variance for housing stock vintage).
      6. Bedrooms / Bathrooms / Property type (smaller deductions)

    Higher-priority factors get larger possible deductions/bonuses, so they
    push less-similar comps further down the ranking. The top ARV_AVG_TOP_N
    by score get averaged for ARV.

    pool_median_ppsf — pre-computed median PPSF across the candidate pool.
    Pass 0 (default) to skip the PPSF tier adjustment (used when pool
    statistics aren't available yet).
    """
    score = 1.0

    # ── 1. SOLD DATE (most important) — max -0.30 ──
    if comp.sold_date:
        try:
            sold_dt = datetime.strptime(comp.sold_date[:10], "%Y-%m-%d")
            days_ago = (datetime.now() - sold_dt).days
            if days_ago < 30:
                score -= 0.00     # fresh
            elif days_ago < 90:
                score -= 0.05
            elif days_ago < 180:
                score -= 0.15
            elif days_ago < 365:
                score -= 0.25
            else:
                score -= 0.30     # stale
        except ValueError:
            score -= 0.15  # bad date = treat as moderately stale
    else:
        score -= 0.15

    # ── 2. PROXIMITY (distance from subject) — max -0.25 ──
    if comp.distance_miles <= 0.5:
        score -= 0.00
    elif comp.distance_miles <= 1.0:
        score -= 0.05
    elif comp.distance_miles <= 1.5:
        score -= 0.10
    elif comp.distance_miles <= 2.0:
        score -= 0.18
    elif comp.distance_miles <= 3.0:
        score -= 0.22
    else:
        score -= 0.25

    # ── 3. PPSF TIER (NEW 2026-06-09) — bias toward renovated comps ──
    # Subject's post-rehab ARV should match what other renovated houses in
    # the area sold for. In a fix-n-flip market this is the top-PPSF tier.
    # Conversely, deeply-distressed comps (tax sales, quitclaims, gut-needed
    # properties) are misleading — penalize them.
    # Tier thresholds:
    #   renovated  → PPSF > pool_median × 1.25 → +0.15 bonus  (ARV-relevant)
    #   distressed → PPSF < pool_median × 0.75 → -0.10 penalty
    #   standard   → in between → neutral
    # AS-IS MODE (2026-07-17): flip the bias — favor typical-condition comps
    # (within ±15% of median PPSF), penalize both renovated AND distressed
    # outliers. Use when subject is rent-ready and you're NOT renovating to
    # top-of-market. Post-rehab ARV logic doesn't apply.
    if pool_median_ppsf and comp.ppsf:
        ratio = comp.ppsf / pool_median_ppsf
        if as_is:
            # As-is: reward middle band, penalize outliers in BOTH directions
            if 0.85 <= ratio <= 1.15:
                score += 0.15
                comp.tier = "as-is (typical)"
            elif ratio > 1.25:
                score -= 0.10  # penalty — likely renovated / different product class
                comp.tier = "renovated (not comparable to as-is)"
            elif ratio < 0.75:
                score -= 0.10  # penalty — distressed / not representative
                comp.tier = "distressed"
            else:
                comp.tier = "standard"
        else:
            # Fix-n-flip (default): reward top-quartile (renovated = post-rehab target)
            if ratio > 1.25:
                score += 0.15
                comp.tier = "renovated"
            elif ratio < 0.75:
                score -= 0.10
                comp.tier = "distressed"
            else:
                comp.tier = "standard"

    # ── 4. SQFT — max -0.15 ──
    if subject.sqft and comp.sqft:
        sqft_diff_pct = abs(subject.sqft - comp.sqft) / subject.sqft
        if sqft_diff_pct <= 0.10:
            score -= 0.00
        elif sqft_diff_pct <= 0.20:
            score -= 0.05
        elif sqft_diff_pct <= 0.30:
            score -= 0.10
        else:
            score -= 0.15

    # ── 5. YEAR BUILT (SOFTENED 2026-06-09) — max -0.10, ±10yr no penalty ──
    # Real-estate convention: a 10-year vintage delta is functionally
    # negligible for housing stock built post-WWII. Drop the harsh penalty.
    if subject.year_built and comp.year_built:
        age_diff = abs(subject.year_built - comp.year_built)
        if age_diff <= 10:
            score -= 0.00       # within typical market variance
        elif age_diff <= 20:
            score -= 0.03
        elif age_diff <= 30:
            score -= 0.06
        else:
            score -= 0.10
    # else: missing year_built data → neutral (no deduction)

    # ── 6. Secondary factors (smaller weights) ──
    # Bedrooms — max -0.05
    if subject.bedrooms and comp.bedrooms:
        bed_diff = abs(subject.bedrooms - comp.bedrooms)
        if bed_diff >= 2:
            score -= 0.05
        elif bed_diff == 1:
            score -= 0.02

    # Bathrooms — max -0.05
    if subject.bathrooms and comp.bathrooms:
        bath_diff = abs(subject.bathrooms - comp.bathrooms)
        if bath_diff > 1:
            score -= 0.05
        elif bath_diff >= 0.5:
            score -= 0.02

    # Property type mismatch — max -0.10
    if subject.property_type and comp.property_type:
        if subject.property_type.upper() != comp.property_type.upper():
            score -= 0.10

    # Allow score > 1.0 — renovated comps with +0.15 PPSF bonus can exceed
    # 1.0 and that's a useful signal for sorting + display ("this is the
    # most ARV-relevant comp in the pool"). Floor at 0.0 only.
    return max(0.0, score)


# ── Adjustment engine ─────────────────────────────────────────────────

def _calculate_adjustments(subject: SubjectProperty, comp: CompProperty) -> dict:
    """Calculate dollar adjustments from comp to subject property."""
    adjustments = {}

    # Square footage adjustment
    if subject.sqft and comp.sqft:
        sqft_diff = subject.sqft - comp.sqft
        if sqft_diff != 0:
            adj = sqft_diff * ADJ_PER_SQFT
            adjustments["sqft"] = round(adj)

    # Bedroom adjustment
    if subject.bedrooms and comp.bedrooms:
        bed_diff = subject.bedrooms - comp.bedrooms
        if bed_diff != 0:
            adjustments["bedrooms"] = round(bed_diff * ADJ_PER_BEDROOM)

    # Bathroom adjustment
    if subject.bathrooms and comp.bathrooms:
        bath_diff = subject.bathrooms - comp.bathrooms
        if bath_diff != 0:
            adjustments["bathrooms"] = round(bath_diff * ADJ_PER_BATHROOM)

    # Age / year built adjustment
    if subject.year_built and comp.year_built:
        year_diff = subject.year_built - comp.year_built
        if year_diff != 0:
            adjustments["year_built"] = round(year_diff * ADJ_PER_YEAR_BUILT)

    # Lot size adjustment (capped)
    if subject.lot_sqft and comp.lot_sqft:
        lot_diff = subject.lot_sqft - comp.lot_sqft
        if lot_diff != 0:
            adj = lot_diff * ADJ_PER_LOT_SQFT
            adj = max(-ADJ_LOT_MAX, min(ADJ_LOT_MAX, adj))
            adjustments["lot_size"] = round(adj)

    # Garage adjustment
    garage_diff = subject.garage_spaces - comp.garage_spaces
    if garage_diff != 0:
        adjustments["garage"] = round(garage_diff * ADJ_PER_GARAGE)

    # Market conditions (time) adjustment
    if comp.sold_date:
        try:
            sold_dt = datetime.strptime(comp.sold_date[:10], "%Y-%m-%d")
            months_ago = (datetime.now() - sold_dt).days / 30.0
            if months_ago > 1:
                adj = comp.sold_price * MARKET_CONDITION_PCT_PER_MONTH * months_ago
                adjustments["market_conditions"] = round(adj)
        except ValueError:
            pass

    return adjustments


def _apply_adjustments(comp: CompProperty, adjustments: dict) -> float:
    """Apply adjustments to comp's sold price and return adjusted price."""
    total_adj = sum(adjustments.values())
    return comp.sold_price + total_adj


# ── Two-Bucket classification ────────────────────────────────────────

def _classify_bucket(comp: CompProperty) -> str:
    """Classify comp into Bucket A or Bucket B.

    Bucket A: Non-disclosure baseline comps — properties with limited price
              transparency (typical in TN as a non-disclosure state).
              Uses Zillow/MLS-reported data as proxy.
    Bucket B: Disclosure/verified comps — properties with confirmed sale
              prices from MLS (listed and sold through agent).

    In practice for TN (non-disclosure state), all comps come through
    Zillow/MLS so we classify based on data completeness:
    - Bucket B: Has complete MLS data (sold through agent, DOM tracked)
    - Bucket A: Limited data (off-market sale, FSBO, etc.)
    """
    if comp.days_on_market > 0:
        return "B"  # Was listed on MLS — has disclosure data
    return "A"  # No DOM data — likely off-market/non-disclosure sale


# ── ARV calculation ───────────────────────────────────────────────────

def calculate_arv(subject: SubjectProperty, comps: list[CompProperty],
                  as_is: bool = False, max_comp_year: int = None) -> ARVResult:
    """Calculate ARV using Top-N-by-Similarity mean of RAW SOLD prices (revised 2026-05-31).

    Methodology:
      1. Score + rank ALL comps by similarity (sold-date > proximity > year > sqft)
      2. Display top TARGET_COMPS (=10) on the report
      3. ARV = average of SOLD prices of the top ARV_AVG_TOP_N (=5) most similar comps
         → replaces prior "middle-3 trimmed mean" — similarity ranking is now
           strong enough (dual-source + tight 1mi/6mo window) that price-trim
           is no longer needed
      4. Property-specific adjustments still computed for reference column
      5. Confidence bands from price spread WITHIN the top-N used for ARV

    as_is: When True, flip PPSF-tier bias in _score_similarity to favor typical-
    condition comps (median PPSF band) instead of top-quartile renovated comps.
    Use for rent-ready deals where subject is habitable and you're NOT
    renovating to top-of-market. Default False (fix-n-flip renovation-driven ARV).

    max_comp_year: When set, hard-cap comps by build year. Applied AFTER
    year_built enrichment (Zillow /search doesn't return year_built, so we
    can't filter until we enrich). Enriches 2× the normal pool so post-filter
    we still have ~10 comps for scoring. Use when subject is a resale and
    nearby new-construction is inflating the pool with builder-premium comps.
    """
    if not comps:
        return ARVResult(confidence="none", confidence_reason="No comparable sales found")

    # Compute pool median PPSF for tier classification (2026-06-09).
    # Drives the "renovated +0.15 / distressed -0.10" bonus in _score_similarity.
    pool_ppsf_values = [c.ppsf for c in comps if c.ppsf > 0]
    pool_median_ppsf = (
        sorted(pool_ppsf_values)[len(pool_ppsf_values) // 2]
        if pool_ppsf_values else 0.0
    )

    # Score and sort by similarity (with PPSF tier bonus applied)
    for comp in comps:
        comp.similarity_score = _score_similarity(subject, comp, pool_median_ppsf, as_is=as_is)
    comps.sort(key=lambda c: c.similarity_score, reverse=True)

    # Enrichment pool: 2× normal size when max_comp_year is set so post-filter
    # we still have ~10 comps. Otherwise stay at MAX_COMPS (~10s enrichment cost).
    enrich_size = MAX_COMPS * 2 if max_comp_year is not None else MAX_COMPS
    enrich_pool = comps[:enrich_size]

    # Enrich with yearBuilt (/search doesn't include it — needs an extra call
    # per comp to /property-details-address). Adds ~10s but lets the Comp
    # Analysis tab show comp age, which matters for visual comp-vetting.
    _enrich_comp_year_built(enrich_pool)

    # ── Optional new-construction filter (2026-07-17) ──
    # Drops comps built after max_comp_year. Comps with unknown year_built
    # are KEPT (can't confirm they're new construction — visual review on
    # the report catches them). If filter over-drops, warn.
    if max_comp_year is not None:
        before = len(enrich_pool)
        enrich_pool = [c for c in enrich_pool if not c.year_built or c.year_built <= max_comp_year]
        dropped = before - len(enrich_pool)
        if dropped:
            logger.info("max_comp_year=%d filter dropped %d/%d comps (built after cap)",
                        max_comp_year, dropped, before)
        if len(enrich_pool) < MIN_COMPS:
            logger.warning("Post-filter comp count %d below minimum %d — consider widening --radius/--months",
                           len(enrich_pool), MIN_COMPS)

    # Re-score after year_built enrichment (year_built now factors in) + take top 10
    for comp in enrich_pool:
        comp.similarity_score = _score_similarity(subject, comp, pool_median_ppsf, as_is=as_is)
    enrich_pool.sort(key=lambda c: c.similarity_score, reverse=True)
    selected = enrich_pool[:MAX_COMPS]  # top 10 for display

    # Reflect final ordering + max_comp_year filter in the caller's list so
    # downstream (pkg.comps, vintage sanity check in deal_analyzer) sees the
    # SAME pool that drove ARV — not the raw pre-filter pool. Anything
    # dropped by the filter is discarded here too.
    comps[:] = enrich_pool

    # Apply adjustments + bucket classification (adjustments kept for reference column)
    for comp in selected:
        comp.bucket = _classify_bucket(comp)
        comp.adjustments = _calculate_adjustments(subject, comp)
        comp.adjusted_price = _apply_adjustments(comp, comp.adjustments)

    bucket_a = [c for c in selected if c.bucket == "A"]
    bucket_b = [c for c in selected if c.bucket == "B"]

    # PPSF values for reporting — uses sold_price (matches ARV basis)
    ppsf_values = [comp.sold_price / comp.sqft for comp in selected if comp.sqft]

    # ── ARV = mean of SOLD prices of top ARV_AVG_TOP_N most similar comps ──
    top_for_arv = [c for c in selected[:ARV_AVG_TOP_N] if c.sold_price > 0]
    if not top_for_arv:
        return ARVResult(confidence="none", confidence_reason="No valid sold prices")

    arv_prices = [c.sold_price for c in top_for_arv]
    arv_mid = sum(arv_prices) / len(arv_prices)

    # Confidence bands based on spread WITHIN the top-N used for ARV
    if len(arv_prices) > 1:
        spread = max(arv_prices) - min(arv_prices)
        spread_pct = (spread / arv_mid * 100) if arv_mid else 0
    else:
        spread = 0
        spread_pct = 0

    # Conservative bands — low end is intentionally more conservative
    if spread_pct < 10:
        arv_low = arv_mid * 0.95
        arv_high = arv_mid * 1.05
        confidence = "high"
        confidence_reason = f"Tight comp spread ({spread_pct:.0f}%), {len(selected)} comps"
    elif spread_pct < 20:
        arv_low = arv_mid * 0.90
        arv_high = arv_mid * 1.08
        confidence = "medium"
        confidence_reason = f"Moderate comp spread ({spread_pct:.0f}%), {len(selected)} comps"
    else:
        arv_low = arv_mid * 0.85
        arv_high = arv_mid * 1.10
        confidence = "low"
        confidence_reason = f"Wide comp spread ({spread_pct:.0f}%) — verify with local knowledge"

    # Fewer comps = lower confidence
    if len(selected) < MIN_COMPS:
        confidence = "low"
        confidence_reason = f"Only {len(selected)} comps found (minimum {MIN_COMPS} recommended)"

    avg_adj = sum(abs(sum(c.adjustments.values())) for c in selected) / len(selected) if selected else 0

    return ARVResult(
        arv_low=round(arv_low),
        arv_mid=round(arv_mid),
        arv_high=round(arv_high),
        confidence=confidence,
        confidence_reason=confidence_reason,
        ppsf_avg=round(sum(ppsf_values) / len(ppsf_values), 2) if ppsf_values else 0,
        ppsf_range=(round(min(ppsf_values), 2), round(max(ppsf_values), 2)) if ppsf_values else (0, 0),
        comp_count=len(selected),
        bucket_a_count=len(bucket_a),
        bucket_b_count=len(bucket_b),
        avg_adjustment=round(avg_adj),
        spread_pct=round(spread_pct, 1),
    )


# ── Excel report generation ──────────────────────────────────────────

# Styles
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_TITLE_FONT = Font(name="Calibri", bold=True, size=16, color="2F5496")
_SUBTITLE_FONT = Font(name="Calibri", bold=True, size=12, color="333333")
_MONEY_FMT = '#,##0'
_PCT_FMT = '0.0%'
_LABEL_FONT = Font(name="Calibri", size=11, color="555555")
_VALUE_FONT = Font(name="Calibri", bold=True, size=13, color="222222")
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_GREEN_FONT = Font(name="Calibri", bold=True, color="006100")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    """Write a styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _auto_column_widths(ws, min_width: int = 12, max_width: int = 35) -> None:
    """Auto-size columns based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def _fmt_money(value: float) -> str:
    """Format number as currency string."""
    if not value:
        return "$0"
    return f"${value:,.0f}"


def generate_comp_report(subject: SubjectProperty, comps: list[CompProperty],
                         arv: ARVResult, output_path: str) -> str:
    """Generate a 7-tab Excel workbook comp report.

    Tabs:
    1. Executive Summary — subject property + ARV range
    2. Subject Property — full detail
    3. Comparable Sales — all comps with distance, date, similarity
    4. Adjustments Detail — per-comp adjustment breakdown
    5. Market Analysis — PPSF trends, DOM, market direction
    6. ARV Calculation — Two-Bucket weighted result with confidence bands
    7. Sources & Notes
    """
    wb = Workbook()

    # ── Tab 1: Executive Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"

    ws.cell(row=1, column=1, value="Comp Analysis Report").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=f"{subject.address}, {subject.city}, {subject.state} {subject.zip_code}").font = _SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = _LABEL_FONT

    row = 5
    summary_data = [
        ("ARV (Low)", _fmt_money(arv.arv_low)),
        ("ARV (Mid — Recommended)", _fmt_money(arv.arv_mid)),
        ("ARV (High)", _fmt_money(arv.arv_high)),
        ("", ""),
        ("Confidence Level", arv.confidence.upper()),
        ("Confidence Reason", arv.confidence_reason),
        ("", ""),
        ("Avg PPSF", f"${arv.ppsf_avg:,.2f}"),
        ("PPSF Range", f"${arv.ppsf_range[0]:,.2f} — ${arv.ppsf_range[1]:,.2f}"),
        ("Comps Analyzed", str(arv.comp_count)),
        ("Bucket A (Non-Disclosure)", str(arv.bucket_a_count)),
        ("Bucket B (Disclosure/MLS)", str(arv.bucket_b_count)),
        ("Avg Gross Adjustment", _fmt_money(arv.avg_adjustment)),
        ("Comp Spread", f"{arv.spread_pct:.1f}%"),
        ("", ""),
        ("Subject Zestimate", _fmt_money(subject.zestimate)),
        ("Subject Property Type", subject.property_type),
        ("Subject Sqft", f"{subject.sqft:,}" if subject.sqft else "N/A"),
        ("Subject Bed/Bath", f"{subject.bedrooms}bd / {subject.bathrooms}ba"),
        ("Subject Year Built", str(subject.year_built) if subject.year_built else "N/A"),
    ]
    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = _VALUE_FONT
        if label == "Confidence Level":
            if arv.confidence == "high":
                cell.fill = _GREEN_FILL
                cell.font = _GREEN_FONT
            elif arv.confidence == "medium":
                cell.fill = _YELLOW_FILL
            else:
                cell.fill = _RED_FILL
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35

    # ── Tab 2: Subject Property ──────────────────────────────────────
    ws2 = wb.create_sheet("Subject Property")
    ws2.cell(row=1, column=1, value="Subject Property Details").font = _TITLE_FONT

    props = [
        ("Address", subject.address),
        ("City", subject.city),
        ("State", subject.state),
        ("ZIP", subject.zip_code),
        ("Latitude", str(subject.latitude) if subject.latitude else ""),
        ("Longitude", str(subject.longitude) if subject.longitude else ""),
        ("Property Type", subject.property_type),
        ("Square Feet", f"{subject.sqft:,}" if subject.sqft else ""),
        ("Bedrooms", str(subject.bedrooms)),
        ("Bathrooms", str(subject.bathrooms)),
        ("Year Built", str(subject.year_built) if subject.year_built else ""),
        ("Lot Size (sqft)", f"{subject.lot_sqft:,}" if subject.lot_sqft else ""),
        ("Garage Spaces", str(subject.garage_spaces)),
        ("Zestimate", _fmt_money(subject.zestimate)),
        ("MLS Status", subject.mls_status),
        ("Last Sold Date", subject.last_sold_date),
        ("Last Sold Price", _fmt_money(subject.last_sold_price)),
    ]
    for i, (label, value) in enumerate(props, 3):
        ws2.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws2.cell(row=i, column=2, value=value).font = _VALUE_FONT
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 30

    # ── Tab 3: Comparable Sales ──────────────────────────────────────
    ws3 = wb.create_sheet("Comparable Sales")
    ws3.cell(row=1, column=1, value="Comparable Sales").font = _TITLE_FONT

    comp_headers = ["#", "Address", "City", "ZIP", "Distance (mi)", "Sold Price",
                    "Sold Date", "Sqft", "Bed", "Bath", "Year Built", "PPSF",
                    "Similarity", "Bucket", "Adjusted Price"]
    _write_header_row(ws3, 3, comp_headers)

    for i, comp in enumerate(comps[:MAX_COMPS], 1):
        row = i + 3
        values = [
            i, comp.address, comp.city, comp.zip_code,
            comp.distance_miles, comp.sold_price, comp.sold_date,
            comp.sqft, comp.bedrooms, comp.bathrooms, comp.year_built,
            round(comp.ppsf, 2), f"{comp.similarity_score:.0%}",
            comp.bucket, comp.adjusted_price,
        ]
        for col, val in enumerate(values, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            if col in (6, 15):  # Money columns
                cell.number_format = _MONEY_FMT
            cell.border = _THIN_BORDER

    _auto_column_widths(ws3)

    # ── Tab 4: Adjustments Detail ────────────────────────────────────
    ws4 = wb.create_sheet("Adjustments Detail")
    ws4.cell(row=1, column=1, value="Per-Comp Adjustment Breakdown").font = _TITLE_FONT

    adj_types = ["sqft", "bedrooms", "bathrooms", "year_built", "lot_size", "garage", "market_conditions"]
    adj_headers = ["Comp #", "Address", "Sold Price"] + [a.replace("_", " ").title() for a in adj_types] + ["Total Adj", "Adjusted Price"]
    _write_header_row(ws4, 3, adj_headers)

    for i, comp in enumerate(comps[:MAX_COMPS], 1):
        row = i + 3
        values = [i, comp.address, comp.sold_price]
        total_adj = 0
        for adj_type in adj_types:
            adj_val = comp.adjustments.get(adj_type, 0)
            total_adj += adj_val
            values.append(adj_val)
        values.append(total_adj)
        values.append(comp.adjusted_price)
        for col, val in enumerate(values, 1):
            cell = ws4.cell(row=row, column=col, value=val)
            if col >= 3:
                cell.number_format = _MONEY_FMT
            cell.border = _THIN_BORDER

    _auto_column_widths(ws4)

    # ── Tab 5: Market Analysis ───────────────────────────────────────
    ws5 = wb.create_sheet("Market Analysis")
    ws5.cell(row=1, column=1, value="Market Analysis").font = _TITLE_FONT

    # PPSF analysis
    ws5.cell(row=3, column=1, value="Price Per Square Foot Analysis").font = _SUBTITLE_FONT
    ppsf_data = [
        ("Average PPSF", f"${arv.ppsf_avg:,.2f}"),
        ("PPSF Range", f"${arv.ppsf_range[0]:,.2f} — ${arv.ppsf_range[1]:,.2f}"),
        ("Subject Implied Value (Avg PPSF)", _fmt_money(arv.ppsf_avg * subject.sqft) if subject.sqft else "N/A"),
    ]
    for i, (label, value) in enumerate(ppsf_data, 4):
        ws5.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws5.cell(row=i, column=2, value=value).font = _VALUE_FONT

    # Days on market
    ws5.cell(row=8, column=1, value="Days on Market Analysis").font = _SUBTITLE_FONT
    dom_values = [c.days_on_market for c in comps[:MAX_COMPS] if c.days_on_market > 0]
    if dom_values:
        ws5.cell(row=9, column=1, value="Average DOM").font = _LABEL_FONT
        ws5.cell(row=9, column=2, value=f"{sum(dom_values) / len(dom_values):.0f} days").font = _VALUE_FONT
        ws5.cell(row=10, column=1, value="Median DOM").font = _LABEL_FONT
        sorted_dom = sorted(dom_values)
        median_dom = sorted_dom[len(sorted_dom) // 2]
        ws5.cell(row=10, column=2, value=f"{median_dom} days").font = _VALUE_FONT

    # Market direction
    ws5.cell(row=12, column=1, value="Market Direction").font = _SUBTITLE_FONT
    ws5.cell(row=13, column=1, value="Monthly Appreciation Rate").font = _LABEL_FONT
    ws5.cell(row=13, column=2, value=f"{MARKET_CONDITION_PCT_PER_MONTH * 100:.1f}%").font = _VALUE_FONT
    ws5.cell(row=14, column=1, value="Annualized Appreciation").font = _LABEL_FONT
    ws5.cell(row=14, column=2, value=f"{MARKET_CONDITION_PCT_PER_MONTH * 12 * 100:.1f}%").font = _VALUE_FONT

    ws5.column_dimensions["A"].width = 35
    ws5.column_dimensions["B"].width = 25

    # ── Tab 6: ARV Calculation ───────────────────────────────────────
    ws6 = wb.create_sheet("ARV Calculation")
    ws6.cell(row=1, column=1, value="Two-Bucket ARV Calculation").font = _TITLE_FONT

    ws6.cell(row=3, column=1, value="Methodology").font = _SUBTITLE_FONT
    ws6.cell(row=4, column=1, value="Bucket A (Non-Disclosure): 30% weight — Off-market/FSBO sales with limited price transparency").font = _LABEL_FONT
    ws6.cell(row=5, column=1, value="Bucket B (Disclosure/MLS): 70% weight — Agent-listed sales with confirmed pricing").font = _LABEL_FONT
    ws6.cell(row=6, column=1, value="Tennessee is a non-disclosure state. All data sourced via Zillow/MLS.").font = _LABEL_FONT

    ws6.cell(row=8, column=1, value="Bucket A Comps").font = _SUBTITLE_FONT
    bucket_a = [c for c in comps[:MAX_COMPS] if c.bucket == "A"]
    bucket_b = [c for c in comps[:MAX_COMPS] if c.bucket == "B"]
    if bucket_a:
        avg_a = sum(c.adjusted_price for c in bucket_a) / len(bucket_a)
        ws6.cell(row=9, column=1, value=f"Count: {len(bucket_a)}  |  Avg Adjusted: {_fmt_money(avg_a)}").font = _VALUE_FONT
    else:
        ws6.cell(row=9, column=1, value="No Bucket A comps").font = _LABEL_FONT

    ws6.cell(row=11, column=1, value="Bucket B Comps").font = _SUBTITLE_FONT
    if bucket_b:
        avg_b = sum(c.adjusted_price for c in bucket_b) / len(bucket_b)
        ws6.cell(row=12, column=1, value=f"Count: {len(bucket_b)}  |  Avg Adjusted: {_fmt_money(avg_b)}").font = _VALUE_FONT
    else:
        ws6.cell(row=12, column=1, value="No Bucket B comps").font = _LABEL_FONT

    ws6.cell(row=14, column=1, value="Final ARV").font = _SUBTITLE_FONT
    arv_display = [
        ("ARV Low (Conservative)", _fmt_money(arv.arv_low)),
        ("ARV Mid (Recommended)", _fmt_money(arv.arv_mid)),
        ("ARV High (Optimistic)", _fmt_money(arv.arv_high)),
        ("Confidence", arv.confidence.upper()),
        ("Reason", arv.confidence_reason),
    ]
    for i, (label, value) in enumerate(arv_display, 15):
        ws6.cell(row=i, column=1, value=label).font = _LABEL_FONT
        ws6.cell(row=i, column=2, value=value).font = _VALUE_FONT

    ws6.column_dimensions["A"].width = 40
    ws6.column_dimensions["B"].width = 30

    # ── Tab 7: Sources & Notes ───────────────────────────────────────
    ws7 = wb.create_sheet("Sources & Notes")
    ws7.cell(row=1, column=1, value="Sources & Notes").font = _TITLE_FONT

    notes = [
        "Data Source: OpenWeb Ninja Real-Time Zillow Data API",
        "Comparable sales sourced from Zillow's similar-sale-homes endpoint",
        "",
        "Adjustment Methodology:",
        f"  Square Footage: ${ADJ_PER_SQFT:,.0f} per sqft difference",
        f"  Bedrooms: ${ADJ_PER_BEDROOM:,.0f} per bedroom difference",
        f"  Bathrooms: ${ADJ_PER_BATHROOM:,.0f} per bathroom difference",
        f"  Year Built: ${ADJ_PER_YEAR_BUILT:,.0f} per year difference",
        f"  Lot Size: ${ADJ_PER_LOT_SQFT:,.2f} per sqft (capped at ${ADJ_LOT_MAX:,.0f})",
        f"  Garage: ${ADJ_PER_GARAGE:,.0f} per stall difference",
        f"  Market Conditions: {MARKET_CONDITION_PCT_PER_MONTH * 100:.1f}% per month appreciation",
        "",
        "Two-Bucket Weighting:",
        "  Bucket A (Non-Disclosure): 30% weight",
        "  Bucket B (Disclosure/MLS): 70% weight",
        "",
        "Confidence Bands:",
        "  High (<10% spread): ±5% of mid ARV",
        "  Medium (10-20% spread): -10%/+8% of mid ARV",
        "  Low (>20% spread): -15%/+10% of mid ARV",
        "",
        "Conservative bias: low-end ARV is intentionally wider.",
        "A high ARV that doesn't hold up kills your deal.",
        "A conservative ARV that comes in low leaves room for upside.",
        "",
        f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Region: Knoxville / East Tennessee",
    ]
    for i, note in enumerate(notes, 3):
        ws7.cell(row=i, column=1, value=note).font = _LABEL_FONT

    ws7.column_dimensions["A"].width = 70

    # Save
    wb.save(output_path)
    logger.info("Comp report saved to %s", output_path)
    return output_path


# ── Main entry point ──────────────────────────────────────────────────

def run_comp_analysis(address: str, city: str = "", state: str = "",
                      zip_code: str = "", radius: float = DEFAULT_RADIUS_MILES,
                      months: int = DEFAULT_MONTHS_BACK,
                      output_path: str = "") -> dict:
    """Run a full comp analysis for a property and generate the report.

    Returns a dict with ARV results and the output file path. Empty
    state falls back to DEFAULT_PROPERTY_STATE (currently AL).
    """
    if not state:
        from state_resolver import DEFAULT_PROPERTY_STATE
        state = DEFAULT_PROPERTY_STATE
    logger.info("Starting comp analysis for: %s %s %s %s", address, city, state, zip_code)

    # Step 1: Fetch subject property details
    subject = fetch_subject_property(address, city, state, zip_code)
    if not subject:
        logger.error("Could not fetch subject property data")
        return {"error": "Could not fetch subject property data"}

    logger.info("Subject: %s — %s sqft, %dbd/%sba, built %s, Zestimate %s",
                subject.address, f"{subject.sqft:,}" if subject.sqft else "?",
                subject.bedrooms, subject.bathrooms,
                subject.year_built or "?", _fmt_money(subject.zestimate))

    # Step 2: Fetch comparable sales
    comps = fetch_comparable_sales(subject, radius, months)
    if not comps:
        logger.warning("No comparable sales found — try expanding radius or time window")
        return {"error": "No comparable sales found", "subject": subject}

    # Step 3: Calculate ARV
    arv = calculate_arv(subject, comps)
    logger.info("ARV: %s (low) / %s (mid) / %s (high) — %s confidence",
                _fmt_money(arv.arv_low), _fmt_money(arv.arv_mid),
                _fmt_money(arv.arv_high), arv.confidence)

    # Step 4: Generate report
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_address = "".join(c if c.isalnum() or c == "-" else "_" for c in address)[:40]
        output_path = str(config.OUTPUT_DIR / f"comp_report_{safe_address}_{timestamp}.xlsx")

    report_path = generate_comp_report(subject, comps, arv, output_path)

    return {
        "subject": subject,
        "comps": comps[:MAX_COMPS],
        "arv": arv,
        "report_path": report_path,
    }
